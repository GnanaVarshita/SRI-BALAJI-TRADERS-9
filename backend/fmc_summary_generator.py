import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import PyPDF2
import re
import datetime
from pathlib import Path

ACTIVITY_MAPPING = {
    'Crop Show/Field Days': 'FD',
    'Field Days': 'FD',
    'Organized Farmer Meeting/Village Meeting': 'OFM',
    'Organized Farmer Meeting': 'OFM',
    'Harvest Days': 'HD',
    'Harvest Day': 'HD',
    'Large Farmer Meeting': 'LFM',
    'Video Shoot / Farmer testimonial': 'Video Shoot',
    'Video Shoot': 'Video Shoot',
    'Demo Activity': 'Demo Activity',
    'Skill Development Training': 'Skill Development Training',
    'Other Branding Activity': 'Other Branding Activity',
}

def clean_activity_name(raw_activity):
    raw_clean = raw_activity.strip()
    return ACTIVITY_MAPPING.get(raw_clean, raw_clean)

def extract_pdf_data(path):
    try:
        with open(path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            text = ''
            for page in reader.pages:
                text += page.extract_text() + '\n'
        
        text = text.replace('\xa0', ' ').replace('\xad', '-')
        text_before_grand_total = text.split('Grand Total')[0]
        
        po_no_match = re.search(r'PO No\s*(500BB[\d]+)', text_before_grand_total)
        po_no = po_no_match.group(1) if po_no_match else ''
        
        # Extract PO Date
        date_match = re.search(r'PO\s*Date\s*([\d]{1,2}/[A-Za-z]{3}/[\d]{4}|\d{2}/\d{2}/\d{4}|\d{2}-\d{2}-\d{4}|\d{1,2}/[A-Za-z]+/\d{4})', text, re.IGNORECASE)
        po_date = ''
        if date_match:
            raw_date = date_match.group(1).strip()
            try:
                dt_obj = datetime.datetime.strptime(raw_date, '%d/%b/%Y')
                po_date = dt_obj.strftime('%d-%m-%Y')
            except Exception:
                po_date = raw_date
        else:
            m2 = re.search(r'Date\s*([\d]{1,2}[/-][A-Za-z0-9]{2,3}[/-][\d]{4})', text, re.IGNORECASE)
            if m2:
                po_date = m2.group(1).strip()

        activities_list = []
        blocks = text_before_grand_total.split('Activity Type')[1:]
        for block in blocks:
            m = re.search(r'^\s*(.*?)Crop\n(.*?)Product\n(.*?)State\n', block, re.S)
            val_match = re.search(r'Approx value \(Rs\.\)/Amount\s+([\d.]+)', block)
            if m and val_match:
                raw_act = m.group(1).strip().replace('\n', ' ')
                crop_val = m.group(2).strip().replace('\n', ' ')
                prod_val = m.group(3).strip().replace('\n', ' ')
                budget_val = float(val_match.group(1))
                
                activities_list.append({
                    'Activity': clean_activity_name(raw_act),
                    'Crop': crop_val,
                    'Product': prod_val,
                    'Budget': budget_val
                })
        return po_no, po_date, activities_list
    except Exception as e:
        print(f"Error extracting {path}: {e}")
        return "", "", []

from excel_parser import load_any_workbook

def generate_fmc_summary(input_folder_path, output_excel_path, territory, am_name):
    input_folder = Path(input_folder_path)
    if not input_folder.is_dir():
        raise ValueError(f"Input path is not a directory: {input_folder_path}")

    output_path = Path(output_excel_path)

    # 1. Extract data from all PDFs in input folder
    pdf_po_map = {}
    for pdf_file in input_folder.rglob('*.pdf'):
        po_no, po_date, activities = extract_pdf_data(pdf_file)
        if po_no and activities:
            if po_no not in pdf_po_map:
                pdf_po_map[po_no] = {'date': po_date, 'activities': []}
            pdf_po_map[po_no]['activities'].extend(activities)

    # 2. Check if Excel file already exists
    if output_path.exists():
        wb = load_any_workbook(output_path)
        if 'Sheet1' in wb.sheetnames:
            ws = wb['Sheet1']
        else:
            ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

    ws.views.sheetView[0].showGridLines = True

    # Styles
    title_font = Font(name='Calibri', size=14, bold=True, color='00B050')
    header_font = Font(name='Calibri', size=11, bold=True, color='00B050')
    regular_font = Font(name='Calibri', size=11)
    bold_font = Font(name='Calibri', size=11, bold=True)
    
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _center():
        return Alignment(horizontal='center', vertical='center')

    # Ensure Header Rows (Rows 1 to 3) are set up
    if ws.max_row < 3 or ws.cell(3, 1).value != 'S.No':
        # Row 1: Title
        ws.merge_cells('E1:G1')
        ws['E1'] = f"FMC {territory} Budget"
        ws['E1'].font = title_font
        ws['E1'].alignment = _center()

        # Row 2: AM Name
        ws.merge_cells('E2:G2')
        ws['E2'] = f"AM : {am_name}"
        ws['E2'].font = title_font
        ws['E2'].alignment = _center()

        # Row 3: Headers
        headers = ['S.No', 'Date', 'PO Number', 'Product', 'Crop', 'Activities', 'Budget', 'Spent Budget', 'Balance']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    else:
        # Check if existing sheet is missing Date column at Column 2
        if ws.cell(3, 2).value != 'Date':
            ws.insert_cols(2)
            ws.cell(row=3, column=2, value='Date').font = header_font
            ws.cell(row=3, column=2).border = border
            ws.cell(row=3, column=2).alignment = Alignment(horizontal='center', vertical='center')
            
            # Fill date for existing rows from pdf_po_map
            curr_po = ""
            for r in range(4, ws.max_row + 1):
                po_val = ws.cell(row=r, column=3).value
                if po_val and str(po_val).strip() != '':
                    curr_po = str(po_val).strip()
                if curr_po and curr_po in pdf_po_map:
                    ws.cell(row=r, column=2, value=pdf_po_map[curr_po]['date']).alignment = _center()
                    ws.cell(row=r, column=2).font = regular_font
                    ws.cell(row=r, column=2).border = border

        # Update AM Name in case user changed selection
        ws['E2'] = f"AM : {am_name}"

    # Read existing PO Numbers and max S.No from sheet
    existing_po_numbers = set()
    current_max_sno = 0
    
    # Check if last row is a Total row and remove it so we can append below data
    last_row = ws.max_row
    if last_row >= 4:
        val_col6 = ws.cell(row=last_row, column=6).value
        if val_col6 and str(val_col6).strip().lower() == 'total':
            ws.delete_rows(last_row)

    # Read existing data rows (Row 4 to current max_row)
    for r in range(4, ws.max_row + 1):
        sno_val = ws.cell(row=r, column=1).value
        po_val = ws.cell(row=r, column=3).value
        
        if sno_val is not None and str(sno_val).strip() != '':
            try:
                sno_int = int(sno_val)
                if sno_int > current_max_sno:
                    current_max_sno = sno_int
            except ValueError:
                pass
                
        if po_val and str(po_val).strip() != '':
            existing_po_numbers.add(str(po_val).strip())

    # 3. Filter out new POs (avoid duplicates)
    new_po_keys = [po for po in sorted(pdf_po_map.keys()) if po not in existing_po_numbers]

    # 4. Append new POs
    current_row = ws.max_row + 1
    if current_row < 4:
        current_row = 4

    next_sno = current_max_sno + 1

    for po_no in new_po_keys:
        po_entry = pdf_po_map[po_no]
        po_date = po_entry['date']
        acts = po_entry['activities']
        for idx, act in enumerate(acts):
            sno = next_sno if idx == 0 else ""
            po_date_str = po_date if idx == 0 else ""
            po_num_str = po_no if idx == 0 else ""

            ws.cell(row=current_row, column=1, value=sno).border = border
            ws.cell(row=current_row, column=1).alignment = _center()

            ws.cell(row=current_row, column=2, value=po_date_str).border = border
            ws.cell(row=current_row, column=2).alignment = _center()

            ws.cell(row=current_row, column=3, value=po_num_str).border = border
            ws.cell(row=current_row, column=4, value=act['Product']).border = border
            ws.cell(row=current_row, column=5, value=act['Crop']).border = border
            ws.cell(row=current_row, column=6, value=act['Activity']).border = border

            budget_cell = ws.cell(row=current_row, column=7, value=act['Budget'])
            budget_cell.border = border
            budget_cell.number_format = '#,##0'

            ws.cell(row=current_row, column=8, value="").border = border
            ws.cell(row=current_row, column=9, value="").border = border

            for c in range(1, 10):
                ws.cell(row=current_row, column=c).font = regular_font
                if c >= 3 and c <= 6:
                    ws.cell(row=current_row, column=c).alignment = Alignment(vertical='center')
                elif c >= 7:
                    ws.cell(row=current_row, column=c).alignment = Alignment(horizontal='right', vertical='center')

            current_row += 1

        next_sno += 1

    # 5. Add / Re-add Total Row at the bottom
    for c in range(1, 6):
        ws.cell(row=current_row, column=c).border = border

    ws.cell(row=current_row, column=6, value="Total").font = bold_font
    ws.cell(row=current_row, column=6).border = border

    if current_row > 4:
        total_budget_cell = ws.cell(row=current_row, column=7, value=f"=SUM(G4:G{current_row-1})")
    else:
        total_budget_cell = ws.cell(row=current_row, column=7, value=0)

    total_budget_cell.font = bold_font
    total_budget_cell.border = border
    total_budget_cell.number_format = '#,##0'

    ws.cell(row=current_row, column=8).border = border
    ws.cell(row=current_row, column=9).border = border

    # Adjust column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15

    wb.save(output_path)
    wb.close()

    return len(new_po_keys)


