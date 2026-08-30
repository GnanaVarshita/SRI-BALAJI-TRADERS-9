import os
import re
import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import sys
sys.path.append(str(Path(__file__).resolve().parent))
try:
    from excel_parser import load_any_workbook
except ImportError:
    from .excel_parser import load_any_workbook

def clean_str(val):
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in ["none", "null", "nan"]:
        return ""
    return s

def parse_num(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0

def parse_date_to_obj(val):
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ["%d-%m-%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d\\%m\\%Y", "%d/%m/%y", "%d-%m-%y"]:
        try:
            dt = datetime.datetime.strptime(s, fmt)
            return dt.date()
        except ValueError:
            pass
    return None

def format_date_str(date_obj):
    if not date_obj:
        return ""
    if isinstance(date_obj, (datetime.datetime, datetime.date)):
        return date_obj.strftime("%d-%m-%Y")
    d = parse_date_to_obj(date_obj)
    if d:
        return d.strftime("%d-%m-%Y")
    return str(date_obj).strip()

def calculate_receivable_date(invoice_date_val, days_credit=45):
    """
    Calculates the receivable date by adding credit period (default 45 days) to invoice date.
    e.g. 09-04-2026 + 45 days -> 24-05-2026
    """
    d = parse_date_to_obj(invoice_date_val)
    if d:
        rec_d = d + datetime.timedelta(days=days_credit)
        return rec_d.strftime("%d-%m-%Y")
    return ""

def normalize_po(po_str):
    if not po_str:
        return ""
    return re.sub(r'[\s\-_]', '', str(po_str)).upper()

def get_details_styles():
    thin = Side(style='thin', color='A6A6A6')
    black_thin = Side(style='thin', color='000000')
    black_medium = Side(style='medium', color='000000')
    black_double = Side(style='double', color='000000')

    thin_border = Border(left=black_thin, right=black_thin, top=black_thin, bottom=black_thin)
    box_border = Border(left=black_thin, right=black_thin, top=black_thin, bottom=black_thin)
    summary_border = Border(left=black_thin, right=black_thin, top=black_thin, bottom=black_double)

    header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    return {
        'thin_border': thin_border,
        'box_border': box_border,
        'summary_border': summary_border,
        'header_fill': header_fill,
        'font_title': Font(name='Calibri', size=11, bold=True),
        'font_summary': Font(name='Calibri', size=10, bold=True),
        'font_header_green': Font(name='Calibri', size=10, bold=True, color='008000'),
        'font_header_red': Font(name='Calibri', size=10, bold=True, color='C00000'),
        'font_header_brown': Font(name='Calibri', size=10, bold=True, color='993300'),
        'font_header_black': Font(name='Calibri', size=10, bold=True, color='000000'),
        'font_bold': Font(name='Calibri', size=10, bold=True),
        'font_regular': Font(name='Calibri', size=10),
        'align_center': Alignment(horizontal='center', vertical='center'),
        'align_center_wrap': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'align_left': Alignment(horizontal='left', vertical='center'),
        'align_right': Alignment(horizontal='right', vertical='center'),
    }

def create_or_load_details_of_bills_wb(file_path, financial_year="APRIL 2026 to MARCH 2027"):
    """
    Initializes a new Details of Bills workbook if not present or empty, matching SS1 structure.
    """
    target = Path(file_path).resolve()
    styles = get_details_styles()

    if target.exists():
        wb = load_any_workbook(target)
        ws = wb.active
        if ws.max_row >= 4 and any(clean_str(ws.cell(4, c).value) for c in range(1, 10)):
            return wb, ws, False
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            ws = wb["Sheet"]
            ws.title = "Sheet1"
        else:
            ws = wb.create_sheet(title="Sheet1", index=0)

    # Initialize full SS1 structure
    ws.views.sheetView[0].showGridLines = True

    # Column widths based on SS1
    col_widths = {
        'A': 8,   # I.V NO
        'B': 13,  # DATE
        'C': 12,  # AREA
        'D': 16,  # PO NUMBER
        'E': 10,  # BUDGET
        'F': 14,  # PRODUCT
        'G': 14,  # CROP
        'H': 14,  # ACTIVITY
        'I': 14,  # ZDGM
        'J': 16,  # Tbm
        'K': 10,  # NO.OF.Ac
        'L': 14,  # AMOUNT
        'M': 14,  # SERVICE IV
        'N': 15,  # TOTAL IV
        'O': 14,  # Grand Total
        'P': 15,  # RECEIVABLE DATE
        'Q': 14,  # RECEIVED
        'R': 12,  # TDS
        'S': 14,  # Received Date
        'T': 16,  # Receivable Amount
        'U': 16,  # Total Amount
    }
    for col_let, w in col_widths.items():
        ws.column_dimensions[col_let].width = w

    # Row 1 height
    ws.row_dimensions[1].height = 15

    # Row 2: Financial Year title
    ws.row_dimensions[2].height = 22
    ws.merge_cells('H2:K2')
    c_title = ws['H2']
    c_title.value = financial_year
    c_title.font = styles['font_title']
    c_title.alignment = styles['align_center']

    # Row 3: Top Summary Totals Formulas
    ws.row_dimensions[3].height = 20
    ws['L3'] = "=SUM(L5:L5)"
    ws['M3'] = "=SUM(M5:M5)"
    ws['N3'] = "=SUM(N5:N5)"
    ws['O3'] = "=SUM(O5:O5)"
    ws['Q3'] = "=SUM(Q5:Q5)"
    ws['R3'] = "=SUM(R5:R5)"
    ws['T3'] = "=SUM(T5:T5)"
    ws['U3'] = "=SUM(U5:U5)"

    for c_idx in [12, 13, 14, 15, 17, 18, 20, 21]:
        cell = ws.cell(3, c_idx)
        cell.font = styles['font_summary']
        cell.alignment = styles['align_right']
        cell.border = styles['box_border']
        cell.number_format = '#,##0.00'

    # Row 4: Column Headers (with SS1 styling)
    ws.row_dimensions[4].height = 26
    headers_spec = [
        (1, "I.V NO", styles['font_header_green']),
        (2, "DATE", styles['font_header_red']),
        (3, "AREA", styles['font_header_brown']),
        (4, "PO NUMBER", styles['font_header_red']),
        (5, "BUDGET", styles['font_header_red']),
        (6, "PRODUCT", styles['font_header_red']),
        (7, "CROP", styles['font_header_red']),
        (8, "ACTIVITY", styles['font_header_red']),
        (9, "ZDGM", styles['font_header_red']),
        (10, "Tbm", styles['font_header_black']),
        (11, "NO.OF\n.Ac", styles['font_header_red']),
        (12, "AMOUNT", styles['font_header_red']),
        (13, "SERVICE IV", styles['font_header_red']),
        (14, "TOTAL IV", styles['font_header_red']),
        (15, "Grand Total", styles['font_header_red']),
        (16, "RECEIVABL\nE DATE", styles['font_header_red']),
        (17, "RECEIVED", styles['font_header_black']),
        (18, "TDS", styles['font_header_red']),
        (19, "Receive\nd Date", styles['font_header_red']),
        (20, "Receivable\nAmount", styles['font_header_red']),
        (21, "Total Amount", styles['font_header_black']),
    ]

    for c_idx, text, font_style in headers_spec:
        cell = ws.cell(4, c_idx, text)
        cell.font = font_style
        cell.alignment = styles['align_center_wrap']
        cell.border = styles['box_border']

    return wb, ws, True

def extract_invoice_file_data(invoice_file_path):
    """
    Parses a generated Invoice Excel workbook taking primary reference from Sheet 1 (metadata, particulars & totals),
    and grouping activities by TBM & Activity where NO.OF.Activities is the count of activities/days (NOT farmers).
    """
    inv_file = Path(invoice_file_path).resolve()
    if not inv_file.exists():
        return None

    wb = load_any_workbook(inv_file)
    if "Sheet1" not in wb.sheetnames:
        wb.close()
        return None

    ws1 = wb["Sheet1"]
    
    # 1. Extract Invoice Metadata from Sheet 1
    invoice_no = ""
    invoice_date = ""
    po_number = ""
    area = ""
    product = ""
    crop = ""
    zdgm = ""
    grand_total = 0.0
    service_charge_pct = 5.0

    # Scan rows 6 to 30 for metadata in Sheet 1
    for r in range(6, min(30, ws1.max_row + 1)):
        g_val = clean_str(ws1.cell(r, 7).value).lower()
        i_val = clean_str(ws1.cell(r, 9).value)

        if "invoive no" in g_val or "invoice no" in g_val:
            invoice_no = i_val
        elif "invoice date" in g_val:
            invoice_date = format_date_str(i_val)
        elif "po no" in g_val or "po:" in g_val or "new gen po" in g_val or "corteva po" in g_val:
            po_number = i_val
        elif "area" in g_val:
            area = i_val
        elif "product" in g_val:
            product = i_val
        elif "crop" in g_val:
            crop = i_val
        elif "zdgm" in g_val or "amm" in g_val:
            zdgm = i_val

    if not invoice_no:
        invoice_no = inv_file.stem

    short_iv = invoice_no
    m_iv = re.search(r'SBT\d{4}(\d+)', invoice_no, re.I)
    if m_iv:
        short_iv = str(int(m_iv.group(1)))
    else:
        m_num = re.search(r'(\d+)$', invoice_no)
        if m_num:
            short_iv = str(int(m_num.group(1)))

    # Scan for service charge percent if present in Sheet 1
    for r in range(25, min(45, ws1.max_row + 1)):
        e_val = clean_str(ws1.cell(r, 5).value)
        if "%" in e_val:
            try:
                service_charge_pct = float(e_val.replace("%", "").strip())
            except ValueError:
                pass

    # Determine Budget type from PO format
    norm_po = normalize_po(po_number)
    if norm_po.startswith("5"):
        budget_type = "Brand"
    elif norm_po.startswith("48"):
        budget_type = "MA"
    else:
        budget_type = "MKTG"

    # 2. Extract Activities & TBM groups
    # Check if Sheet 2 has TBM-level tables to get exact TBM names and activity counts
    items = []
    if "Sheet2" in wb.sheetnames:
        ws2 = wb["Sheet2"]
        tbm_act_groups = {}  # key: (tbm, activity) -> {'tbm': ..., 'activity': ..., 'qty': count, 'amount': sum}
        max_r = ws2.max_row
        r = 1
        while r <= max_r:
            c1_val = clean_str(ws2.cell(r, 1).value).lower()
            c2_val = clean_str(ws2.cell(r, 2).value).lower()

            if ("s.no" in c1_val or "si no" in c1_val or "sl" in c1_val) and "date" in c2_val:
                hdr_r = r
                data_r = hdr_r + 1
                while data_r <= max_r:
                    d_c1 = clean_str(ws2.cell(data_r, 1).value)
                    d_c15 = clean_str(ws2.cell(data_r, 15).value).lower()

                    if d_c1.lower() == 'total' or d_c15 == 'total' or 'total' in d_c1.lower():
                        break
                    if "activities expenses" in d_c1.lower() or "iv no" in d_c1.lower():
                        break

                    row_vals = [ws2.cell(data_r, c).value for c in range(1, 18)]
                    if not any(row_vals):
                        data_r += 1
                        continue

                    item_tbm = clean_str(ws2.cell(data_r, 4).value)
                    item_act = clean_str(ws2.cell(data_r, 9).value)
                    item_amt = parse_num(ws2.cell(data_r, 16).value)
                    if item_amt == 0.0:
                        calc_amt = sum(parse_num(ws2.cell(data_r, c).value) for c in range(12, 16))
                        if calc_amt > 0:
                            item_amt = calc_amt

                    if item_act or item_tbm or item_amt > 0:
                        grp_key = (item_tbm, item_act)
                        if grp_key not in tbm_act_groups:
                            tbm_act_groups[grp_key] = {
                                'tbm': item_tbm,
                                'activity': item_act,
                                'qty': 0,
                                'amount': 0.0
                            }
                        tbm_act_groups[grp_key]['qty'] += 1  # Count of activity events (NOT farmers!)
                        tbm_act_groups[grp_key]['amount'] += item_amt

                    data_r += 1
                r = data_r
            else:
                r += 1

        for grp in tbm_act_groups.values():
            amt = grp['amount']
            s_iv = round(amt * (1 + service_charge_pct / 100.0), 2)
            t_iv = round(s_iv * 1.18, 2)
            items.append({
                'date': invoice_date,
                'area': area,
                'po_number': po_number,
                'budget': budget_type,
                'product': product,
                'crop': crop,
                'activity': grp['activity'],
                'zdgm': zdgm,
                'tbm': grp['tbm'],
                'qty': grp['qty'], # Exact number of activities
                'amount': amt,
                'service_iv': s_iv,
                'total_iv': t_iv
            })

    # If no Sheet 2 TBM groups found, read directly from Sheet 1 Particulars (Rows 25 to 33)
    if not items:
        for r_part in range(25, 34):
            part_name = clean_str(ws1.cell(r_part, 2).value)
            part_qty = int(parse_num(ws1.cell(r_part, 9).value))
            part_amt = parse_num(ws1.cell(r_part, 10).value)

            if part_qty > 0 or part_amt > 0:
                act_clean = re.sub(r'(?i)\s*activities\s*expenses\s*', '', part_name).strip()
                if not act_clean:
                    act_clean = "Marketing Activities"

                # Check if amount includes service charge (Corteva) or is raw (FMC)
                if norm_po.startswith("48"):
                    s_iv = part_amt
                    raw_amt = round(s_iv / (1 + service_charge_pct / 100.0), 2)
                else:
                    raw_amt = part_amt
                    s_iv = round(raw_amt * (1 + service_charge_pct / 100.0), 2)
                
                t_iv = round(s_iv * 1.18, 2)
                items.append({
                    'date': invoice_date,
                    'area': area,
                    'po_number': po_number,
                    'budget': budget_type,
                    'product': product,
                    'crop': crop,
                    'activity': act_clean,
                    'zdgm': zdgm,
                    'tbm': "",
                    'qty': part_qty, # From Sheet 1 NO.OF.Days /Act
                    'amount': raw_amt,
                    'service_iv': s_iv,
                    'total_iv': t_iv
                })

    # Scan for Grand Total in Sheet1
    for r in range(25, min(50, ws1.max_row + 1)):
        g_val = clean_str(ws1.cell(r, 7).value).lower()
        if "grand total" in g_val:
            raw_gt = ws1.cell(r, 10).value
            if isinstance(raw_gt, (int, float)):
                grand_total = float(raw_gt)
            break

    # If grand_total was not a direct number, calculate from items
    if grand_total == 0.0 and items:
        tot_sub = sum(it.get('service_iv', 0.0) for it in items)
        tot_cgst = round(tot_sub * 0.09, 2)
        tot_sgst = round(tot_sub * 0.09, 2)
        grand_total = round(tot_sub + tot_cgst + tot_sgst)

    wb.close()

    receivable_date = calculate_receivable_date(invoice_date, days_credit=45)

    return {
        'invoice_no': invoice_no,
        'short_iv': short_iv,
        'invoice_date': invoice_date,
        'po_number': po_number,
        'area': area,
        'product': product,
        'crop': crop,
        'zdgm': zdgm,
        'grand_total': grand_total,
        'receivable_date': receivable_date,
        'service_charge_pct': service_charge_pct,
        'items': items
    }

def update_budget_po_summary_cards(cards_excel_path, invoice_records_list):
    """
    Scans the Budget PO Summary Cards Excel (e.g. Nandyala FMC Budget.xlsx)
    and populates I.V NO (Column A) and DATE (Column B) for matching PO cards.
    """
    cards_file = Path(cards_excel_path).resolve()
    if not cards_file.exists():
        return {"success": False, "message": f"Cards file not found: {cards_excel_path}"}

    wb = load_any_workbook(cards_file)
    styles = get_details_styles()

    cards_updated = 0
    updated_pos = set()

    for inv_data in invoice_records_list:
        target_po = inv_data.get('po_number', '').strip()
        norm_target_po = normalize_po(target_po)
        short_iv = inv_data.get('short_iv', '')
        inv_date = inv_data.get('invoice_date', '')

        if not norm_target_po:
            continue

        # Look for PO in all sheets
        for sname in wb.sheetnames:
            ws = wb[sname]
            max_r = ws.max_row

            for r in range(1, max_r + 1):
                # Card PO title in Col A (e.g. A1, A20, etc.)
                c1_val = clean_str(ws.cell(r, 1).value)
                norm_c1 = normalize_po(c1_val)

                if norm_target_po in norm_c1 or norm_c1 == norm_target_po:
                    # Found the PO card header at row r.
                    # Locate table under this card (usually 6-7 rows below, starting with 'I.V NO' in Col A)
                    for tbl_r in range(r + 1, min(r + 15, max_r + 1)):
                        cell_a_hdr = clean_str(ws.cell(tbl_r, 1).value).upper().replace(".", "").replace(" ", "")
                        if "IVNO" in cell_a_hdr:
                            # Table data starts at tbl_r + 1
                            card_data_r = tbl_r + 1
                            
                            # Write IV NO in Col A (1) and DATE in Col B (2)
                            cell_iv = ws.cell(card_data_r, 1, short_iv)
                            cell_iv.font = styles['font_header_green']
                            cell_iv.alignment = styles['align_center']

                            cell_dt = ws.cell(card_data_r, 2, inv_date)
                            cell_dt.font = styles['font_header_green']
                            cell_dt.alignment = styles['align_center']

                            cards_updated += 1
                            updated_pos.add(target_po)
                            break

    wb.save(cards_file)
    wb.close()

    return {
        "success": True,
        "cardsUpdated": cards_updated,
        "updatedPos": list(updated_pos)
    }

def scan_and_append_invoices(
    details_excel_path,
    invoices_folder_path,
    budget_cards_path=None,
    financial_year="APRIL 2026 to MARCH 2027"
):
    """
    Main orchestration function:
    1. Loads or initializes Details of Bills workbook (Sheet1).
    2. Scans invoices folder and extracts bill details taking reference from Sheet 1.
    3. Appends new / missing invoices into Sheet1.
    4. Recalculates top summary formulas in Row 3.
    5. Syncs IV No and Date into Budget PO Summary Cards if provided.
    """
    invoices_folder = Path(invoices_folder_path).resolve()
    if not invoices_folder.exists() or not invoices_folder.is_dir():
        raise FileNotFoundError(f"Invoices folder not found: {invoices_folder_path}")

    # 1. Load or Create Details of Bills workbook
    wb, ws, is_new = create_or_load_details_of_bills_wb(details_excel_path, financial_year)
    styles = get_details_styles()

    # 2. Collect existing invoice numbers in Sheet 1 to prevent duplicates
    existing_ivs = set()
    for r in range(5, ws.max_row + 1):
        iv_val = clean_str(ws.cell(r, 1).value)
        if iv_val:
            existing_ivs.add(iv_val.upper())
            m_num = re.search(r'(\d+)$', iv_val)
            if m_num:
                existing_ivs.add(str(int(m_num.group(1))))

    # Helper function for natural numerical sorting of invoice filenames
    def get_invoice_sort_key(p):
        m = re.search(r'(\d+)', p.stem)
        return int(m.group(1)) if m else 999999

    # 3. Scan all invoice excel files in folder with natural numerical sort
    invoice_files = sorted(list(invoices_folder.glob("*.xlsx")), key=get_invoice_sort_key)
    if not invoice_files:
        raise ValueError(f"No invoice Excel (.xlsx) files found in: {invoices_folder_path}")

    parsed_invoices = []
    appended_invoices = []
    skipped_invoices = []
    total_new_rows = 0

    current_r = max(5, ws.max_row + 1)
    # If Sheet1 was already populated, find actual first truly empty row
    for check_r in range(5, ws.max_row + 2):
        if not any(ws.cell(check_r, c).value for c in range(1, 15)):
            current_r = check_r
            break

    for inv_path in invoice_files:
        inv_data = extract_invoice_file_data(inv_path)
        if not inv_data:
            continue

        parsed_invoices.append(inv_data)
        short_iv = inv_data['short_iv']
        full_iv = inv_data['invoice_no']

        # Check if already present in Details of Bills strictly by invoice number
        if short_iv.upper() in existing_ivs or full_iv.upper() in existing_ivs:
            skipped_invoices.append(short_iv)
            continue

        # Append invoice activity rows
        items = inv_data['items']
        num_items = len(items)
        iv_start_r = current_r

        for idx, item in enumerate(items, start=1):
            ws.row_dimensions[current_r].height = 18

            # Col A (1): I.V NO (Green bold) on first row of invoice group
            if idx == 1:
                ws.cell(current_r, 1, short_iv).font = styles['font_header_green']
                ws.cell(current_r, 1).alignment = styles['align_center']
            else:
                ws.cell(current_r, 1, "").alignment = styles['align_center']

            # Col B (2): DATE
            ws.cell(current_r, 2, item.get('date', '')).alignment = styles['align_center']
            
            # Col C (3): AREA
            ws.cell(current_r, 3, item.get('area', '')).alignment = styles['align_left']

            # Col D (4): PO NUMBER
            ws.cell(current_r, 4, item.get('po_number', '')).alignment = styles['align_center']

            # Col E (5): BUDGET
            ws.cell(current_r, 5, item.get('budget', 'Brand')).alignment = styles['align_center']

            # Col F (6): PRODUCT
            ws.cell(current_r, 6, item.get('product', '')).alignment = styles['align_left']

            # Col G (7): CROP
            ws.cell(current_r, 7, item.get('crop', '')).alignment = styles['align_left']

            # Col H (8): ACTIVITY
            ws.cell(current_r, 8, item.get('activity', '')).alignment = styles['align_left']

            # Col I (9): ZDGM
            ws.cell(current_r, 9, item.get('zdgm', '')).alignment = styles['align_left']

            # Col J (10): Tbm
            ws.cell(current_r, 10, item.get('tbm', '')).alignment = styles['align_left']

            # Col K (11): NO.OF.Ac (Number of Activities / Days - NOT farmers)
            c_qty = ws.cell(current_r, 11, item.get('qty', 1))
            c_qty.alignment = styles['align_center']
            c_qty.number_format = '#,##0'

            # Col L (12): AMOUNT
            c_amt = ws.cell(current_r, 12, item.get('amount', 0.0))
            c_amt.alignment = styles['align_right']
            c_amt.number_format = '#,##0'

            # Col M (13): SERVICE IV
            c_siv = ws.cell(current_r, 13, item.get('service_iv', 0.0))
            c_siv.alignment = styles['align_right']
            c_siv.number_format = '#,##0.00'

            # Col N (14): TOTAL IV
            c_tiv = ws.cell(current_r, 14, item.get('total_iv', 0.0))
            c_tiv.alignment = styles['align_right']
            c_tiv.number_format = '#,##0.00'

            # Col O (15): Grand Total on last row of invoice group
            if idx == num_items:
                c_gt = ws.cell(current_r, 15, inv_data.get('grand_total', 0.0))
                c_gt.font = styles['font_bold']
                c_gt.alignment = styles['align_right']
                c_gt.number_format = '#,##0'

            # Col P (16): RECEIVABLE DATE
            ws.cell(current_r, 16, inv_data.get('receivable_date', '')).alignment = styles['align_center']

            # Cols Q to U (17-21): Tracking fields (empty initially with borders)
            for c in range(17, 22):
                ws.cell(current_r, c, "")

            # Apply borders and regular fonts
            for c in range(1, 22):
                cell = ws.cell(current_r, c)
                if c not in [1, 15]:
                    cell.font = styles['font_regular']
                cell.border = styles['box_border']

            current_r += 1
            total_new_rows += 1

        appended_invoices.append(short_iv)
        existing_ivs.add(short_iv.upper())

    # 4. Refresh Top Summary Totals in Row 3
    final_max_r = max(5, current_r - 1)
    ws['L3'] = f"=SUM(L5:L{final_max_r})"
    ws['M3'] = f"=SUM(M5:M{final_max_r})"
    ws['N3'] = f"=SUM(N5:N{final_max_r})"
    ws['O3'] = f"=SUM(O5:O{final_max_r})"
    ws['Q3'] = f"=SUM(Q5:Q{final_max_r})"
    ws['R3'] = f"=SUM(R5:R{final_max_r})"
    ws['T3'] = f"=SUM(T5:T{final_max_r})"
    ws['U3'] = f"=SUM(U5:U{final_max_r})"

    # Save Details of Bills workbook
    out_details_path = Path(details_excel_path).resolve()
    out_details_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_details_path)
    wb.close()

    # 5. Sync to Budget PO Summary Cards if path provided
    cards_sync_res = None
    if budget_cards_path and str(budget_cards_path).strip():
        cards_sync_res = update_budget_po_summary_cards(budget_cards_path, parsed_invoices)

    return {
        "success": True,
        "isNew": is_new,
        "detailsExcelPath": str(out_details_path),
        "totalInvoicesFound": len(invoice_files),
        "appendedInvoices": appended_invoices,
        "skippedInvoices": skipped_invoices,
        "totalRowsAdded": total_new_rows,
        "totalRowsInSheet": final_max_r - 4,
        "cardsSync": cards_sync_res,
        "message": f"Successfully processed {len(invoice_files)} invoices. Appended {len(appended_invoices)} new invoices ({total_new_rows} rows) to Details of Bills."
    }
