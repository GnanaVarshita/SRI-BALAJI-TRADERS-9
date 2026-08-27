import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from excel_parser import validate_budget_sheet, load_any_workbook

def generate_quotations(filepath, company, contact, designation, territory, date_str):
    """
    Generates product quotation sheets and appends them to the excel workbook.
    """
    # 1. First run validation to fetch rows and check products
    validation_res = validate_budget_sheet(filepath)
    if not validation_res["rows"]:
        raise ValueError("No data rows found in the uploaded sheet to process.")
        
    rows = validation_res["rows"]
    
    # Get unique products list
    products = sorted(list(set(r["product"] for r in rows if r["product"])))
    
    # Load original workbook for modification
    wb = load_any_workbook(filepath)
    
    # Styles definition
    gold_header_fill = PatternFill(start_color="F5E6CC", end_color="F5E6CC", fill_type="solid") # Soft Gold
    light_gold_fill = PatternFill(start_color="FAF6EE", end_color="FAF6EE", fill_type="solid") # Warm accent
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    total_border = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )
    
    green_font = Font(name="Segoe UI", size=10, bold=True, italic=True, color="008000") # Bold Italic Green
    bold_gold_title_font = Font(name="Segoe UI", size=12, bold=True, color="AA7C11") # Gold title
    
    # Remove any previously generated sheets that match product names to avoid duplicates
    for p in products:
        if p in wb.sheetnames:
            del wb[p]

    # For each product, create a styled sheet
    for product in products:
        # Filter rows for this product, skipping rows with blank or zero activity quantities
        prod_rows = [r for r in rows if r["product"] == product and r["qty"] is not None and r["qty"] > 0]
        if not prod_rows:
            continue
            
        ws = wb.create_sheet(title=product)
        ws.views.sheetView[0].showGridLines = True
        
        # Write Title block (Row 9)
        ws["B9"] = f"Quotation for {territory} Kharif Budget"
        ws["B9"].font = bold_gold_title_font
        
        ws["F9"] = f"Date : {date_str}"
        ws["F9"].font = green_font
        ws["F9"].alignment = Alignment(horizontal="right")
        
        # Write Address Block (Row 11-15)
        ws["B11"] = "TO :"
        ws["B11"].font = Font(name="Segoe UI", size=10, bold=True)
        
        ws["B12"] = company
        ws["B12"].font = Font(name="Segoe UI", size=10, bold=True, color="008000")
        
        ws["B13"] = contact
        ws["B13"].font = Font(name="Segoe UI", size=10, color="008000")
        
        ws["B14"] = designation
        ws["B14"].font = Font(name="Segoe UI", size=10, color="008000")
        
        ws["B15"] = territory
        ws["B15"].font = Font(name="Segoe UI", size=10, color="008000")
        
        # Subject line (Row 17)
        ws["B17"] = f"Sub: Quotation for {product} Product Budget Activities {territory} Territory"
        ws["B17"].font = Font(name="Segoe UI", size=10, bold=True, italic=True)
        
        # Write Table Headers (Row 18)
        headers = ["Type", "PRODUCT", "crop", "Type of Activity", "Event Cost", "Marketing Budget for DG Activities", "Total Budget Allocation"]
        for c_idx, h_text in enumerate(headers, start=2): # Start on Column B
            cell = ws.cell(row=18, column=c_idx, value=h_text)
            cell.font = Font(name="Segoe UI", size=10, bold=True)
            cell.fill = gold_header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        ws.row_dimensions[18].height = 28
        
        # Write Data rows (Row 19 onwards)
        current_row = 19
        for row_data in prod_rows:
            ws.row_dimensions[current_row].height = 20
            
            # Map values to columns B to H
            ws.cell(row=current_row, column=2, value=row_data["type"]).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=3, value=row_data["product"]).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=4, value=row_data["crop"]).alignment = Alignment(horizontal="center")
            
            ws.cell(row=current_row, column=5, value=row_data["activity"]).alignment = Alignment(horizontal="left", wrap_text=True)
            
            ws.cell(row=current_row, column=6, value=row_data["event_cost"]).alignment = Alignment(horizontal="right")
            ws.cell(row=current_row, column=6).number_format = '0.00'
            
            ws.cell(row=current_row, column=7, value=row_data["qty"]).alignment = Alignment(horizontal="center")
            
            # Formula for budget allocation: =F{row}*G{row}
            ws.cell(row=current_row, column=8, value=f"=F{current_row}*G{current_row}").alignment = Alignment(horizontal="right")
            ws.cell(row=current_row, column=8).number_format = '0.00'
            
            # Apply thin borders and default fonts
            for col in range(2, 9):
                cell = ws.cell(row=current_row, column=col)
                cell.font = Font(name="Segoe UI", size=10)
                cell.border = thin_border
                
            current_row += 1
            
        # Write Total row
        ws.row_dimensions[current_row].height = 22
        ws.cell(row=current_row, column=5, value="Total").font = Font(name="Segoe UI", size=10, bold=True)
        ws.cell(row=current_row, column=5).alignment = Alignment(horizontal="right")
        
        # Formula for Total Quantity: =SUM(G19:G{current_row-1})
        tot_qty_cell = ws.cell(row=current_row, column=7, value=f"=SUM(G19:G{current_row-1})")
        tot_qty_cell.font = Font(name="Segoe UI", size=10, bold=True, color="FF0000")
        tot_qty_cell.alignment = Alignment(horizontal="center")
        tot_qty_cell.border = total_border
        
        # Formula for Total Allocation: =SUM(H19:H{current_row-1})
        tot_alloc_cell = ws.cell(row=current_row, column=8, value=f"=SUM(H19:H{current_row-1})")
        tot_alloc_cell.font = Font(name="Segoe UI", size=10, bold=True, color="FF0000")
        tot_alloc_cell.alignment = Alignment(horizontal="right")
        tot_alloc_cell.border = total_border
        tot_alloc_cell.number_format = '0.00'
        
        # Apply borders to other columns in total row
        for col in [2, 3, 4, 6]:
            ws.cell(row=current_row, column=col).border = Border(top=Side(style='thin', color='CCCCCC'))
            
        # Add Notes / Terms & Conditions block (Row total_row + 3)
        notes_start = current_row + 3
        ws.cell(row=notes_start, column=2, value="*Rate of the Activity is including Vendor service Charges 5%").font = green_font
        ws.cell(row=notes_start+1, column=2, value="*Quotation Validity is for 10days").font = green_font
        ws.cell(row=notes_start+2, column=2, value="*Vendor Service tax will be extra").font = green_font
        
        # Signature block (Columns F & G, Row total_row + 5)
        sig_start = current_row + 5
        ws.cell(row=sig_start, column=6, value="Thanks & Regards").font = green_font
        ws.cell(row=sig_start+1, column=6, value="SriBalajiTraders").font = green_font
        ws.cell(row=sig_start+2, column=6, value="Radhadevi Kamisetty").font = green_font
        ws.cell(row=sig_start+3, column=6, value="Proprietor").font = green_font
        
        # Align signature right
        for row_offset in range(4):
            ws.cell(row=sig_start+row_offset, column=6).alignment = Alignment(horizontal="center")
            
        # Merge signature columns for neat display (F & G)
        ws.merge_cells(start_row=sig_start, start_column=6, end_row=sig_start, end_column=7)
        ws.merge_cells(start_row=sig_start+1, start_column=6, end_row=sig_start+1, end_column=7)
        ws.merge_cells(start_row=sig_start+2, start_column=6, end_row=sig_start+2, end_column=7)
        ws.merge_cells(start_row=sig_start+3, start_column=6, end_row=sig_start+3, end_column=7)

        # Set specific Column Widths
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 8   # Type
        ws.column_dimensions['C'].width = 15  # PRODUCT
        ws.column_dimensions['D'].width = 10  # crop
        ws.column_dimensions['E'].width = 45  # Type of Activity
        ws.column_dimensions['F'].width = 12  # Event Cost
        ws.column_dimensions['G'].width = 15  # DG Qty
        ws.column_dimensions['H'].width = 18  # Total Allocation

    # Save changes
    wb.save(filepath)
    wb.close()
