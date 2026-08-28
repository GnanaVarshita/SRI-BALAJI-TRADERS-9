import openpyxl
import re
import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import sys
sys.path.append(str(Path(__file__).resolve().parent))
try:
    from excel_parser import load_any_workbook
except ImportError:
    from .excel_parser import load_any_workbook

def generate_po_summary(input_path, output_path, po_number, date_str, contact, territory):
    """
    Parses product worksheets from the input Quotation workbook
    and generates a brand-new PO Summary workbook with tracking sheets.
    """
    wb_input = load_any_workbook(input_path)

    wb_output = openpyxl.Workbook()
    default_sheet = wb_output.active
    wb_output.remove(default_sheet)

    # --- Styling helpers ---
    def _side(color='000000'):
        return Side(style='thin', color=color)

    def _medium_side(color='000000'):
        return Side(style='medium', color=color)

    def _border(color='000000'):
        s = _medium_side(color)
        return Border(left=s, right=s, top=s, bottom=s)

    thin_border = _border('000000')  # medium-weight black borders for the table

    total_border = Border(
        top=_medium_side('000000'),
        bottom=Side(style='double', color='000000'),
        left=_medium_side('000000'),
        right=_medium_side('000000')
    )

    header_fill  = PatternFill(start_color='C2D69B', end_color='C2D69B', fill_type='solid')
    summary_fill = PatternFill(start_color='EBF1DE', end_color='EBF1DE', fill_type='solid')

    green_font    = Font(name='Calibri', size=10, bold=True,  color='008000')
    red_bold_font = Font(name='Calibri', size=10, bold=True,  color='FF0000')
    blue_bold     = Font(name='Calibri', size=10, bold=True,  color='0000FF')
    bold_font     = Font(name='Calibri', size=10, bold=True)
    regular_font  = Font(name='Calibri', size=10)
    title_font    = Font(name='Calibri', size=11, bold=True)

    def _center(wrap=False):
        return Alignment(horizontal='center', vertical='center', wrap_text=wrap)

    def _right():
        return Alignment(horizontal='right', vertical='center')

    def _left():
        return Alignment(horizontal='left', vertical='center')

    # --- Loop sheets ---
    first_sheet_name = wb_input.sheetnames[0]
    for name in wb_input.sheetnames:
        if (name == first_sheet_name
                or name.strip().lower() in ['sheet1', 'processed_emails']
                or name.startswith('Sheet')):
            continue

        ws_input = wb_input[name]

        # 1. Parse activities
        activities = []
        header_r = 18
        val_e18 = ws_input.cell(row=18, column=5).value
        if val_e18 is None or 'activity' not in str(val_e18).lower():
            for r in range(1, 26):
                val_col5 = ws_input.cell(row=r, column=5).value
                if val_col5 and 'activity' in str(val_col5).lower():
                    header_r = r
                    break

        crop_val = 'All Crops'
        for r_idx in range(header_r + 1, 100):
            val_e = ws_input.cell(row=r_idx, column=5).value
            if val_e is None or str(val_e).strip().lower() == 'total' or str(val_e).strip() == '':
                break
            row_crop = ws_input.cell(row=r_idx, column=4).value
            if row_crop:
                crop_val = str(row_crop).strip()
            rate_val   = float(ws_input.cell(row=r_idx, column=6).value or 0.0)
            qty_val    = float(ws_input.cell(row=r_idx, column=7).value or 0.0)
            activities.append({
                'activity': str(val_e).strip(),
                'rate':     rate_val,
                'qty':      qty_val,
                'budget':   rate_val * qty_val
            })

        if not activities:
            continue

        ws = wb_output.create_sheet(title=name)
        ws.views.sheetView[0].showGridLines = True  # Show default gridlines

        suffix = 'MA'
        m = re.search(r'\(([^)]+)\)', name)
        if m:
            suffix = m.group(1).strip()

        n_act         = len(activities)
        ACT_START_COL = 12
        AMOUNT_COL    = ACT_START_COL + n_act
        AMOUNT_LET    = get_column_letter(AMOUNT_COL)
        T_COL, U_COL, V_COL, W_COL, X_COL, Y_COL = 20, 21, 22, 23, 24, 25

        # 2. Row heights
        for r in range(1, 32):
            ws.row_dimensions[r].height = 16
        ws.row_dimensions[6].height  = 18
        ws.row_dimensions[7].height  = 18
        ws.row_dimensions[10].height = 28
        ws.row_dimensions[11].height = 1
        ws.row_dimensions[29].height = 22

        # 3. Column widths
        for col, width in [('A',8),('B',11),('C',10),('D',13),('E',8),
                           ('F',13),('G',9),('H',14),('I',10),('J',10),('K',10),
                           ('T',22),('U',13),('V',12),('W',12),('X',12),('Y',12)]:
            ws.column_dimensions[col].width = width
        for i in range(n_act):
            ws.column_dimensions[get_column_letter(ACT_START_COL + i)].width = 11
        ws.column_dimensions[AMOUNT_LET].width = 11

        # 4. PO Details
        ws.merge_cells('A6:B6')
        ws['A6'].value = po_number or ''
        ws['A6'].font  = blue_bold
        ws['A6'].alignment = _center()

        ws.merge_cells('A7:C7')
        ws['A7'].value = suffix
        ws['A7'].font  = green_font
        ws['A7'].alignment = _center()

        ws['E6'].value = 'PRODUCT'
        ws['E6'].font  = bold_font
        ws['E6'].alignment = _right()

        ws.merge_cells('F6:H6')
        ws['F6'].value = name
        ws['F6'].font  = green_font
        ws['F6'].alignment = _center()

        ws['I6'].value = 'CROP'
        ws['I6'].font  = bold_font
        ws['I6'].alignment = _right()

        ws['J6'].value = crop_val
        ws['J6'].font  = green_font
        ws['J6'].alignment = _left()

        ws['L6'].value = 'DATE'
        ws['L6'].font  = bold_font
        ws['L6'].alignment = _right()

        ws.merge_cells('M6:N6')
        ws['M6'].value = date_str or datetime.date.today().strftime('%d-%m-%Y')
        ws['M6'].font  = green_font
        ws['M6'].alignment = _center()

        ws.merge_cells('D7:F9')
        ws['D7'].value = f'{contact} - {territory}'
        ws['D7'].font  = title_font
        ws['D7'].alignment = _center(wrap=True)

        # 5. Rates grid T3:U
        for i, act in enumerate(activities):
            r = 3 + i
            ws.cell(r, T_COL).value = act['activity']
            ws.cell(r, T_COL).font  = regular_font
            ws.cell(r, U_COL).value = act['budget']
            ws.cell(r, U_COL).font  = regular_font
            ws.cell(r, U_COL).number_format = '#,##0'

        # 6. GST totals row (sits AT row 10, above the merged header)
        last_rate_row = 3 + n_act - 1
        ws.cell(10, T_COL).value     = 'TOTAL'
        ws.cell(10, T_COL).font      = bold_font
        ws.cell(10, T_COL).alignment = _right()
        ws.cell(10, T_COL).border    = total_border

        ws.cell(10, U_COL).value         = f'=SUM(U3:U{last_rate_row})'
        ws.cell(10, U_COL).font          = red_bold_font
        ws.cell(10, U_COL).number_format = '#,##0'
        ws.cell(10, U_COL).border        = total_border

        ws.merge_cells('V10:W10')
        ws.cell(10, V_COL).value     = 'With GST'
        ws.cell(10, V_COL).font      = green_font
        ws.cell(10, V_COL).alignment = _center()

        ws.cell(10, X_COL).value         = '=U10*18%+U10'
        ws.cell(10, X_COL).font          = green_font
        ws.cell(10, X_COL).number_format = '#,##0'
        ws.cell(10, X_COL).border        = total_border

        # 7. Table header rows 10 & 11 (merged per column)
        FIXED_HEADERS = [
            'I.V NO', 'DATE', 'AREA', 'PO NUMBER', 'BUDGET TYPE',
            'PRODUCT', 'CROP', 'ACTIVITY', 'ZDGM', 'TBM', 'NO.OF.Activities'
        ]

        def write_header(col, label):
            cl = get_column_letter(col)
            ws.merge_cells(f'{cl}10:{cl}11')
            c = ws.cell(10, col)
            c.value     = label
            c.font      = bold_font
            c.fill      = header_fill
            c.alignment = _center(wrap=True)
            c.border    = thin_border

        for c_idx, hdr in enumerate(FIXED_HEADERS, start=1):
            write_header(c_idx, hdr)
        for i in range(n_act):
            write_header(ACT_START_COL + i, f'=T{3+i}')
        write_header(AMOUNT_COL, 'Amount')

        # 8. Data rows 12-28
        for r in range(12, 29):
            for c in range(1, AMOUNT_COL + 1):
                cell = ws.cell(r, c)
                cell.border    = thin_border
                cell.alignment = _center()
                cell.font      = regular_font
                if ACT_START_COL <= c < AMOUNT_COL:
                    cell.number_format = '#,##0'

            parts = [get_column_letter(ACT_START_COL + i) + str(r) for i in range(n_act)]
            amt = ws.cell(r, AMOUNT_COL)
            amt.value         = '=' + '+'.join(parts) if parts else 0
            amt.number_format = '#,##0'
            amt.font          = regular_font
            amt.border        = thin_border
            amt.alignment     = _center()

        # 9. Totals row 29
        for c in range(1, AMOUNT_COL + 1):
            ws.cell(29, c).border    = thin_border
            ws.cell(29, c).alignment = _center()

        for i in range(n_act):
            col    = ACT_START_COL + i
            cl     = get_column_letter(col)
            cell   = ws.cell(29, col)
            cell.value         = f'=SUM({cl}12:{cl}28)'
            cell.font          = red_bold_font
            cell.number_format = '#,##0'
            cell.border        = total_border

        amt_tot = ws.cell(29, AMOUNT_COL)
        amt_tot.value         = f'=SUM({AMOUNT_LET}12:{AMOUNT_LET}28)'
        amt_tot.font          = red_bold_font
        amt_tot.number_format = '#,##0'
        amt_tot.border        = total_border

        # 10. Summary block T15:Y
        summary_hdrs = ['Activities','BUDGET','SPENT','SV Charges','TOTAL IV','BALANCE']
        for j, sh in enumerate(summary_hdrs):
            c = ws.cell(15, T_COL + j)
            c.value     = sh
            c.font      = green_font
            c.fill      = summary_fill
            c.alignment = _center()
            c.border    = thin_border

        for i in range(n_act):
            r      = 16 + i
            cl     = get_column_letter(ACT_START_COL + i)
            rate_r = 3 + i

            ws.cell(r, T_COL).value     = f'=T{rate_r}'
            ws.cell(r, T_COL).font      = regular_font
            ws.cell(r, T_COL).border    = thin_border

            ws.cell(r, U_COL).value         = f'=U{rate_r}'
            ws.cell(r, U_COL).font          = regular_font
            ws.cell(r, U_COL).number_format = '#,##0'
            ws.cell(r, U_COL).border        = thin_border

            ws.cell(r, V_COL).value         = f'={cl}29'
            ws.cell(r, V_COL).font          = regular_font
            ws.cell(r, V_COL).number_format = '#,##0'
            ws.cell(r, V_COL).border        = thin_border

            ws.cell(r, W_COL).value         = f'=V{r}*5%'
            ws.cell(r, W_COL).font          = regular_font
            ws.cell(r, W_COL).number_format = '#,##0.00'
            ws.cell(r, W_COL).border        = thin_border

            ws.cell(r, X_COL).value         = f'=V{r}+W{r}'
            ws.cell(r, X_COL).font          = regular_font
            ws.cell(r, X_COL).number_format = '#,##0.00'
            ws.cell(r, X_COL).border        = thin_border

            ws.cell(r, Y_COL).value         = f'=U{r}-X{r}'
            ws.cell(r, Y_COL).font          = regular_font
            ws.cell(r, Y_COL).number_format = '#,##0.00'
            ws.cell(r, Y_COL).border        = thin_border

        tot_row = 16 + n_act
        ws.cell(tot_row, T_COL).value     = 'TOTAL'
        ws.cell(tot_row, T_COL).font      = bold_font
        ws.cell(tot_row, T_COL).border    = total_border

        for col in [U_COL, V_COL, W_COL, X_COL, Y_COL]:
            cl   = get_column_letter(col)
            cell = ws.cell(tot_row, col)
            cell.value         = f'=SUM({cl}16:{cl}{tot_row-1})'
            cell.font          = red_bold_font
            cell.number_format = '#,##0'
            cell.border        = total_border

    wb_output.save(output_path)
    wb_output.close()
    wb_input.close()
