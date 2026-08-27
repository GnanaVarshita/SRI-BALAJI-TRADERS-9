import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import re

def parse_pos_from_step1_sheet(ws):
    pos = []
    current_po = None
    
    for r in range(4, ws.max_row + 1):
        val_c6 = ws.cell(r, 6).value
        if val_c6 and str(val_c6).strip().lower() == 'total':
            break
            
        date_val = ws.cell(r, 2).value
        po_num = ws.cell(r, 3).value
        prod = ws.cell(r, 4).value
        crop = ws.cell(r, 5).value
        act = ws.cell(r, 6).value
        budget = ws.cell(r, 7).value
        
        if po_num and str(po_num).strip() != '':
            current_po = {
                'po_number': str(po_num).strip(),
                'po_date': str(date_val).strip() if date_val else '',
                'products': [str(prod).strip()] if prod and str(prod).strip() else [],
                'crops': [str(crop).strip()] if crop and str(crop).strip() else [],
                'activities': []
            }
            pos.append(current_po)
        elif current_po:
            if prod and str(prod).strip() and str(prod).strip() not in current_po['products']:
                current_po['products'].append(str(prod).strip())
            if crop and str(crop).strip() and str(crop).strip() not in current_po['crops']:
                current_po['crops'].append(str(crop).strip())
            
        if current_po and act:
            current_po['activities'].append({
                'activity': str(act).strip(),
                'budget': budget or 0
            })
            
    for p in pos:
        p['product'] = ' / '.join(p['products']) if p.get('products') else ''
        p['crop'] = ' / '.join(p['crops']) if p.get('crops') else ''
        
    return pos

def write_po_summary_block(ws, start_r, po_data, territory='Nandyal', am_name='Madhavareddy', date_str='01-08-2026'):
    # Styles matching media_1787816358686.png exactly
    red_bold = Font(name='Calibri', size=11, bold=True, color='FF0000')
    green_bold = Font(name='Calibri', size=11, bold=True, color='008000')
    blue_bold = Font(name='Calibri', size=11, bold=True, color='002060')
    green_header_font = Font(name='Calibri', size=10, bold=True, color='008000')
    regular_font = Font(name='Calibri', size=10)
    
    light_green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    def _center():
        return Alignment(horizontal='center', vertical='center')

    po_no = po_data['po_number']
    prefix = po_no[3:5] if len(po_no) >= 5 else 'BB'
    product = po_data['product']
    crops = po_data.get('crops', [po_data.get('crop', '')])
    activities = po_data['activities']
    total_budget = sum(a['budget'] for a in activities)

    # --- ROW 1 of Block (start_r) ---
    # A: PO Number
    ws.cell(start_r, 1, po_no).font = green_bold

    # D: PRODUCT, G: Product Value
    ws.cell(start_r, 4, "PRODUCT").font = red_bold
    ws.cell(start_r, 7, product).font = green_bold

    # H: CROP, I: Crop 1 Value
    ws.cell(start_r, 8, "CROP").font = red_bold
    if len(crops) > 0 and crops[0]:
        ws.cell(start_r, 9, crops[0]).font = green_bold

    # J: DATE, L: Date Value
    ws.cell(start_r, 10, "DATE").font = red_bold
    ws.cell(start_r, 12, date_str).font = green_bold

    # List activities in Cols O & P starting at start_r
    for idx, act_item in enumerate(activities):
        r_pos = start_r + idx
        ws.cell(r_pos, 15, act_item['activity']).font = green_bold
        b_act = ws.cell(r_pos, 16, act_item['budget'])
        b_act.font = green_bold
        b_act.number_format = '#,##0'

    # --- ROW 2 of Block (start_r + 1) ---
    # D: MA or BB
    ws.cell(start_r + 1, 4, prefix).font = green_bold

    # I: Crop 2 Value (if present)
    if len(crops) > 1 and crops[1]:
        crop2_str = ' / '.join(crops[1:])
        ws.cell(start_r + 1, 9, crop2_str).font = green_bold

    # --- ROW 3 of Block (start_r + 2) ---
    # G: Territory Name (e.g. Nandyal / Nellore)
    c_terr = ws.cell(start_r + 2, 7, territory.title())
    c_terr.font = green_bold
    c_terr.alignment = _center()

    # --- ROW 4 of Block (start_r + 3) ---
    # D: AM Name
    ws.cell(start_r + 3, 4, am_name).font = blue_bold

    # O: Total, P: Total Budget, Q: WITH GST, S: Total with GST
    ws.cell(start_r + 3, 15, "Total").font = green_bold

    c_tot = ws.cell(start_r + 3, 16, total_budget)
    c_tot.font = green_bold
    c_tot.number_format = '#,##0'

    ws.cell(start_r + 3, 17, "WITH GST").font = green_bold

    c_gst = ws.cell(start_r + 3, 19, f"=P{start_r + 3}*1.18")
    c_gst.font = green_bold
    c_gst.number_format = '#,##0'

    # --- ROWS 5 & 6 of Block (start_r + 4 & start_r + 5): Headers ---
    # D6: PO
    ws.cell(start_r + 5, 4, "PO").font = red_bold

    # Main Table Headers in Row 7 (start_r + 6)
    main_headers = {
        1: ("I.V NO", green_bold),
        2: ("DATE", green_bold),
        3: ("AREA", green_bold),
        4: ("NUMBER", red_bold),
        5: ("BUDGET", red_bold),
        6: ("Product", red_bold),
        7: ("Crop", red_bold),
        8: ("ACTIVITY", red_bold),
        9: ("RBM", red_bold),
        10: ("MIE", red_bold)
    }
    for col_i, (h_txt, h_font) in main_headers.items():
        cell = ws.cell(start_r + 6, col_i, h_txt)
        cell.font = h_font
        cell.border = border
        cell.alignment = _center()

    # Light Green Header Box for NO.OF.Activities (Cols K to N, Rows start_r + 5 & start_r + 6)
    j5 = ws.cell(start_r + 5, 11, "NO.OF.Ac")
    j5.font = red_bold
    j5.fill = light_green_fill
    j5.border = border
    j5.alignment = _center()

    j6 = ws.cell(start_r + 6, 11, "tivities")
    j6.font = red_bold
    j6.fill = light_green_fill
    j6.border = border
    j6.alignment = _center()

    k_act0 = activities[0]['activity'] if len(activities) > 0 else "FD"
    k5 = ws.cell(start_r + 5, 12, k_act0)
    k5.font = red_bold
    k5.fill = light_green_fill
    k5.border = border
    k5.alignment = _center()

    ws.cell(start_r + 6, 12).fill = light_green_fill
    ws.cell(start_r + 6, 12).border = border

    k_act1 = activities[1]['activity'] if len(activities) > 1 else 0
    l5 = ws.cell(start_r + 5, 13, k_act1)
    l5.font = red_bold
    l5.fill = light_green_fill
    l5.border = border
    l5.alignment = _center()

    ws.cell(start_r + 6, 13, 0).fill = light_green_fill
    ws.cell(start_r + 6, 13).border = border
    ws.cell(start_r + 6, 13).alignment = _center()

    m5 = ws.cell(start_r + 5, 14, "Amount")
    m5.font = red_bold
    m5.fill = light_green_fill
    m5.border = border
    m5.alignment = _center()

    ws.cell(start_r + 6, 14).fill = light_green_fill
    ws.cell(start_r + 6, 14).border = border

    # --- ROW 8 of Block (start_r + 7): Right Summary Table Headers (Cols O to T) ---
    right_headers = ['Activities', 'BUDGET', 'SPENT', 'SV Charges', 'TOTAL IV', 'BALANCE']
    for c_idx, h_text in enumerate(right_headers, 15):
        cell = ws.cell(start_r + 7, c_idx, h_text)
        cell.font = green_header_font
        cell.border = border
        cell.alignment = _center()

    # --- Main Table Empty Grid (Cols A to N, Rows start_r + 6 to start_r + 15) ---
    for r_offset in range(6, 16):
        for c in range(1, 15):
            ws.cell(start_r + r_offset, c).border = border

    # --- Right Summary Table Data Rows (Rows start_r + 8 to start_r + 15, Cols O to T) ---
    first_data_r = start_r + 8
    last_data_r = start_r + 15

    for idx in range(8):
        curr_r = first_data_r + idx
        if idx < len(activities):
            act_item = activities[idx]
            ws.cell(curr_r, 15, act_item['activity']).font = red_bold if idx == 0 else regular_font
            
            b_cell = ws.cell(curr_r, 16, act_item['budget'])
            b_cell.font = green_bold if idx == 0 else regular_font
            b_cell.number_format = '#,##0'
        else:
            ws.cell(curr_r, 15, 0).font = red_bold if idx == 0 else regular_font
            ws.cell(curr_r, 16, 0).font = green_bold if idx == 0 else regular_font

        ws.cell(curr_r, 17, 0).font = regular_font # SPENT
        ws.cell(curr_r, 18, 0).font = regular_font # SV Charges
        ws.cell(curr_r, 19, 0).font = regular_font # TOTAL IV
        
        bal_cell = ws.cell(curr_r, 20, f"=P{curr_r}-S{curr_r}")
        bal_cell.font = green_bold if idx == 0 else regular_font
        bal_cell.number_format = '#,##0'

        for c in range(15, 21):
            ws.cell(curr_r, c).border = border
            if c >= 16:
                ws.cell(curr_r, c).alignment = Alignment(horizontal='right', vertical='center')

    # --- Right Summary Table Total Row (Row start_r + 16, Cols O to T) ---
    tot_r = start_r + 16
    
    t_lbl = ws.cell(tot_r, 15, "TOTAL")
    t_lbl.font = red_bold
    t_lbl.border = border

    t_bud = ws.cell(tot_r, 16, f"=SUM(P{first_data_r}:P{last_data_r})")
    t_bud.font = green_bold
    t_bud.border = border
    t_bud.number_format = '#,##0'

    t_spent = ws.cell(tot_r, 17, f"=SUM(Q{first_data_r}:Q{last_data_r})")
    t_spent.font = regular_font
    t_spent.border = border

    t_sv = ws.cell(tot_r, 18, f"=SUM(R{first_data_r}:R{last_data_r})")
    t_sv.font = regular_font
    t_sv.border = border

    t_tot_iv = ws.cell(tot_r, 19, f"=SUM(S{first_data_r}:S{last_data_r})")
    t_tot_iv.font = regular_font
    t_tot_iv.border = border

    t_bal = ws.cell(tot_r, 20, f"=P{start_r + 3}-S{tot_r}")
    t_bal.font = red_bold
    t_bal.border = border
    t_bal.number_format = '#,##0'

    # --- Footer Zeros Row (Row start_r + 17, Cols L, M, N) ---
    for c in [12, 13, 14]:
        cell = ws.cell(start_r + 17, c, 0)
        cell.font = red_bold

def count_cards_on_sheet(ws):
    count = 0
    for b_idx in range(11):
        r = 1 + b_idx * 19
        if r <= ws.max_row:
            val = ws.cell(r, 1).value
            if val and str(val).strip().startswith('500'):
                count += 1
            else:
                break
        else:
            break
    return count

def get_existing_card_pos(wb):
    card_pos = set()
    for sheetname in wb.sheetnames:
        if sheetname == 'Sheet1':
            continue
        ws = wb[sheetname]
        for b_idx in range(11):
            r = 1 + b_idx * 19
            if r <= ws.max_row:
                po_val = ws.cell(r, 1).value
                if po_val and str(po_val).strip().startswith('500'):
                    card_pos.add(str(po_val).strip())
    return card_pos

from excel_parser import load_any_workbook

def generate_fmc_step2_summaries(excel_file_path, territory='Nandyal', am_name='Madhavareddy'):
    excel_path = Path(excel_file_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found at: {excel_file_path}")

    wb = load_any_workbook(excel_path)
    
    if 'Sheet1' not in wb.sheetnames:
        raise ValueError("Sheet1 not found in the uploaded Excel file. Please run Step 1 first.")
        
    ws_master = wb['Sheet1']
    pos = parse_pos_from_step1_sheet(ws_master)

    if not pos:
        raise ValueError("No PO data found in Sheet1.")

    existing_pos = get_existing_card_pos(wb)
    new_pos = [p for p in pos if p['po_number'] not in existing_pos]

    if not new_pos:
        card_sheets = [s for s in wb.sheetnames if s != 'Sheet1']
        wb.close()
        return 0, len(card_sheets)

    new_cards_added = len(new_pos)
    card_sheet_names = [s for s in wb.sheetnames if s != 'Sheet1']

    col_widths = {
        'A': 11, 'B': 11, 'C': 11, 'D': 13, 'E': 10,
        'F': 12, 'G': 12, 'H': 12, 'I': 10, 'J': 10,
        'K': 12, 'L': 14, 'M': 14, 'N': 12, 'O': 25,
        'P': 12, 'Q': 10, 'R': 12, 'S': 12, 'T': 12
    }

    def apply_col_widths(ws):
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

    # Fill remaining slots on the last card sheet if < 11 cards
    if card_sheet_names:
        last_sheet_name = card_sheet_names[-1]
        ws_last = wb[last_sheet_name]
        count_last = count_cards_on_sheet(ws_last)

        if count_last < 11:
            space = 11 - count_last
            to_add_last = new_pos[:space]
            new_pos = new_pos[space:]

            for i, po_item in enumerate(to_add_last):
                b_idx = count_last + i
                start_r = 1 + b_idx * 19
                po_date_val = po_item.get('po_date', '').strip()
                kwargs = {'territory': territory, 'am_name': am_name}
                if po_date_val:
                    kwargs['date_str'] = po_date_val
                write_po_summary_block(ws_last, start_r, po_item, **kwargs)

            first_po_val = str(ws_last.cell(1, 1).value).strip()
            first_po_short = first_po_val[-5:]
            last_po_short = to_add_last[-1]['po_number'][-5:]
            new_sheet_name = f"{first_po_short}-{last_po_short}"
            if new_sheet_name != last_sheet_name:
                ws_last.title = new_sheet_name

    # Create new card sheets for remaining new POs
    if new_pos:
        chunk_size = 11
        po_chunks = [new_pos[i:i + chunk_size] for i in range(0, len(new_pos), chunk_size)]

        for chunk in po_chunks:
            first_po_short = chunk[0]['po_number'][-5:]
            last_po_short = chunk[-1]['po_number'][-5:]
            sheet_name = f"{first_po_short}-{last_po_short}"

            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
                
            ws = wb.create_sheet(title=sheet_name)
            ws.views.sheetView[0].showGridLines = True

            for b_idx, po_item in enumerate(chunk):
                start_r = 1 + b_idx * 19
                po_date_val = po_item.get('po_date', '').strip()
                kwargs = {'territory': territory, 'am_name': am_name}
                if po_date_val:
                    kwargs['date_str'] = po_date_val
                write_po_summary_block(ws, start_r, po_item, **kwargs)

            apply_col_widths(ws)

    wb.save(excel_path)
    wb.close()
    final_card_sheets = [s for s in wb.sheetnames if s != 'Sheet1']
    return new_cards_added, len(final_card_sheets)

