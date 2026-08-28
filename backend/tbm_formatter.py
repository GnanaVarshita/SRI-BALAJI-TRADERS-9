import os
import re
import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import xlrd
except ImportError:
    xlrd = None

def clean_str(val):
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in ["none", "null", "nan"]:
        return ""
    return s

def parse_num(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0

def parse_date_intelligent(val):
    """
    Intelligently parses dates from datetime, date, Excel serial numbers,
    or diverse string formats (with slashes, dashes, backslashes, dots, month names),
    and normalizes them to DD-MM-YYYY.
    """
    if val is None or str(val).strip() == "":
        return ""
    
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%d-%m-%Y")
    
    s = str(val).strip()
    if not s:
        return ""

    # Check if numeric Excel serial date (e.g. 45447)
    try:
        num = float(s)
        if 35000 <= num <= 65000:
            dt = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(num))
            return dt.strftime("%d-%m-%Y")
    except ValueError:
        pass

    # Normalize delimiters
    s_norm = s.replace("\\\\", "/").replace("\\", "/").replace(".", "/").replace("-", "/").strip()

    # Try standard string formats
    for fmt in [
        "%Y/%m/%d %H:%M:%S", "%Y/%m/%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y",
        "%d/%b/%Y", "%d/%B/%Y", "%d/%b/%y", "%d/%B/%y",
        "%b/%d/%Y", "%B/%d/%Y", "%b/%d/%y", "%B/%d/%y"
    ]:
        try:
            dt = datetime.datetime.strptime(s_norm, fmt)
            return dt.strftime("%d-%m-%Y")
        except ValueError:
            pass

    # Regex for day/month/year components
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})", s_norm)
    if m:
        p1, p2, p3 = int(m.group(1)), int(m.group(2)), int(m.group(3))
        yr = p3 if p3 >= 100 else (2000 + p3)
        if p1 > 12:
            day, month = p1, p2
        elif p2 > 12:
            day, month = p2, p1
        else:
            # Default DD/MM/YYYY for Indian context
            day, month = p1, p2
        try:
            dt = datetime.datetime(yr, month, day)
            return dt.strftime("%d-%m-%Y")
        except Exception:
            pass

    return s

FIELD_KEYWORDS = {
    'sl_no': ['sl no', 'sl.no', 's no', 's.no', 'sno', 'si no', 'si.no', 's. no', 'sl. no', 'slno'],
    'date': ['date'],
    'zdgm': ['zdgm', 'zdgl', 'area manager', 'adgl', 'adg', 'dm', 'zdsm', 'manager'],
    'tbm': ['tbm name', 'tbm', 'name of the tbm', 'tbm/sc/so', 'sc/so', 'tbm / sc / so'],
    'mdo': ['mdo name', 'mdo'],
    'territory': ['territory', 'tbm territory', 'area', 'place', 'location'],
    'product': ['product', 'item', 'brand', 'product name'],
    'crop': ['crop', 'crops'],
    'activity': ['type of activity', 'activity name', 'activity', 'activities'],
    'village': ['village name', 'village', 'villages', 'town'],
    'farmers': ['no.of farmers', 'no of farmers', 'no. of farmers', 'n0 of farmers', 'no of rarmers', 'farmers attended', 'farmers', 'no of farmer', 'no.of farmer'],
    'tent': ['tent/hall /chairs expenses', 'tent/hall/chairs', 'tent/hall suppliers charges', 'tent/ hall', 'tent', 'chairs', 'chairs/ table/tent', 'suppliers/charges', 'suppliers charges', 'suppliers', 'hall charges', 'tent charges'],
    'food': ['food expense', 'food expenses', 'food/snacks', 'food', 'expenses', 'food expences', 'snacks', 'tiffin'],
    'transport': ['transport', 'trans port', 'auto charges', 'auto', 'travelling', 'travel'],
    'others': ['others/gifts', 'others / gifts', 'others / gift', 'others', 'gifts', 'saplaires', 'other', 'gift'],
    'total': ['total amount', 'total', 'amount'],
    'po_number': ['po number', 'po.no', 'ponumber', 'po #', 'po', 'po no']
}

def extract_activities_from_sheet(ws, default_tbm_name="", default_territory=""):
    """
    Scans worksheet for ALL tables / header rows (supporting single or multiple tables on the same sheet).
    Accurately maps columns, avoids mixing banner titles with headers, and extracts all activity records.
    """
    all_activities = []
    header_rows = []
    max_r = ws.max_row or 100
    max_c = min(ws.max_column or 35, 45)

    # 1. Detect all header rows across the sheet
    for r in range(1, max_r + 1):
        col_map = {}
        score = 0
        is_banner = False

        for c in range(1, max_c + 1):
            txt = clean_str(ws.cell(r, c).value).lower()
            if 'activities expenses by' in txt or 'marketing activities' in txt or 'activities expenses' in txt:
                is_banner = True
                break

        if is_banner:
            continue

        for c in range(1, max_c + 1):
            curr_txt = clean_str(ws.cell(r, c).value).lower()
            if not curr_txt:
                continue

            # Prioritize sl_no for column 1 or sl/sno labels
            if any(kw == curr_txt or kw in curr_txt for kw in FIELD_KEYWORDS['sl_no']):
                if 'sl_no' not in col_map:
                    col_map['sl_no'] = c
                    score += 1
                    continue

            for field, kw_list in FIELD_KEYWORDS.items():
                if field == 'sl_no':
                    continue
                if field not in col_map:
                    if any(kw == curr_txt or kw in curr_txt for kw in kw_list):
                        col_map[field] = c
                        score += 1
                        break

        # A valid table header row matches at least 3 standard columns including date, product, activity, etc.
        if score >= 3 and any(k in col_map for k in ['date', 'product', 'activity', 'mdo', 'zdgm', 'territory', 'sl_no', 'farmers']):
            header_rows.append((r, col_map))

    if not header_rows:
        slip_act = parse_slip_format(ws, default_tbm_name, default_territory)
        if slip_act:
            return [slip_act]
        return []

    # 2. Extract data rows for each detected table header
    for i, (hr, cmap) in enumerate(header_rows):
        next_hr = header_rows[i + 1][0] if i + 1 < len(header_rows) else max_r + 1

        for r in range(hr + 1, next_hr):
            row_vals = [ws.cell(r, c).value for c in range(1, max_c + 1)]
            if not any(row_vals):
                continue

            # Check if this row is a TOTAL / Summary row
            row_str = ' '.join(clean_str(v).lower() for v in row_vals)
            has_data_content = any(
                clean_str(ws.cell(r, cmap.get(k, 0)).value)
                for k in ['product', 'activity', 'date', 'village']
                if k in cmap and cmap.get(k, 0) <= max_c
            )
            if 'total' in row_str and not has_data_content:
                # Reached bottom of current table
                break

            raw_date = ws.cell(r, cmap.get('date', 0)).value if 'date' in cmap else None
            date_val = parse_date_intelligent(raw_date)

            prod_val = clean_str(ws.cell(r, cmap.get('product', 0)).value) if 'product' in cmap else ''
            act_val = clean_str(ws.cell(r, cmap.get('activity', 0)).value) if 'activity' in cmap else ''
            crop_val = clean_str(ws.cell(r, cmap.get('crop', 0)).value) if 'crop' in cmap else ''
            vlg_val = clean_str(ws.cell(r, cmap.get('village', 0)).value) if 'village' in cmap else ''
            zdgm_val = clean_str(ws.cell(r, cmap.get('zdgm', 0)).value) if 'zdgm' in cmap else ''
            mdo_val = clean_str(ws.cell(r, cmap.get('mdo', 0)).value) if 'mdo' in cmap else ''
            terr_val = clean_str(ws.cell(r, cmap.get('territory', 0)).value) if 'territory' in cmap else ''
            if not terr_val:
                terr_val = default_territory

            tbm_val = clean_str(ws.cell(r, cmap.get('tbm', 0)).value) if 'tbm' in cmap else ''
            # Guard against TBM column mapping to Sl No (digits)
            if not tbm_val or tbm_val.isdigit():
                tbm_val = default_tbm_name

            farm_val = parse_num(ws.cell(r, cmap.get('farmers', 0)).value) if 'farmers' in cmap else 0.0
            tent_val = parse_num(ws.cell(r, cmap.get('tent', 0)).value) if 'tent' in cmap else 0.0
            food_val = parse_num(ws.cell(r, cmap.get('food', 0)).value) if 'food' in cmap else 0.0
            trans_val = parse_num(ws.cell(r, cmap.get('transport', 0)).value) if 'transport' in cmap else 0.0
            oth_val = parse_num(ws.cell(r, cmap.get('others', 0)).value) if 'others' in cmap else 0.0
            tot_val = parse_num(ws.cell(r, cmap.get('total', 0)).value) if 'total' in cmap else 0.0
            if tot_val == 0.0:
                tot_val = tent_val + food_val + trans_val + oth_val

            # Row must be a genuine activity row with at least one descriptive field
            if not prod_val and not act_val and not date_val and not vlg_val and not crop_val and tot_val == 0.0:
                continue
            if not prod_val and not act_val and not date_val and not vlg_val and not crop_val and not (zdgm_val and terr_val):
                continue

            # Extract PO number directly from row
            po_val = clean_str(ws.cell(r, cmap.get('po_number', 0)).value) if 'po_number' in cmap else ''
            if not po_val:
                for c in range(1, max_c + 1):
                    v = clean_str(ws.cell(r, c).value)
                    m = re.search(r'5\d{2}[A-Z0-9]{8,20}', v, re.I)
                    if m:
                        po_val = m.group(0).upper()
                        break

            all_activities.append({
                'date': date_val,
                'zdgm': zdgm_val,
                'tbm': tbm_val,
                'mdo': mdo_val,
                'territory': terr_val,
                'product': prod_val,
                'crop': crop_val,
                'activity': act_val,
                'village': vlg_val,
                'farmers': int(farm_val),
                'tent': tent_val,
                'food': food_val,
                'transport': trans_val,
                'others': oth_val,
                'total': tot_val,
                'po_number': po_val
            })

    return all_activities

def parse_slip_format(ws, default_tbm_name="", default_territory=""):
    kv = {}
    for r in range(1, 25):
        k = clean_str(ws.cell(r, 2).value).lower()
        v = ws.cell(r, 4).value
        if k:
            if 'date' in k: kv['date'] = parse_date_intelligent(v)
            elif 'product' in k: kv['product'] = clean_str(v)
            elif 'crop' in k: kv['crop'] = clean_str(v)
            elif 'activity' in k: kv['activity'] = clean_str(v)
            elif 'place' in k or 'territory' in k: kv['territory'] = clean_str(v)
            elif 'farmers' in k: kv['farmers'] = int(parse_num(v))
            elif 'food' in k: kv['food'] = parse_num(v)
            elif 'chairs' in k or 'tent' in k: kv['tent'] = parse_num(v)
            elif 'auto' in k or 'transport' in k: kv['transport'] = parse_num(v)
            elif 'others' in k: kv['others'] = parse_num(v)
            elif 'tbm' in k: kv['tbm'] = clean_str(v)
            elif 'po' in k: kv['po_number'] = clean_str(v)

    if kv.get('activity') or kv.get('product') or kv.get('date'):
        tbm_val = kv.get('tbm') or default_tbm_name
        terr_val = kv.get('territory') or default_territory
        tent = kv.get('tent', 0.0)
        food = kv.get('food', 0.0)
        transport = kv.get('transport', 0.0)
        others = kv.get('others', 0.0)
        total = tent + food + transport + others
        return {
            'date': kv.get('date', ''),
            'zdgm': '',
            'tbm': tbm_val,
            'mdo': '',
            'territory': terr_val,
            'product': kv.get('product', ''),
            'crop': kv.get('crop', ''),
            'activity': kv.get('activity', ''),
            'village': '',
            'farmers': kv.get('farmers', 0),
            'tent': tent,
            'food': food,
            'transport': transport,
            'others': others,
            'total': total,
            'po_number': kv.get('po_number', '')
        }
    return None

def create_green_styles():
    """
    Creates styles matching the screenshot layout with bold green headings and clean borders.
    """
    green_color = '006100'  # Dark Excel green
    
    thin_border = Border(
        left=Side(style='thin', color='A6A6A6'),
        right=Side(style='thin', color='A6A6A6'),
        top=Side(style='thin', color='A6A6A6'),
        bottom=Side(style='thin', color='A6A6A6')
    )

    total_border = Border(
        left=Side(style='thin', color='A6A6A6'),
        right=Side(style='thin', color='A6A6A6'),
        top=Side(style='thin', color='006100'),
        bottom=Side(style='double', color='006100')
    )

    title_font = Font(name='Calibri', size=11, bold=True, color=green_color)
    hdr_font = Font(name='Calibri', size=10, bold=True, color=green_color)
    green_text_font = Font(name='Calibri', size=10, color=green_color)
    green_bold_font = Font(name='Calibri', size=10, bold=True, color=green_color)
    total_lbl_font = Font(name='Calibri', size=11, bold=True, color=green_color)
    total_val_font = Font(name='Calibri', size=11, bold=True, color=green_color)

    return {
        'thin_border': thin_border,
        'total_border': total_border,
        'title_font': title_font,
        'hdr_font': hdr_font,
        'green_text_font': green_text_font,
        'green_bold_font': green_bold_font,
        'total_lbl_font': total_lbl_font,
        'total_val_font': total_val_font
    }

def build_grouped_tables(activities):
    """
    Groups activities hierarchically:
    1. By PO Number (POs first, then unassigned/NO_PO)
    2. Within PO, by (Product, Activity)
    Returns an ordered list of groups:
    [
      {
        'po_number': '500882025128285',
        'product': 'Keenali',
        'activity': 'Field Day',
        'territory': 'Kavali',
        'tbm': 'Surendra',
        'rows': [...]
      },
      ...
    ]
    """
    grouped_map = {}
    for r in activities:
        po_num = str(r.get('po_number', '')).strip().upper()
        prod = str(r.get('product', '')).strip().title() or "General Product"
        act = str(r.get('activity', '')).strip().title() or "General Activity"
        terr = str(r.get('territory', '')).strip().title()
        tbm = str(r.get('tbm', '')).strip().title()

        group_key = (po_num, prod, act)
        if group_key not in grouped_map:
            grouped_map[group_key] = {
                'po_number': po_num,
                'product': prod,
                'activity': act,
                'territory': terr,
                'tbm': tbm,
                'rows': []
            }
        
        # update territory/tbm if found
        if terr and not grouped_map[group_key]['territory']:
            grouped_map[group_key]['territory'] = terr
        if tbm and not grouped_map[group_key]['tbm']:
            grouped_map[group_key]['tbm'] = tbm

        grouped_map[group_key]['rows'].append(r)

    # Sort groups: POs first (alphabetical), NO_PO at the end, then Product, then Activity
    def sort_key(item):
        po, prod, act = item[0]
        is_no_po = 1 if not po else 0
        return (is_no_po, po, prod, act)

    sorted_groups = [v for k, v in sorted(grouped_map.items(), key=sort_key)]
    return sorted_groups

def write_formatted_group_table(ws, start_row, group_data, styles):
    """
    Writes a single formatted table for a (PO, Product, Activity) group matching the screenshot.
    Returns (next_start_r, tot_row).
    """
    tbm_name = group_data.get('tbm', '')
    territory = group_data.get('territory', '')
    po_number = group_data.get('po_number', '')
    rows = group_data.get('rows', [])

    # Format Title banner: "Activities Expenses by {Territory} TBM {TBM Name}"
    title_parts = ["Activities Expenses by"]
    if territory:
        title_parts.append(territory)
    if tbm_name and not tbm_name.isdigit():
        title_parts.append(f"TBM {tbm_name}")
    elif not territory:
        title_parts.append("TBM")
    
    title_text = " ".join(title_parts)

    title_row = start_row
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=17)
    title_cell = ws.cell(title_row, 1, title_text)
    title_cell.font = styles['title_font']
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[title_row].height = 24

    for c in range(1, 18):
        ws.cell(title_row, c).border = styles['thin_border']

    # Column Headers (Row 2 of table)
    hdr_row = title_row + 1
    headers = [
        "SI No.", "Date", "ZDGM", "TBM", "MDO", "Territory", "Product", "Crop", "Activity", "Village",
        "No.of Farmers", "Tent/Hall Suppliers Charges", "Food Expenses", "Transport", "Others/Gifts", "Total", "PO Number"
    ]
    ws.row_dimensions[hdr_row].height = 32
    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws.cell(hdr_row, col_idx, h_text)
        cell.font = styles['hdr_font']
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = styles['thin_border']

    # Data Rows
    current_r = hdr_row + 1
    data_start_r = current_r

    for idx, item in enumerate(rows, start=1):
        ws.row_dimensions[current_r].height = 20

        c_sno = ws.cell(current_r, 1, idx)
        c_sno.font = styles['green_bold_font']
        c_sno.alignment = Alignment(horizontal='center', vertical='center')

        c_date = ws.cell(current_r, 2, item.get('date', ''))
        c_date.font = styles['green_text_font']
        c_date.alignment = Alignment(horizontal='center', vertical='center')

        # Text fields
        field_map = [
            (3, 'zdgm', 'left'),
            (4, 'tbm', 'left'),
            (5, 'mdo', 'left'),
            (6, 'territory', 'left'),
            (7, 'product', 'left'),
            (8, 'crop', 'left'),
            (9, 'activity', 'left'),
            (10, 'village', 'left'),
        ]
        for col_idx, key, align_h in field_map:
            val = item.get(key, '')
            if key == 'tbm' and (not val or val.isdigit()):
                val = tbm_name
            if key == 'territory' and not val:
                val = territory
            c_txt = ws.cell(current_r, col_idx, val)
            c_txt.font = styles['green_text_font']
            c_txt.alignment = Alignment(horizontal=align_h, vertical='center')

        c_farmers = ws.cell(current_r, 11, item.get('farmers', 0))
        c_farmers.font = styles['green_text_font']
        c_farmers.alignment = Alignment(horizontal='center', vertical='center')
        c_farmers.number_format = '#,##0'

        for c_offset, key in enumerate(['tent', 'food', 'transport', 'others'], start=12):
            amt_val = item.get(key, 0.0)
            c_amt = ws.cell(current_r, c_offset, amt_val if amt_val > 0 else "")
            c_amt.font = styles['green_text_font']
            c_amt.alignment = Alignment(horizontal='right', vertical='center')
            if amt_val > 0:
                c_amt.number_format = '#,##0'

        # Row Total formula
        c_tot = ws.cell(current_r, 16, f"=SUM(L{current_r}:O{current_r})")
        c_tot.font = styles['green_text_font']
        c_tot.alignment = Alignment(horizontal='right', vertical='center')
        c_tot.number_format = '#,##0'

        row_po = item.get('po_number') or po_number or ''
        c_po = ws.cell(current_r, 17, row_po)
        c_po.font = styles['green_text_font']
        c_po.alignment = Alignment(horizontal='center', vertical='center')

        for c in range(1, 18):
            ws.cell(current_r, c).border = styles['thin_border']

        current_r += 1

    data_end_r = current_r - 1

    # Empty row below table data
    empty_row = current_r
    ws.row_dimensions[empty_row].height = 18
    current_r += 1

    # Total Row (Row 25 style in screenshot)
    tot_row = current_r
    ws.row_dimensions[tot_row].height = 22

    lbl_cell = ws.cell(tot_row, 15, "Total")
    lbl_cell.font = styles['total_lbl_font']
    lbl_cell.alignment = Alignment(horizontal='right', vertical='center')

    sum_tot_cell = ws.cell(tot_row, 16, f"=SUM(P{data_start_r}:P{data_end_r})")
    sum_tot_cell.font = styles['total_val_font']
    sum_tot_cell.alignment = Alignment(horizontal='right', vertical='center')
    sum_tot_cell.number_format = '#,##0'

    # Next table starts 3 rows below
    next_start_r = tot_row + 3
    return next_start_r, tot_row

def write_po_and_grand_totals_block(ws, start_row, po_totals, styles):
    """
    Writes a summary table at the end of Sheet2 listing each PO total and the Grand Total.
    Layout matches user screenshot:
      PO Number (Col 15)  |  Total of this PO (Col 16)
      Grand Total (Col 15)|  Sum of this (Col 16)
    """
    current_r = start_row
    po_summary_rows = []

    for po_num, tot_rows in po_totals.items():
        ws.row_dimensions[current_r].height = 22
        display_po = po_num if po_num != "NO_PO" else "No PO"

        c_po = ws.cell(current_r, 15, display_po)
        c_po.font = styles['green_bold_font']
        c_po.alignment = Alignment(horizontal='right', vertical='center')
        c_po.border = styles['thin_border']

        refs = [f"P{r}" for r in tot_rows]
        sum_formula = f"=SUM({','.join(refs)})" if len(refs) > 1 else f"={refs[0]}"

        c_tot = ws.cell(current_r, 16, sum_formula)
        c_tot.font = styles['total_val_font']
        c_tot.alignment = Alignment(horizontal='right', vertical='center')
        c_tot.number_format = '#,##0'
        c_tot.border = styles['thin_border']

        po_summary_rows.append(current_r)
        current_r += 1

    # Grand Total Row
    ws.row_dimensions[current_r].height = 24
    c_g_lbl = ws.cell(current_r, 15, "Grand Total")
    c_g_lbl.font = styles['total_lbl_font']
    c_g_lbl.alignment = Alignment(horizontal='right', vertical='center')
    c_g_lbl.border = styles['total_border']

    if po_summary_rows:
        g_formula = f"=SUM(P{po_summary_rows[0]}:P{po_summary_rows[-1]})"
    else:
        g_formula = "=0"

    c_g_val = ws.cell(current_r, 16, g_formula)
    c_g_val.font = styles['total_val_font']
    c_g_val.alignment = Alignment(horizontal='right', vertical='center')
    c_g_val.number_format = '#,##0'
    c_g_val.border = styles['total_border']

    return current_r + 2

def format_tbm_workbook(file_path, default_tbm_name="", default_territory=""):
    """
    Reads Sheet 1 (original raw table) of an Excel file, extracts all activity records,
    groups them by (PO, Product, Activity), formats them on Sheet 2, and appends
    a PO Totals + Grand Total summary block at the end of Sheet 2.
    """
    file_path = Path(file_path).resolve()
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    # Determine default TBM name and territory from folder structure if not provided
    if not default_tbm_name:
        default_tbm_name = file_path.parent.name if file_path.parent.name != "TBM s Summary" else ""
    if not default_territory:
        for part in file_path.parts:
            p_up = part.upper()
            if any(k in p_up for k in ['KURNOOL', 'NELLORE', 'NANDYAL', 'NANDYALA', 'SURYAPET', 'KAVALI', 'ALLAGADDA', 'ADONI']):
                default_territory = part.replace("-FMC", "").replace(" POs", "").title()
                break

    wb = openpyxl.load_workbook(file_path)

    # 1. Extract raw activities from the FIRST sheet (or first non-empty sheet)
    first_sheet_name = wb.sheetnames[0]
    ws_raw = wb[first_sheet_name]
    activities = extract_activities_from_sheet(ws_raw, default_tbm_name, default_territory)

    # If first sheet returned no activities and there are other sheets, search other raw sheets
    if not activities and len(wb.sheetnames) > 1:
        for sname in wb.sheetnames[1:]:
            if "formatted" not in sname.lower() and sname != "Sheet2":
                ws_alt = wb[sname]
                alt_acts = extract_activities_from_sheet(ws_alt, default_tbm_name, default_territory)
                if alt_acts:
                    activities.extend(alt_acts)
                    break

    if not activities:
        wb.close()
        return {
            "success": False,
            "message": f"No valid activity records found in {file_path.name}",
            "file": file_path.name,
            "groupsCount": 0,
            "activitiesCount": 0
        }

    # 2. Build hierarchical groups: PO -> Product & Activity
    groups = build_grouped_tables(activities)

    # 3. Setup the SECOND sheet without touching Sheet 1
    # Sheet names: If Sheet2 exists, use it. If there is only 1 sheet, create Sheet2 at index 1.
    target_sheet_name = "Sheet2"
    if len(wb.sheetnames) >= 2:
        second_sheet_name = wb.sheetnames[1]
        wb.remove(wb[second_sheet_name])
        ws_target = wb.create_sheet(title=second_sheet_name, index=1)
    else:
        ws_target = wb.create_sheet(title=target_sheet_name, index=1)

    ws_target.views.sheetView[0].showGridLines = True

    # Setup column widths
    col_widths = {
        'A': 8,   # SI No.
        'B': 13,  # Date
        'C': 18,  # ZDGM
        'D': 18,  # TBM
        'E': 16,  # MDO
        'F': 14,  # Territory
        'G': 15,  # Product
        'H': 12,  # Crop
        'I': 14,  # Activity
        'J': 16,  # Village
        'K': 14,  # No.of Farmers
        'L': 24,  # Tent/Hall Suppliers Charges
        'M': 15,  # Food Expenses
        'N': 12,  # Transport
        'O': 20,  # Others/Gifts (Wide enough for PO Numbers / Grand Total labels)
        'P': 16,  # Total
        'Q': 18   # PO Number
    }
    for col_let, w in col_widths.items():
        ws_target.column_dimensions[col_let].width = w

    styles = create_green_styles()

    current_r = 1
    po_totals = {}  # maps po_number -> list of table total row numbers

    for group_data in groups:
        po_num = str(group_data.get('po_number', '')).strip().upper() or "NO_PO"
        current_r, tbl_tot_r = write_formatted_group_table(ws_target, current_r, group_data, styles)
        if po_num not in po_totals:
            po_totals[po_num] = []
        po_totals[po_num].append(tbl_tot_r)

    # 4. Write Each PO Total and Grand Total block at the end of Sheet2
    write_po_and_grand_totals_block(ws_target, current_r, po_totals, styles)

    wb.save(file_path)
    wb.close()

    return {
        "success": True,
        "message": f"Formatted {len(activities)} activities across {len(groups)} table(s) on second sheet of {file_path.name}",
        "file": file_path.name,
        "filePath": str(file_path),
        "groupsCount": len(groups),
        "activitiesCount": len(activities),
        "poTotals": list(po_totals.keys())
    }

def format_all_tbm_summaries_in_folder(tbm_folder_path):
    """
    Batch formats all TBM Excel summaries in a folder (and its TBM subfolders).
    """
    tbm_dir = Path(tbm_folder_path).resolve()
    if not tbm_dir.exists() or not tbm_dir.is_dir():
        raise ValueError(f"TBM Summary folder path does not exist: {tbm_folder_path}")

    # Find all Excel files inside TBM folder and its subfolders
    excel_files = [
        f for f in tbm_dir.rglob("*")
        if f.is_file() 
        and f.suffix.lower() in ['.xlsx', '.xlsm'] 
        and not f.name.startswith('~$') 
        and "-All-TBMs-Summary" not in f.name
    ]

    if not excel_files:
        return {
            "success": False,
            "message": f"No Excel files found in {tbm_dir.name}",
            "processedFiles": 0,
            "totalActivities": 0,
            "details": []
        }

    results = []
    total_acts = 0
    total_groups = 0
    success_count = 0

    for ef in excel_files:
        tbm_folder_name = ef.parent.name if ef.parent != tbm_dir else ""
        try:
            res = format_tbm_workbook(ef, default_tbm_name=tbm_folder_name)
            results.append(res)
            if res.get("success"):
                success_count += 1
                total_acts += res.get("activitiesCount", 0)
                total_groups += res.get("groupsCount", 0)
        except Exception as e:
            results.append({
                "success": False,
                "file": ef.name,
                "message": f"Error formatting {ef.name}: {e}"
            })

    return {
        "success": True,
        "message": f"Successfully formatted {success_count} / {len(excel_files)} TBM summary file(s) ({total_acts} activities in {total_groups} tables) into their second sheets!",
        "processedFiles": success_count,
        "totalFiles": len(excel_files),
        "totalActivities": total_acts,
        "totalTables": total_groups,
        "details": results
    }

if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        target = sys.argv[1]
        p = Path(target)
        if p.is_dir():
            res = format_all_tbm_summaries_in_folder(p)
            print(res["message"])
        elif p.is_file():
            res = format_tbm_workbook(p)
            print(res["message"])
    else:
        print("Usage: python tbm_formatter.py <tbm_summary_folder_or_file>")
