import os
import re
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure backend directory is in python path
sys.path.append(str(Path(__file__).resolve().parent))
try:
    from excel_parser import load_any_workbook
except ImportError:
    from .excel_parser import load_any_workbook

ACTIVITY_NORMALIZATION = {
    'FIELD DAYS': 'FD',
    'FIELD DAY': 'FD',
    'CROP SHOW/FIELD DAYS': 'FD',
    'CROP SHOW / FIELD DAYS': 'FD',
    'FD': 'FD',
    'ORGANIZED FARMER MEETING/VILLAGE MEETING': 'OFM',
    'ORGANIZED FARMER MEETING / VILLAGE MEETING': 'OFM',
    'ORGANIZED FARMER MEETING': 'OFM',
    'VILLAGE MEETING': 'OFM',
    'OFM': 'OFM',
    'HARVEST DAYS': 'HD',
    'HARVEST DAY': 'HD',
    'HD': 'HD',
    'LARGE FARMER MEETING': 'LFM',
    'LFM': 'LFM',
    'DEMO ACTIVITY': 'DA',
    'DEMONSTRATION ACTIVITY': 'DA',
    'DA': 'DA',
    'OTHER BRANDING ACTIVITY': 'OBA',
    'OBA': 'OBA',
    'VIDEO SHOOT': 'VIDEO SHOOT',
    'VIDEO SHOOT / FARMER TESTIMONIAL': 'VIDEO SHOOT',
    'FARMER TESTIMONIAL': 'VIDEO SHOOT',
    'SKILL DEVELOPMENT TRAINING': 'SKILL DEVELOPMENT TRAINING',
    'SDT': 'SKILL DEVELOPMENT TRAINING'
}

def normalize_activity(act):
    if not act:
        return ""
    s = str(act).strip().upper()
    return ACTIVITY_NORMALIZATION.get(s, s)

def extract_tbm_spent_summary(tbm_summary_path):
    """
    Extracts spent records by PO and TBM from the Consolidated Master TBM Summary workbook.
    Reads from 'TBM Amount Summary' sheet or falls back to extracting across all activity sheets.
    Returns: dict mapping PO Number -> list of {tbm, product, crop, activity, num_activities, total_amount}
    """
    tbm_file = Path(tbm_summary_path).resolve()
    if not tbm_file.exists():
        raise FileNotFoundError(f"TBM Summary file not found at: {tbm_summary_path}")

    wb = load_any_workbook(tbm_file)
    tbm_data = {}

    # 1. Prefer 'TBM Amount Summary' sheet
    target_sheet = None
    for sname in wb.sheetnames:
        if 'tbm amount summary' in sname.lower() or 'amount summary' in sname.lower():
            target_sheet = wb[sname]
            break

    if target_sheet is not None:
        header_row = 2
        for r in range(1, min(10, target_sheet.max_row + 1)):
            c1 = str(target_sheet.cell(r, 1).value or '').strip().lower()
            c2 = str(target_sheet.cell(r, 2).value or '').strip().lower()
            if 'po' in c1 or 'po' in c2 or 'tbm' in c2:
                header_row = r
                break

        for r in range(header_row + 1, target_sheet.max_row + 1):
            po_val = str(target_sheet.cell(r, 1).value or '').strip().upper()
            if not po_val or 'TOTAL' in po_val or 'GRAND' in po_val:
                continue

            tbm_val = str(target_sheet.cell(r, 2).value or '').strip()
            prod_val = str(target_sheet.cell(r, 3).value or '').strip()
            crop_val = str(target_sheet.cell(r, 4).value or '').strip()
            act_val = str(target_sheet.cell(r, 5).value or '').strip()
            
            try:
                num_act_val = int(target_sheet.cell(r, 6).value or 0)
            except Exception:
                num_act_val = 0

            try:
                amt_val = float(str(target_sheet.cell(r, 7).value or '0').replace(',', '').strip())
            except Exception:
                amt_val = 0.0

            if po_val not in tbm_data:
                tbm_data[po_val] = []

            tbm_data[po_val].append({
                'po': po_val,
                'tbm': tbm_val,
                'product': prod_val,
                'crop': crop_val,
                'activity': act_val,
                'num_activities': num_act_val,
                'total_amount': amt_val
            })
    else:
        # Fallback: scan all activity sheets
        try:
            import tbm_formatter
            for sname in wb.sheetnames:
                acts = tbm_formatter.extract_activities_from_sheet(wb[sname])
                for a in acts:
                    po_val = str(a.get('po_number', '')).strip().upper()
                    if not po_val:
                        continue
                    if po_val not in tbm_data:
                        tbm_data[po_val] = []
                    tbm_data[po_val].append({
                        'po': po_val,
                        'tbm': a.get('tbm', ''),
                        'product': a.get('product', ''),
                        'crop': a.get('crop', ''),
                        'activity': a.get('activity', ''),
                        'num_activities': 1,
                        'total_amount': float(a.get('total', 0.0))
                    })
        except Exception as e:
            print(f"Fallback extraction warning: {e}")

    wb.close()
    return tbm_data

def write_sheet_end_summary_table(ws, start_row=240):
    """
    Writes the Sheet Summary Table at row 240 (Cols H to M) summarizing
    all POs, products, crops, activities, budgets (=P{r}), and balances (=T{r}) on this sheet.
    """
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name='Calibri', size=10, bold=True)
    regular_font = Font(name='Calibri', size=10)
    bold_font = Font(name='Calibri', size=10, bold=True)

    entries = []
    for b in range(11):
        r = 1 + b * 19
        if r > ws.max_row:
            break
        po_val = ws.cell(r, 1).value
        if not po_val or not str(po_val).strip().startswith('500'):
            continue
        po_str = str(po_val).strip()
        prod = str(ws.cell(r, 7).value or '').strip()
        crop1 = str(ws.cell(r, 9).value or '').strip()
        crop2 = str(ws.cell(r + 1, 9).value or '').strip()

        for idx in range(8):
            curr_r = r + 8 + idx
            act_val = ws.cell(curr_r, 15).value
            if act_val and str(act_val).strip() not in ['0', 'None', '']:
                act_str = str(act_val).strip()
                c_crop = crop1
                if idx == 1 and crop2:
                    c_crop = crop2
                entries.append({
                    'po': po_str,
                    'prod': prod,
                    'crop': c_crop,
                    'act': act_str,
                    'budget_ref': f'P{curr_r}',
                    'bal_ref': f'T{curr_r}'
                })

    if not entries:
        return

    # Clear prior table area (rows 240 to 285, cols 8 to 13)
    for cr in range(start_row, start_row + 45):
        for cc in range(8, 14):
            ws.cell(cr, cc).value = None
            ws.cell(cr, cc).border = Border()

    # Headers in row start_row
    headers = ['Pos', 'Product', 'Crop', 'Activity', 'Budget', 'Balance']
    for c_i, h_text in enumerate(headers, start=8):
        c = ws.cell(start_row, c_i, h_text)
        c.font = header_font
        c.border = border
        c.alignment = Alignment(horizontal='right' if h_text in ['Budget', 'Balance'] else 'center', vertical='center')

    # Data rows
    for e_idx, e in enumerate(entries):
        row_r = start_row + 1 + e_idx

        # Pos
        c = ws.cell(row_r, 8, e['po'])
        c.font = regular_font
        c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')

        # Product
        c = ws.cell(row_r, 9, e['prod'])
        c.font = regular_font
        c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')

        # Crop
        c = ws.cell(row_r, 10, e['crop'])
        c.font = regular_font
        c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')

        # Activity
        c = ws.cell(row_r, 11, e['act'])
        c.font = regular_font
        c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')

        # Budget
        c = ws.cell(row_r, 12, f"={e['budget_ref']}")
        c.font = regular_font
        c.border = border
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.number_format = '#,##0'

        # Balance
        c = ws.cell(row_r, 13, f"={e['bal_ref']}")
        c.font = regular_font
        c.border = border
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.number_format = '#,##0.00'

    # TOTAL row
    tot_r = start_row + 1 + len(entries)

    c = ws.cell(tot_r, 8, 'TOTAL')
    c.font = bold_font
    c.border = border
    c.alignment = Alignment(horizontal='center', vertical='center')

    for cc in [9, 10, 11]:
        c = ws.cell(tot_r, cc, '')
        c.border = border

    c = ws.cell(tot_r, 12, f"=SUM(L{start_row+1}:L{tot_r-1})")
    c.font = bold_font
    c.border = border
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = '#,##0'

    c = ws.cell(tot_r, 13, f"=SUM(M{start_row+1}:M{tot_r-1})")
    c.font = bold_font
    c.border = border
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = '#,##0.00'

def sync_fmc_cards_workbook(wb_cards, tbm_data, service_charge_percent=5.0):
    """
    Synchronizes FMC PO Summary Cards (where cards are stacked in blocks of 19 rows across sheets).
    Updates left main table data rows, bottom totals row, right summary table with formulas,
    and updates Sheet1 master overview table with linked formulas and balances.
    """
    sv_rate = float(service_charge_percent) / 100.0
    sv_percent_str = f"{float(service_charge_percent):.4f}".rstrip('0').rstrip('.')

    regular_font = Font(name='Calibri', size=10)
    green_bold = Font(name='Calibri', size=11, bold=True, color='008000')
    red_bold = Font(name='Calibri', size=11, bold=True, color='FF0000')

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    card_links = {}  # maps (po_number, normalized_activity) -> (sheet_name, summary_row_idx)
    updated_cards_count = 0

    for sname in wb_cards.sheetnames:
        if sname.lower() == 'sheet1':
            continue
        ws = wb_cards[sname]

        for b in range(11):
            r = 1 + b * 19
            if r > ws.max_row:
                break

            po_val = str(ws.cell(r, 1).value or '').strip().upper()
            if not po_val or not po_val.startswith('500'):
                continue

            prod_card = str(ws.cell(r, 7).value or '').strip()
            crop1_card = str(ws.cell(r, 9).value or '').strip()
            crop2_card = str(ws.cell(r + 1, 9).value or '').strip()
            crop_card = f"{crop1_card} / {crop2_card}".strip(" /") if crop2_card else crop1_card
            area_card = str(ws.cell(r + 2, 7).value or '').strip()
            am_card = str(ws.cell(r + 3, 4).value or '').strip()
            prefix = po_val[3:5] if len(po_val) >= 5 else 'BB'
            budget_type = 'Brand' if prefix == 'BB' else 'Promo'

            # 1. Identify Activity Columns in row (r + 5)
            act_col_map = {}  # norm_act -> col_idx
            last_act_col = 12
            for c in range(12, 16):
                h_val = ws.cell(r + 5, c).value
                if h_val and str(h_val).strip().lower() == 'amount':
                    last_act_col = c
                    break
                elif h_val:
                    act_col_map[normalize_activity(h_val)] = c
                    last_act_col = c + 1

            amount_col = last_act_col

            # 2. Populate Main Table Data Rows (rows r + 7 to r + 15)
            first_data_r = r + 7
            last_data_r = r + 15
            table_tot_r = r + 17

            # Clear previous data rows
            for d_r in range(first_data_r, last_data_r + 1):
                for c in range(1, amount_col + 1):
                    ws.cell(d_r, c).value = None

            tbm_rows_for_po = tbm_data.get(po_val, [])
            if tbm_rows_for_po:
                updated_cards_count += 1

            for idx, t_item in enumerate(tbm_rows_for_po):
                if idx >= 9:
                    break
                row_idx = first_data_r + idx

                # Leave I.V NO and DATE blank for invoice verification
                ws.cell(row_idx, 1).value = None
                ws.cell(row_idx, 2).value = None

                ws.cell(row_idx, 3, area_card).font = regular_font
                ws.cell(row_idx, 4, po_val).font = regular_font
                ws.cell(row_idx, 5, budget_type).font = regular_font
                ws.cell(row_idx, 6, t_item['product'] or prod_card).font = regular_font
                ws.cell(row_idx, 7, t_item['crop'] or crop_card).font = regular_font
                ws.cell(row_idx, 8, t_item['activity']).font = regular_font
                ws.cell(row_idx, 9, am_card).font = regular_font
                ws.cell(row_idx, 10, t_item['tbm']).font = regular_font
                ws.cell(row_idx, 11, t_item['num_activities']).font = regular_font

                norm_act = normalize_activity(t_item['activity'])
                target_col = act_col_map.get(norm_act, 12)

                for c in range(12, amount_col):
                    if c == target_col:
                        amt_cell = ws.cell(row_idx, c, t_item['total_amount'])
                        amt_cell.font = regular_font
                        amt_cell.number_format = '#,##0'
                    else:
                        ws.cell(row_idx, c, None)

                act_cell_refs = [f"{get_column_letter(c)}{row_idx}" for c in range(12, amount_col)]
                amt_formula = f"=SUM({','.join(act_cell_refs)})" if act_cell_refs else 0
                c_amt = ws.cell(row_idx, amount_col, amt_formula)
                c_amt.font = regular_font
                c_amt.number_format = '#,##0'

            # 3. Bottom Totals row for main table (row r + 17)
            for c in range(12, amount_col):
                cl = get_column_letter(c)
                c_tot = ws.cell(table_tot_r, c, f"=SUM({cl}{first_data_r}:{cl}{last_data_r})")
                c_tot.font = red_bold
                c_tot.number_format = '#,##0'

            amt_cl = get_column_letter(amount_col)
            c_amt_tot = ws.cell(table_tot_r, amount_col, f"=SUM({amt_cl}{first_data_r}:{amt_cl}{last_data_r})")
            c_amt_tot.font = red_bold
            c_amt_tot.number_format = '#,##0'

            # 4. Right Summary Table (Rows r + 8 to r + 15)
            first_sum_r = r + 8
            last_sum_r = r + 15
            sum_tot_r = r + 16

            for idx in range(8):
                curr_r = first_sum_r + idx
                act_name_val = ws.cell(curr_r, 15).value

                if act_name_val and str(act_name_val).strip() not in ['0', 'None', '']:
                    norm_act = normalize_activity(act_name_val)
                    target_col = act_col_map.get(norm_act, 12)
                    t_cl = get_column_letter(target_col)

                    # SPENT -> points to bottom table total for this activity
                    ws.cell(curr_r, 17, f"={t_cl}{table_tot_r}").font = regular_font
                    ws.cell(curr_r, 17).number_format = '#,##0'

                    # SV Charges -> =SPENT * sv_rate
                    ws.cell(curr_r, 18, f"=Q{curr_r}*{sv_rate}").font = regular_font
                    ws.cell(curr_r, 18).number_format = '#,##0.00'

                    # TOTAL IV -> =SPENT + SV Charges
                    ws.cell(curr_r, 19, f"=Q{curr_r}+R{curr_r}").font = regular_font
                    ws.cell(curr_r, 19).number_format = '#,##0.00'

                    # BALANCE -> =BUDGET - TOTAL IV
                    ws.cell(curr_r, 20, f"=P{curr_r}-S{curr_r}").font = regular_font
                    ws.cell(curr_r, 20).number_format = '#,##0.00'

                    card_links[(po_val, norm_act)] = (sname, curr_r)
                else:
                    ws.cell(curr_r, 17, 0).font = regular_font
                    ws.cell(curr_r, 18, 0).font = regular_font
                    ws.cell(curr_r, 19, 0).font = regular_font
                    ws.cell(curr_r, 20, f"=P{curr_r}-S{curr_r}").font = regular_font

            # Total row in Right Summary Table
            ws.cell(sum_tot_r, 15, "TOTAL").font = red_bold
            
            c_sum_b = ws.cell(sum_tot_r, 16, f"=SUM(P{first_sum_r}:P{last_sum_r})")
            c_sum_b.font = green_bold
            c_sum_b.number_format = '#,##0'

            c_sum_sp = ws.cell(sum_tot_r, 17, f"=SUM(Q{first_sum_r}:Q{last_sum_r})")
            c_sum_sp.font = regular_font
            c_sum_sp.number_format = '#,##0'

            c_sum_sv = ws.cell(sum_tot_r, 18, f"=SUM(R{first_sum_r}:R{last_sum_r})")
            c_sum_sv.font = regular_font
            c_sum_sv.number_format = '#,##0.00'

            c_sum_iv = ws.cell(sum_tot_r, 19, f"=SUM(S{first_sum_r}:S{last_sum_r})")
            c_sum_iv.font = regular_font
            c_sum_iv.number_format = '#,##0.00'

            c_sum_bal = ws.cell(sum_tot_r, 20, f"=P{sum_tot_r}-S{sum_tot_r}")
            c_sum_bal.font = red_bold
            c_sum_bal.number_format = '#,##0.00'

        # 5. Write / Update Sheet End Summary Table at Row 240 (Cols H to M)
        write_sheet_end_summary_table(ws, start_row=240)

    # 6. Update Sheet1 Master Overview Table (ss4)
    if 'Sheet1' in wb_cards.sheetnames:
        ws1 = wb_cards['Sheet1']
        curr_po = ''

        for r in range(4, ws1.max_row + 1):
            val_c6 = ws1.cell(r, 6).value
            if val_c6 and str(val_c6).strip().lower() == 'total':
                # Total row in Sheet1
                ws1.cell(r, 7, f"=SUM(G4:G{r-1})")
                ws1.cell(r, 8, f"=SUM(H4:H{r-1})")
                ws1.cell(r, 9, f"=G{r}-H{r}")
                break

            po_cell = ws1.cell(r, 3).value
            if po_cell and str(po_cell).strip():
                curr_po = str(po_cell).strip().upper()

            act_cell = ws1.cell(r, 6).value
            if act_cell and str(act_cell).strip():
                norm_act = normalize_activity(act_cell)
                key = (curr_po, norm_act)

                if key in card_links:
                    card_sheet, card_row = card_links[key]
                    # Link Spent Budget to card's TOTAL IV (Col S)
                    ws1.cell(r, 8, f"='{card_sheet}'!S{card_row}").number_format = '#,##0.00'
                else:
                    ws1.cell(r, 8, 0).number_format = '#,##0.00'

                # Balance = Budget - Spent Budget
                ws1.cell(r, 9, f"=G{r}-H{r}").number_format = '#,##0.00'

    return updated_cards_count

def sync_corteva_cards_workbook(wb_cards, tbm_data, service_charge_percent=5.0):
    """
    Synchronizes Corteva PO Summary Cards (where each product is on its own sheet).
    """
    sv_rate = float(service_charge_percent) / 100.0
    regular_font = Font(name='Calibri', size=10)
    bold_font = Font(name='Calibri', size=10, bold=True)
    red_bold_font = Font(name='Calibri', size=10, bold=True, color='FF0000')

    updated_cards_count = 0

    first_sheet_name = wb_cards.sheetnames[0]
    for name in wb_cards.sheetnames:
        if name == first_sheet_name or name.strip().lower() in ['sheet1', 'processed_emails']:
            continue

        ws = wb_cards[name]
        po_val = str(ws.cell(6, 1).value or '').strip().upper()
        if not po_val:
            continue

        # Find activities from rate grid (Cols T & U, rows 3..)
        act_map = {}
        for r_idx in range(3, 10):
            act_name = ws.cell(r_idx, 20).value
            if act_name and str(act_name).strip() not in ['TOTAL', '']:
                act_map[normalize_activity(act_name)] = 12 + len(act_map)

        if not act_map:
            continue

        n_act = len(act_map)
        amount_col = 12 + n_act
        tbm_rows_for_po = tbm_data.get(po_val, [])
        if tbm_rows_for_po:
            updated_cards_count += 1

        # Clear data rows 12 to 28
        for r in range(12, 29):
            for c in range(1, amount_col + 1):
                ws.cell(r, c).value = None

        area_val = str(ws.cell(7, 4).value or '').split('-')[-1].strip()

        for idx, t_item in enumerate(tbm_rows_for_po):
            if idx >= 17:
                break
            r = 12 + idx

            ws.cell(r, 1).value = None  # I.V NO
            ws.cell(r, 2).value = None  # DATE
            ws.cell(r, 3, area_val).font = regular_font
            ws.cell(r, 4, po_val).font = regular_font
            ws.cell(r, 5, "Marketing").font = regular_font
            ws.cell(r, 6, t_item['product'] or name).font = regular_font
            ws.cell(r, 7, t_item['crop'] or "All Crops").font = regular_font
            ws.cell(r, 8, t_item['activity']).font = regular_font
            ws.cell(r, 9, "ZDGM").font = regular_font
            ws.cell(r, 10, t_item['tbm']).font = regular_font
            ws.cell(r, 11, t_item['num_activities']).font = regular_font

            norm_act = normalize_activity(t_item['activity'])
            target_col = act_map.get(norm_act, 12)

            for c in range(12, amount_col):
                if c == target_col:
                    ws.cell(r, c, t_item['total_amount']).font = regular_font
                    ws.cell(r, c).number_format = '#,##0'
                else:
                    ws.cell(r, c, None)

            parts = [f"{get_column_letter(c)}{r}" for c in range(12, amount_col)]
            ws.cell(r, amount_col, f"=SUM({','.join(parts)})").font = regular_font
            ws.cell(r, amount_col).number_format = '#,##0'

        # Bottom Totals row 29
        for i in range(n_act):
            col = 12 + i
            cl = get_column_letter(col)
            cell = ws.cell(29, col, f"=SUM({cl}12:{cl}28)")
            cell.font = red_bold_font
            cell.number_format = '#,##0'

        amt_cl = get_column_letter(amount_col)
        cell = ws.cell(29, amount_col, f"=SUM({amt_cl}12:{amt_cl}28)")
        cell.font = red_bold_font
        cell.number_format = '#,##0'

        # Right Summary block T16:Y
        for i in range(n_act):
            r = 16 + i
            cl = get_column_letter(12 + i)
            rate_r = 3 + i

            # SPENT in Col V (22)
            ws.cell(r, 22, f"={cl}29").font = regular_font
            ws.cell(r, 22).number_format = '#,##0'

            # SV Charges in Col W (23)
            ws.cell(r, 23, f"=V{r}*{sv_rate}").font = regular_font
            ws.cell(r, 23).number_format = '#,##0.00'

            # TOTAL IV in Col X (24)
            ws.cell(r, 24, f"=V{r}+W{r}").font = regular_font
            ws.cell(r, 24).number_format = '#,##0.00'

            # BALANCE in Col Y (25)
            ws.cell(r, 25, f"=U{r}-X{r}").font = regular_font
            ws.cell(r, 25).number_format = '#,##0.00'

        tot_row = 16 + n_act
        for col_idx in [21, 22, 23, 24, 25]:
            cl = get_column_letter(col_idx)
            cell = ws.cell(tot_row, col_idx, f"=SUM({cl}16:{cl}{tot_row-1})")
            cell.font = red_bold_font
            cell.number_format = '#,##0.00' if col_idx in [23, 24, 25] else '#,##0'

    return updated_cards_count

def sync_tbm_with_cards(cards_excel_path, tbm_summary_excel_path, output_path=None, service_charge_percent=5.0):
    """
    Main entry point to synchronize PO Cards summary with Consolidated TBM Summary.
    Applies service charges %, links formulas, updates balance columns in cards and Sheet1.
    """
    cards_file = Path(cards_excel_path).resolve()
    tbm_file = Path(tbm_summary_excel_path).resolve()

    if not cards_file.exists():
        raise FileNotFoundError(f"Cards Summary file not found at: {cards_excel_path}")
    if not tbm_file.exists():
        raise FileNotFoundError(f"TBM Summary file not found at: {tbm_summary_excel_path}")

    # 1. Extract TBM data
    tbm_data = extract_tbm_spent_summary(tbm_file)

    # 2. Load Cards Workbook
    wb_cards = load_any_workbook(cards_file)

    # Determine if FMC (stacked 19-row cards) or Corteva (product sheets)
    is_fmc_structure = False
    for sname in wb_cards.sheetnames:
        if sname != 'Sheet1':
            ws = wb_cards[sname]
            if ws.cell(1, 1).value and str(ws.cell(1, 1).value).strip().startswith('500'):
                is_fmc_structure = True
                break

    if is_fmc_structure:
        updated_count = sync_fmc_cards_workbook(wb_cards, tbm_data, service_charge_percent)
    else:
        updated_count = sync_corteva_cards_workbook(wb_cards, tbm_data, service_charge_percent)

    save_target = Path(output_path).resolve() if output_path else cards_file
    wb_cards.save(save_target)
    wb_cards.close()

    return {
        "success": True,
        "message": f"Successfully synchronized {updated_count} PO card(s) with TBM summary at {service_charge_percent}% service charge!",
        "updatedCards": updated_count,
        "totalTbmPOs": len(tbm_data),
        "outputPath": str(save_target),
        "serviceChargePercent": service_charge_percent
    }

if __name__ == "__main__":
    if len(sys.argv) > 2:
        c_path = sys.argv[1]
        t_path = sys.argv[2]
        sv_p = float(sys.argv[3]) if len(sys.argv) > 3 else 5.0
        res = sync_tbm_with_cards(c_path, t_path, service_charge_percent=sv_p)
        print(res["message"])
    else:
        print("Usage: python card_sync_engine.py <cards_excel_path> <tbm_summary_excel_path> [service_charge_percent]")
