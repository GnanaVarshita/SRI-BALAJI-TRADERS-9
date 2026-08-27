import openpyxl
from pathlib import Path

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    import pyxlsb
except ImportError:
    pyxlsb = None

def load_any_workbook(filepath):
    """
    Universal workbook loader accepting .xlsx, .xlsm, .xls, .xlsb, and .csv files.
    Returns an openpyxl Workbook object.
    """
    p = Path(filepath)
    ext = p.suffix.lower()

    if ext in ['.xlsx', '.xlsm']:
        return openpyxl.load_workbook(p, data_only=True)

    wb_out = openpyxl.Workbook()
    if "Sheet" in wb_out.sheetnames:
        wb_out.remove(wb_out["Sheet"])

    if ext == '.xls' and xlrd is not None:
        wb_xls = xlrd.open_workbook(p)
        for sname in wb_xls.sheet_names():
            sh_xls = wb_xls.sheet_by_name(sname)
            ws_out = wb_out.create_sheet(title=sname)
            for r in range(sh_xls.nrows):
                for c in range(sh_xls.ncols):
                    val = sh_xls.cell_value(r, c)
                    if sh_xls.cell_type(r, c) == xlrd.XL_CELL_DATE:
                        try:
                            dt = xlrd.xldate_as_datetime(val, wb_xls.datemode)
                            val = dt.strftime('%d/%m/%Y')
                        except Exception:
                            pass
                    ws_out.cell(row=r+1, column=c+1, value=val)
        return wb_out

    elif ext == '.csv':
        import csv
        with open(p, 'r', encoding='utf-8-sig', errors='replace') as f:
            reader = list(csv.reader(f))
            ws_out = wb_out.create_sheet(title="Sheet1")
            for r_idx, row in enumerate(reader, start=1):
                for c_idx, val in enumerate(row, start=1):
                    ws_out.cell(row=r_idx, column=c_idx, value=val)
        return wb_out

    elif ext == '.xlsb' and pyxlsb is not None:
        with pyxlsb.open_workbook(p) as wb_xlsb:
            for sname in wb_xlsb.sheets:
                ws_out = wb_out.create_sheet(title=sname)
                with wb_xlsb.get_sheet(sname) as sheet:
                    for r_idx, row in enumerate(sheet.rows(), start=1):
                        for c_idx, cell in enumerate(row, start=1):
                            ws_out.cell(row=r_idx, column=c_idx, value=cell.v)
        return wb_out

    # Fallback try openpyxl then xlrd
    try:
        return openpyxl.load_workbook(p, data_only=True)
    except Exception:
        if xlrd is not None:
            try:
                wb_xls = xlrd.open_workbook(p)
                for sname in wb_xls.sheet_names():
                    sh_xls = wb_xls.sheet_by_name(sname)
                    ws_out = wb_out.create_sheet(title=sname)
                    for r in range(sh_xls.nrows):
                        for c in range(sh_xls.ncols):
                            ws_out.cell(row=r+1, column=c+1, value=sh_xls.cell_value(r, c))
                return wb_out
            except Exception:
                pass
        raise ValueError(f"Unsupported or corrupted Excel format for file: {p.name}")

def find_headers(sheet):
    """
    Scans the first 20 rows of the sheet to dynamically identify the header row.
    Matches keywords like type, product, crop, activity, event cost, budget.
    """
    for r in range(1, 21):
        row_vals = [str(sheet.cell(row=r, column=col).value).strip() for col in range(1, 15)]
        indices = {}
        for c_idx, val in enumerate(row_vals, start=1):
            if val is None or val == "None":
                continue
            val_lower = val.lower()
            if "type" in val_lower and "activity" not in val_lower:
                indices["type"] = c_idx
            elif "product" in val_lower:
                indices["product"] = c_idx
            elif "crop" in val_lower:
                indices["crop"] = c_idx
            elif "type of activity" in val_lower or ("activity" in val_lower and "budget" not in val_lower):
                indices["activity"] = c_idx
            elif "event cost" in val_lower or "cost" in val_lower:
                indices["event_cost"] = c_idx
            elif "budget for" in val_lower or "marketing budget" in val_lower or "dg activities" in val_lower:
                indices["qty"] = c_idx
            elif "total budget" in val_lower or "allocation" in val_lower:
                indices["allocation"] = c_idx
        
        if len(indices) >= 4:
            default_keys = ["type", "product", "crop", "activity", "event_cost", "qty", "allocation"]
            standard_cols = {
                "type": 1,
                "product": 2,
                "crop": 3,
                "activity": 4,
                "event_cost": 5,
                "qty": 6,
                "allocation": 7
            }
            for k in default_keys:
                if k not in indices:
                    indices[k] = standard_cols[k]
            return r, indices
            
    return 11, {
        "type": 1,
        "product": 2,
        "crop": 3,
        "activity": 4,
        "event_cost": 5,
        "qty": 6,
        "allocation": 7
    }

def validate_budget_sheet(filepath):
    """
    Opens any format Excel sheet (.xlsx, .xls, .xlsm, .xlsb, .csv),
    dynamically locates data headers, parses data rows, and returns a report.
    """
    try:
        wb = load_any_workbook(filepath)
        sheet = wb.active
        
        header_row, col_map = find_headers(sheet)
        errors = []
        rows_data = []
        
        calculated_total_qty = 0
        calculated_total_budget = 0
        
        row_idx = header_row + 1
        while True:
            row_values = [sheet.cell(row=row_idx, column=col).value for col in range(1, 10)]
            if all(v is None or str(v).strip() == "" or str(v).strip().lower() == "none" for v in row_values):
                break
                
            row_type = sheet.cell(row=row_idx, column=col_map["type"]).value
            product = sheet.cell(row=row_idx, column=col_map["product"]).value
            crop = sheet.cell(row=row_idx, column=col_map["crop"]).value
            activity = sheet.cell(row=row_idx, column=col_map["activity"]).value
            
            if product is None and row_type is None:
                break
                
            try:
                event_cost = float(sheet.cell(row=row_idx, column=col_map["event_cost"]).value or 0.0)
                qty = float(sheet.cell(row=row_idx, column=col_map["qty"]).value or 0.0)
                
                cell_alloc = sheet.cell(row=row_idx, column=col_map["allocation"]).value
                if cell_alloc is None:
                    allocation = event_cost * qty
                else:
                    allocation = float(cell_alloc)
            except (ValueError, TypeError):
                errors.append(f"Row {row_idx}: Non-numeric value in budget columns")
                row_idx += 1
                continue

            calculated_total_qty += qty
            calculated_total_budget += allocation
            
            rows_data.append({
                "row": row_idx,
                "type": row_type,
                "product": product,
                "crop": crop,
                "activity": activity,
                "event_cost": event_cost,
                "qty": qty,
                "allocation": allocation
            })
            row_idx += 1

        wb.close()
        return {
            "success": len(errors) == 0,
            "errors": errors,
            "totals": {
                "totalQty": calculated_total_qty,
                "totalBudget": calculated_total_budget
            },
            "rows": rows_data
        }
    except Exception as e:
        return {
            "success": False,
            "errors": [f"Failed to parse Excel file: {e}"],
            "totals": {"totalQty": 0, "totalBudget": 0},
            "rows": []
        }
