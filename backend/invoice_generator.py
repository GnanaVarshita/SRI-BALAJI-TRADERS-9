import os
import re
import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.pagebreak import Break
from openpyxl.utils import get_column_letter

import sys
sys.path.append(str(Path(__file__).resolve().parent))
try:
    from excel_parser import load_any_workbook
except ImportError:
    from .excel_parser import load_any_workbook

# --- Number to Indian Currency Words Converter ---
def num_to_indian_words(amount):
    """
    Converts a number (float or int) to Indian Currency Words format matching exact screenshot style:
    e.g. 126843 -> "One Lakh Twenty Six Thousand Eight Hundred and Forty Three Rupees Only"
    e.g. 15488 -> "Fifteen Thousand Four Hundred and Eighty Eight Rupees Only"
    e.g. 8989 -> "Eight Thousand Nine Hundred and Eighty Nine Rupees Only"
    """
    try:
        amt_float = float(amount)
    except (ValueError, TypeError):
        return ""

    if amt_float == 0:
        return "Zero Rupees Only"

    is_negative = amt_float < 0
    amt_float = abs(amt_float)

    rupees = int(amt_float)
    paise = int(round((amt_float - rupees) * 100))

    units = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
             "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen",
             "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]

    def two_digits_to_words(n):
        if n == 0:
            return ""
        elif n < 20:
            return units[n]
        else:
            t = tens[n // 10]
            u = units[n % 10]
            return f"{t} {u}".strip()

    def three_digits_to_words(n):
        h = n // 100
        rem = n % 100
        words = []
        if h > 0:
            words.append(f"{units[h]} Hundred")
        if rem > 0:
            if h > 0:
                words.append(f"and {two_digits_to_words(rem)}")
            else:
                words.append(two_digits_to_words(rem))
        return " ".join(words).strip()

    parts = []
    # Crores (10,000,000)
    crores = rupees // 10000000
    rupees %= 10000000
    if crores > 0:
        parts.append(f"{two_digits_to_words(crores)} Crore")

    # Lakhs (100,000)
    lakhs = rupees // 100000
    rupees %= 100000
    if lakhs > 0:
        parts.append(f"{two_digits_to_words(lakhs)} Lakh")

    # Thousands (1,000)
    thousands = rupees // 1000
    rupees %= 1000
    if thousands > 0:
        parts.append(f"{two_digits_to_words(thousands)} Thousand")

    # Hundreds and below
    if rupees > 0:
        parts.append(three_digits_to_words(rupees))

    rupees_str = " ".join(parts).strip()
    if not rupees_str:
        rupees_str = "Zero"

    res = f"{rupees_str} Rupees"
    if paise > 0:
        res += f" and {two_digits_to_words(paise)} Paise"
    res += " Only"

    if is_negative:
        res = "Minus " + res

    return res

def clean_str(val):
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in ["none", "null", "nan"]:
        return ""
    return s

def parse_num(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0

def format_date_val(val):
    if val is None:
        return ""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%d-%m-%Y")
    s = str(val).strip()
    if not s:
        return ""
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d\\%m\\%Y", "%d/%m/%y", "%d-%m-%y"]:
        try:
            dt = datetime.datetime.strptime(s, fmt)
            return dt.strftime("%d-%m-%Y")
        except ValueError:
            pass
    return s

def normalize_po(po_str):
    if not po_str:
        return ""
    return re.sub(r'[\s\-_]', '', str(po_str)).upper()

def format_short_iv(invoice_num_str):
    """
    Extracts short display number: e.g. "SBT26270072" -> "72" or "SBT26270067" -> "67"
    """
    s = str(invoice_num_str).strip()
    m = re.search(r'SBT\d{4}(\d+)', s, re.I)
    if m:
        return str(int(m.group(1)))
    m2 = re.search(r'(\d+)$', s)
    if m2:
        return str(int(m2.group(1)))
    return s

def format_full_iv(invoice_num_str, prefix="SBT2627"):
    """
    Formats standard full invoice number: e.g. "72" -> "SBT26270072"
    """
    s = str(invoice_num_str).strip()
    if s.upper().startswith("SBT"):
        return s.upper()
    m = re.search(r'(\d+)$', s)
    if m:
        num = int(m.group(1))
        return f"{prefix}{num:04d}"
    return s

def scan_pos_in_summary(tbm_summary_path):
    """
    Scans the All-TBMs-Summary.xlsx workbook and returns list of unique PO numbers found.
    Handles any length of PO number without trimming.
    """
    tbm_file = Path(tbm_summary_path).resolve()
    if not tbm_file.exists():
        return []

    wb = load_any_workbook(tbm_file)
    pos_found = set()

    for sname in wb.sheetnames:
        ws = wb[sname]
        for r in range(1, ws.max_row + 1):
            for c in range(1, min(ws.max_column + 1, 35)):
                val = clean_str(ws.cell(r, c).value)
                if val:
                    # Corteva PO: starts with 48 (e.g. 4800108503)
                    m_corteva = re.findall(r'\b(48\d{6,30})\b', val)
                    # FMC PO: starts with 5 (e.g. 500BB20260710172, 500BB2026018404)
                    m_fmc = re.findall(r'\b(5[A-Za-z0-9]{8,35})\b', val, re.I)
                    for m in m_corteva + m_fmc:
                        pos_found.add(m.upper())
                    
                    # Direct check on Column 17 or any PO-labeled column
                    if c == 17:
                        val_u = val.upper()
                        if val_u and val_u not in ["PO", "PO NUMBER", "PO NO", "P.O", "P.O.", "NONE", "TOTAL"]:
                            if re.match(r'^[A-Z0-9\-_]{6,35}$', val_u):
                                pos_found.add(val_u)
    wb.close()
    return sorted(list(pos_found))

def extract_tables_for_po(tbm_summary_path, target_po):
    """
    Extracts all activity expense records matching target_po from All-TBMs-Summary.xlsx.
    Performs strict exact PO matching on each row to prevent mixing activities from different POs.
    Returns: (records, metadata)
    """
    tbm_file = Path(tbm_summary_path).resolve()
    if not tbm_file.exists():
        raise FileNotFoundError(f"TBM Summary file not found at: {tbm_summary_path}")

    wb = load_any_workbook(tbm_file)
    norm_target_po = normalize_po(target_po)

    all_matched_records = []
    overall_metadata = {
        'product': '',
        'crop': '',
        'territory': '',
        'area': '',
        'zdgm': '',
        'amm': ''
    }

    for sname in wb.sheetnames:
        if sname.strip().lower() in ['tbm amount summary', 'amount summary']:
            continue

        ws = wb[sname]
        max_r = ws.max_row
        r = 1
        while r <= max_r:
            cell_val = clean_str(ws.cell(r, 1).value)
            if "ACTIVITIES EXPENSES" in cell_val.upper():
                hdr_row_idx = r + 1
                headers = [clean_str(ws.cell(hdr_row_idx, c).value).lower() for c in range(1, 18)]
                if not any('si' in h or 's.no' in h or 'sl' in h or 'date' in h for h in headers):
                    r += 1
                    continue

                data_r = hdr_row_idx + 1
                table_rows = []
                while data_r <= max_r:
                    c1_val = clean_str(ws.cell(data_r, 1).value).lower()
                    c15_val = clean_str(ws.cell(data_r, 15).value).lower()
                    
                    if c1_val == 'total' or c15_val == 'total' or 'total expenses' in c1_val or 'grand total' in c1_val:
                        break
                    if "ACTIVITIES EXPENSES" in c1_val.upper():
                        break

                    row_vals = [ws.cell(data_r, c).value for c in range(1, 18)]
                    if not any(row_vals):
                        data_r += 1
                        continue

                    po_in_row = clean_str(ws.cell(data_r, 17).value)
                    norm_row_po = normalize_po(po_in_row)

                    date_val = format_date_val(ws.cell(data_r, 2).value)
                    zdgm_val = clean_str(ws.cell(data_r, 3).value)
                    tbm_val = clean_str(ws.cell(data_r, 4).value)
                    mdo_val = clean_str(ws.cell(data_r, 5).value)
                    terr_val = clean_str(ws.cell(data_r, 6).value)
                    prod_val = clean_str(ws.cell(data_r, 7).value)
                    crop_val = clean_str(ws.cell(data_r, 8).value)
                    act_val = clean_str(ws.cell(data_r, 9).value)
                    vlg_val = clean_str(ws.cell(data_r, 10).value)
                    farm_val = int(parse_num(ws.cell(data_r, 11).value))
                    tent_val = parse_num(ws.cell(data_r, 12).value)
                    food_val = parse_num(ws.cell(data_r, 13).value)
                    trans_val = parse_num(ws.cell(data_r, 14).value)
                    oth_val = parse_num(ws.cell(data_r, 15).value)
                    tot_val = parse_num(ws.cell(data_r, 16).value)
                    calc_tot = tent_val + food_val + trans_val + oth_val
                    if tot_val == 0.0 and calc_tot > 0.0:
                        tot_val = calc_tot

                    is_match = False
                    if norm_target_po:
                        if norm_row_po:
                            # Strict exact match - prevents mixing activities belonging to different POs
                            if norm_row_po == norm_target_po:
                                is_match = True
                        else:
                            # If row cell is empty, check if the table header explicitly specifies this PO
                            tbl_header_norm = normalize_po(cell_val)
                            if norm_target_po in tbl_header_norm:
                                is_match = True
                    else:
                        is_match = True

                    if is_match and (act_val or prod_val or tot_val > 0):
                        rec = {
                            'date': date_val,
                            'zdgm': zdgm_val,
                            'tbm': tbm_val,
                            'mdo': mdo_val,
                            'territory': terr_val,
                            'product': prod_val,
                            'crop': crop_val,
                            'activity': act_val,
                            'village': vlg_val,
                            'farmers': farm_val,
                            'tent': tent_val,
                            'food': food_val,
                            'transport': trans_val,
                            'others': oth_val,
                            'total': tot_val,
                            'po_number': po_in_row or target_po
                        }
                        table_rows.append(rec)
                        
                        if prod_val and not overall_metadata['product']: overall_metadata['product'] = prod_val
                        if crop_val and not overall_metadata['crop']: overall_metadata['crop'] = crop_val
                        if terr_val and not overall_metadata['territory']: overall_metadata['territory'] = terr_val
                        if zdgm_val and not overall_metadata['zdgm']: overall_metadata['zdgm'] = zdgm_val
                        if zdgm_val and not overall_metadata['amm']: overall_metadata['amm'] = zdgm_val

                    data_r += 1

                if table_rows:
                    all_matched_records.extend(table_rows)

                r = data_r
            else:
                r += 1

    wb.close()
    return all_matched_records, overall_metadata

def get_base_styles():
    thin = Side(style='thin', color='A6A6A6')
    black_thin = Side(style='thin', color='000000')
    black_double = Side(style='double', color='000000')
    black_medium = Side(style='medium', color='000000')

    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    box_border = Border(left=black_thin, right=black_thin, top=black_thin, bottom=black_thin)
    total_border = Border(left=black_thin, right=black_thin, top=black_thin, bottom=black_double)
    header_border = Border(left=black_thin, right=black_thin, top=black_medium, bottom=black_medium)

    # Soft SBT background fill
    sbt_fill = PatternFill(start_color='F2EFE9', end_color='F2EFE9', fill_type='solid')

    return {
        'thin_border': thin_border,
        'box_border': box_border,
        'total_border': total_border,
        'header_border': header_border,
        'sbt_fill': sbt_fill,
        'font_title': Font(name='Calibri', size=11, bold=True),
        'font_header': Font(name='Calibri', size=10, bold=True),
        'font_bold': Font(name='Calibri', size=10, bold=True),
        'font_bold_u': Font(name='Calibri', size=10, bold=True, underline='single'),
        'font_regular': Font(name='Calibri', size=10),
        'font_italic': Font(name='Calibri', size=10, italic=True),
        'font_green_title': Font(name='Calibri', size=11, bold=True, color='008000'),
        'font_green_bold': Font(name='Calibri', size=10, bold=True, color='008000'),
        'font_blue_bold': Font(name='Calibri', size=10, bold=True, color='0000FF'),
        'font_sbt_logo': Font(name='Times New Roman', size=26, bold=True),
        'font_sbt_sub': Font(name='Calibri', size=8, bold=True),
        'align_center': Alignment(horizontal='center', vertical='center'),
        'align_center_wrap': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'align_left': Alignment(horizontal='left', vertical='center'),
        'align_right': Alignment(horizontal='right', vertical='center'),
    }

def setup_page_print_fit(ws, print_area=None, orientation="portrait"):
    """
    Configures worksheet page setup for clean A4 printing with balanced side margins.
    """
    if orientation == "landscape":
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    else:
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if print_area:
        ws.print_area = print_area

    # Center horizontally and set balanced side margins for clean spacing
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35

def write_sbt_header_block(ws, start_row, copy_type, styles):
    """
    Writes the standard Sri Balaji Traders top branding block starting at start_row across Columns A to J.
    """
    r1 = start_row
    r2 = start_row + 1
    r3 = start_row + 2
    r4 = start_row + 3
    r5 = start_row + 4
    r6 = start_row + 5
    r7 = start_row + 6
    r8 = start_row + 7

    # Left company info (Cols A-C)
    ws[f'A{r1}'] = "SRI BALAJI TRADERS"
    ws[f'A{r1}'].font = styles['font_title']
    ws[f'A{r2}'] = "#5/387, Gandhi Road, Proddatur,"
    ws[f'A{r2}'].font = styles['font_regular']
    ws[f'A{r3}'] = "YSR District, (Kadapa)"
    ws[f'A{r3}'].font = styles['font_regular']
    ws[f'A{r4}'] = "Andhra Pradesh, 516360"
    ws[f'A{r4}'].font = styles['font_regular']
    ws[f'A{r5}'] = "GSTIN:37BXDPK0359K1ZA"
    ws[f'A{r5}'].font = styles['font_bold']

    # Center SBT logo block (Cols D-G) with soft fill
    ws.merge_cells(f'D{r1}:G{r4}')
    c_logo = ws[f'D{r1}']
    c_logo.value = "SBT"
    c_logo.font = styles['font_sbt_logo']
    c_logo.alignment = styles['align_center']
    
    ws.merge_cells(f'D{r5}:G{r5}')
    c_logosub = ws[f'D{r5}']
    c_logosub.value = "SRI BALAJI TRADERS"
    c_logosub.font = styles['font_sbt_sub']
    c_logosub.alignment = styles['align_center']

    for r_sbt in range(r1, r5 + 1):
        for c_sbt in range(4, 8):
            ws.cell(r_sbt, c_sbt).fill = styles['sbt_fill']

    # Right Contact Info (Cols I-J)
    ws.merge_cells(f'I{r1}:J{r1}')
    ws[f'I{r1}'] = "Prop: K Radha Devi"
    ws[f'I{r1}'].font = styles['font_bold']
    ws[f'I{r1}'].alignment = styles['align_right']

    ws.merge_cells(f'I{r2}:J{r2}')
    ws[f'I{r2}'] = "Cell: 8328588119"
    ws[f'I{r2}'].font = styles['font_bold']
    ws[f'I{r2}'].alignment = styles['align_right']

    ws.merge_cells(f'I{r3}:J{r3}')
    ws[f'I{r3}'] = "9000491388"
    ws[f'I{r3}'].font = styles['font_bold']
    ws[f'I{r3}'].alignment = styles['align_right']

    # Row 5 solid bottom separator across Columns A to J (Cols 1 to 10)
    for col in range(1, 11):
        cell = ws.cell(r5, col)
        cell.border = Border(bottom=Side(style='medium', color='000000'))

    for r in range(r1, r5 + 1):
        ws.row_dimensions[r].height = 15

    # Row 6 spacer
    ws.row_dimensions[r6].height = 6

    # Titles (Row 7): TAX INVOICE & ORIGINAL / DUPLICATE
    ws.row_dimensions[r7].height = 18
    ws.merge_cells(f'E{r7}:G{r7}')
    ws[f'E{r7}'] = "TAX INVOICE"
    ws[f'E{r7}'].font = Font(name='Calibri', size=11, bold=True, underline='single')
    ws[f'E{r7}'].alignment = styles['align_center']

    ws.merge_cells(f'I{r7}:J{r7}')
    ws[f'I{r7}'] = copy_type
    ws[f'I{r7}'].font = Font(name='Calibri', size=11, bold=True, underline='single')
    ws[f'I{r7}'].alignment = styles['align_center']

    # Row 8 spacer
    ws.row_dimensions[r8].height = 6

def write_bank_and_signature_block(ws, start_row, styles):
    """
    Writes the Bank Details and Authorised Signatory block starting directly at start_row across Cols A to J.
    """
    r_bank = start_row
    r_bnk_name = start_row + 1
    r_acc = start_row + 2
    r_ifsc = start_row + 3
    r_branch = start_row + 4
    r_end = start_row + 5

    # Row 1: Bank details:
    ws[f'A{r_bank}'] = "Bank details:"
    ws[f'A{r_bank}'].font = styles['font_bold_u']
    ws.row_dimensions[r_bank].height = 15

    # Row 2: Bank name: Karnataka Bank & For Sri Balaji Traders
    ws[f'A{r_bnk_name}'] = "Bank name: Karnataka Bank"
    ws[f'A{r_bnk_name}'].font = styles['font_bold']
    ws.merge_cells(f'G{r_bnk_name}:J{r_bnk_name}')
    ws[f'G{r_bnk_name}'] = "For Sri Balaji Traders"
    ws[f'G{r_bnk_name}'].font = styles['font_bold']
    ws[f'G{r_bnk_name}'].alignment = styles['align_center']
    ws.row_dimensions[r_bnk_name].height = 15

    # Row 3: A/c No
    ws[f'A{r_acc}'] = "A/c No"
    ws[f'A{r_acc}'].font = styles['font_bold']
    ws.merge_cells(f'B{r_acc}:E{r_acc}')
    ws[f'B{r_acc}'] = ":6187000600001901"
    ws[f'B{r_acc}'].font = styles['font_bold']
    ws.row_dimensions[r_acc].height = 15

    # Row 4: IFSC
    ws[f'A{r_ifsc}'] = "IFSC"
    ws[f'A{r_ifsc}'].font = styles['font_bold']
    ws.merge_cells(f'B{r_ifsc}:E{r_ifsc}')
    ws[f'B{r_ifsc}'] = ":KARB0000618"
    ws[f'B{r_ifsc}'].font = styles['font_bold']
    ws.row_dimensions[r_ifsc].height = 15

    # Row 5: Branch & Authorised Signatory
    ws[f'A{r_branch}'] = "Branch"
    ws[f'A{r_branch}'].font = styles['font_bold']
    ws.merge_cells(f'B{r_branch}:E{r_branch}')
    ws[f'B{r_branch}'] = ":Proddatur"
    ws[f'B{r_branch}'].font = styles['font_bold']
    
    ws.merge_cells(f'G{r_branch}:J{r_branch}')
    ws[f'G{r_branch}'] = "Authorised Signatory"
    ws[f'G{r_branch}'].font = styles['font_bold']
    ws[f'G{r_branch}'].alignment = styles['align_center']
    ws.row_dimensions[r_branch].height = 18

    # Row 6: bottom padding
    ws.row_dimensions[r_end].height = 12

    # Outer border on bank & signature section
    for r in range(r_bank, r_end + 1):
        ws.cell(r, 1).border = Border(left=Side(style='thin', color='000000'))
        ws.cell(r, 10).border = Border(right=Side(style='thin', color='000000'))
    
    # Close bottom border on Row r_end across A:J
    for c in range(1, 11):
        ws.cell(r_end, c).border = Border(
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000') if c == 1 else None,
            right=Side(style='thin', color='000000') if c == 10 else None
        )
    return r_end

def render_corteva_invoice_block(ws, start_row, copy_type, invoice_no, invoice_date, po_number, metadata, activity_groups, service_charge_pct, styles):
    """
    Renders one full Corteva Tax Invoice block (ORIGINAL or DUPLICATE) starting at start_row.
    Returns: (end_row, r_subtotal, r_grand)
    """
    write_sbt_header_block(ws, start_row, copy_type, styles)

    # Customer & Metadata block (Rows start_row+8 to start_row+16)
    r_cust_start = start_row + 8
    for r_c in range(r_cust_start, r_cust_start + 9):
        ws.row_dimensions[r_c].height = 14

    ws[f'A{r_cust_start}'] = "TO:"
    ws[f'A{r_cust_start}'].font = styles['font_bold']
    ws[f'A{r_cust_start + 1}'] = "Corteva Agriscience India Private Limited"
    ws[f'A{r_cust_start + 1}'].font = styles['font_bold']
    ws[f'A{r_cust_start + 2}'] = "D.NO 12-112 Godown NO.2 Krishna Nagar"
    ws[f'A{r_cust_start + 2}'].font = styles['font_regular']
    ws[f'A{r_cust_start + 3}'] = "Vijayawada"
    ws[f'A{r_cust_start + 3}'].font = styles['font_regular']
    ws[f'A{r_cust_start + 4}'] = "01 520007"
    ws[f'A{r_cust_start + 4}'].font = styles['font_regular']
    ws[f'A{r_cust_start + 5}'] = "INDIA"
    ws[f'A{r_cust_start + 5}'].font = styles['font_regular']
    ws[f'A{r_cust_start + 6}'] = "GSTIN: 37AAACE2462M1ZI"
    ws[f'A{r_cust_start + 6}'].font = styles['font_bold']

    # Right metadata
    ws.merge_cells(f'I{r_cust_start}:J{r_cust_start}')
    ws[f'I{r_cust_start}'] = "MSME"
    ws[f'I{r_cust_start}'].font = styles['font_bold']
    ws[f'I{r_cust_start}'].alignment = styles['align_center']

    ws[f'G{r_cust_start + 1}'] = "Invoive no:"
    ws[f'G{r_cust_start + 1}'].font = styles['font_bold']
    ws.merge_cells(f'I{r_cust_start + 1}:J{r_cust_start + 1}')
    ws[f'I{r_cust_start + 1}'] = invoice_no
    ws[f'I{r_cust_start + 1}'].font = styles['font_bold']
    ws[f'I{r_cust_start + 1}'].alignment = styles['align_center']

    ws[f'G{r_cust_start + 2}'] = "Invoice Date:"
    ws[f'G{r_cust_start + 2}'].font = styles['font_bold']
    ws.merge_cells(f'I{r_cust_start + 2}:J{r_cust_start + 2}')
    ws[f'I{r_cust_start + 2}'] = invoice_date
    ws[f'I{r_cust_start + 2}'].font = styles['font_bold']
    ws[f'I{r_cust_start + 2}'].alignment = styles['align_center']

    ws[f'G{r_cust_start + 3}'] = "Place Of Supply:"
    ws[f'G{r_cust_start + 3}'].font = styles['font_bold']
    ws.merge_cells(f'I{r_cust_start + 3}:J{r_cust_start + 3}')
    ws[f'I{r_cust_start + 3}'] = "Andhra Pradesh"
    ws[f'I{r_cust_start + 3}'].font = styles['font_bold']
    ws[f'I{r_cust_start + 3}'].alignment = styles['align_center']

    ws[f'G{r_cust_start + 4}'] = "Corteva PO No :"
    ws[f'G{r_cust_start + 4}'].font = styles['font_bold']
    ws.merge_cells(f'I{r_cust_start + 4}:J{r_cust_start + 4}')
    ws[f'I{r_cust_start + 4}'] = po_number
    ws[f'I{r_cust_start + 4}'].font = styles['font_green_bold']
    ws[f'I{r_cust_start + 4}'].alignment = styles['align_center']

    ws[f'G{r_cust_start + 5}'] = "Product:"
    ws[f'G{r_cust_start + 5}'].font = styles['font_bold']
    ws.merge_cells(f'I{r_cust_start + 5}:J{r_cust_start + 5}')
    ws[f'I{r_cust_start + 5}'] = metadata.get('product', '')
    ws[f'I{r_cust_start + 5}'].font = styles['font_bold']
    ws[f'I{r_cust_start + 5}'].alignment = styles['align_center']

    ws[f'G{r_cust_start + 6}'] = "Crop:"
    ws[f'G{r_cust_start + 6}'].font = styles['font_bold']
    ws.merge_cells(f'I{r_cust_start + 6}:J{r_cust_start + 6}')
    ws[f'I{r_cust_start + 6}'] = metadata.get('crop', '')
    ws[f'I{r_cust_start + 6}'].font = styles['font_bold']
    ws[f'I{r_cust_start + 6}'].alignment = styles['align_center']

    # Row AREA
    ws[f'G{r_cust_start + 7}'] = "AREA"
    ws[f'G{r_cust_start + 7}'].font = styles['font_bold']
    ws.merge_cells(f'I{r_cust_start + 7}:J{r_cust_start + 7}')
    ws[f'I{r_cust_start + 7}'] = metadata.get('area') or metadata.get('territory', '')
    ws[f'I{r_cust_start + 7}'].font = styles['font_bold']
    ws[f'I{r_cust_start + 7}'].alignment = styles['align_center']

    # Row ZDGM
    ws[f'G{r_cust_start + 8}'] = "ZDGM"
    ws[f'G{r_cust_start + 8}'].font = styles['font_bold']
    ws.merge_cells(f'I{r_cust_start + 8}:J{r_cust_start + 8}')
    ws[f'I{r_cust_start + 8}'] = metadata.get('zdgm', '') or metadata.get('amm', '')
    ws[f'I{r_cust_start + 8}'].font = styles['font_bold']
    ws[f'I{r_cust_start + 8}'].alignment = styles['align_center']

    # Spacer (Row 23 / header gap)
    r_hdr_spacer = r_cust_start + 9
    ws.row_dimensions[r_hdr_spacer].height = 27

    # Particulars Table Header
    r_tbl_hdr = r_hdr_spacer + 1
    ws.row_dimensions[r_tbl_hdr].height = 20
    ws[f'A{r_tbl_hdr}'] = "SI.NO"
    ws[f'A{r_tbl_hdr}'].font = styles['font_header']
    ws[f'A{r_tbl_hdr}'].alignment = styles['align_center']

    ws.merge_cells(f'B{r_tbl_hdr}:G{r_tbl_hdr}')
    ws[f'B{r_tbl_hdr}'] = "PARTICULARS"
    ws[f'B{r_tbl_hdr}'].font = styles['font_header']
    ws[f'B{r_tbl_hdr}'].alignment = styles['align_center']

    ws[f'H{r_tbl_hdr}'] = "HSN/SAC"
    ws[f'H{r_tbl_hdr}'].font = styles['font_header']
    ws[f'H{r_tbl_hdr}'].alignment = styles['align_center']

    ws[f'I{r_tbl_hdr}'] = "NO.OF.Days /Act"
    ws[f'I{r_tbl_hdr}'].font = styles['font_header']
    ws[f'I{r_tbl_hdr}'].alignment = styles['align_center']

    ws[f'J{r_tbl_hdr}'] = "Amount"
    ws[f'J{r_tbl_hdr}'].font = styles['font_header']
    ws[f'J{r_tbl_hdr}'].alignment = styles['align_center']

    for c in range(1, 11):
        ws.cell(r_tbl_hdr, c).border = styles['box_border']

    # Particulars Data Rows (9 rows)
    start_act_r = r_tbl_hdr + 1
    num_act_rows = max(9, len(activity_groups))
    end_act_r = start_act_r + num_act_rows - 1

    curr_r = start_act_r
    subtotal_sum = 0.0
    total_qty = 0

    for idx, (act_name, act_info) in enumerate(activity_groups.items(), start=1):
        ws.row_dimensions[curr_r].height = 15
        ws.cell(curr_r, 1, idx).alignment = styles['align_center']
        ws.cell(curr_r, 1).font = styles['font_bold']
        
        ws.merge_cells(start_row=curr_r, start_column=2, end_row=curr_r, end_column=7)
        ws.cell(curr_r, 2, f"{act_name} Activities Expenses").font = styles['font_bold']
        ws.cell(curr_r, 2).alignment = styles['align_left']

        if idx == 1:
            ws.cell(curr_r, 8, "998596").alignment = styles['align_center']
            ws.cell(curr_r, 8).font = styles['font_bold']

        q = act_info.get('qty', len(act_info.get('rows', [])))
        total_qty += q
        c_q = ws.cell(curr_r, 9, q)
        c_q.alignment = styles['align_center']
        c_q.font = styles['font_bold']

        raw_amt = act_info.get('raw_amount', 0.0)
        amt_with_sc = raw_amt * (1 + service_charge_pct / 100.0)
        subtotal_sum += amt_with_sc

        c_amt = ws.cell(curr_r, 10, amt_with_sc)
        c_amt.alignment = styles['align_right']
        c_amt.font = styles['font_bold']
        c_amt.number_format = '#,##0.00'

        for c in range(1, 11):
            ws.cell(curr_r, c).border = styles['box_border']

        curr_r += 1

    # Fill remaining empty rows up to end_act_r
    while curr_r <= end_act_r:
        ws.row_dimensions[curr_r].height = 15
        ws.merge_cells(start_row=curr_r, start_column=2, end_row=curr_r, end_column=7)
        
        c_q = ws.cell(curr_r, 9, 0)
        c_q.alignment = styles['align_center']
        c_q.font = styles['font_bold']

        c_amt = ws.cell(curr_r, 10, 0.0)
        c_amt.alignment = styles['align_right']
        c_amt.font = styles['font_bold']
        c_amt.number_format = '#,##0.00'

        for c in range(1, 11):
            ws.cell(curr_r, c).border = styles['box_border']
        curr_r += 1

    # Total Days / Subtotal Row
    tot_days_r = end_act_r + 1
    ws.row_dimensions[tot_days_r].height = 18
    ws.merge_cells(start_row=tot_days_r, start_column=2, end_row=tot_days_r, end_column=7)

    c_tot_q = ws.cell(tot_days_r, 9, f"=SUM(I{start_act_r}:I{end_act_r})")
    c_tot_q.alignment = styles['align_center']
    c_tot_q.font = styles['font_bold']

    c_tot_amt = ws.cell(tot_days_r, 10, f"=SUM(J{start_act_r}:J{end_act_r})")
    c_tot_amt.alignment = styles['align_right']
    c_tot_amt.font = styles['font_bold']
    c_tot_amt.number_format = '#,##0.00'

    for c in range(1, 11):
        ws.cell(tot_days_r, c).border = Border(
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000'),
            left=Side(style='thin', color='000000') if c in [1, 8, 9, 10] else None,
            right=Side(style='thin', color='000000') if c in [7, 8, 9, 10] else None
        )

    # Taxes & Grand Total Block
    r_subtotal = tot_days_r + 1
    ws.row_dimensions[r_subtotal].height = 16
    ws.merge_cells(f'G{r_subtotal}:H{r_subtotal}')
    ws[f'G{r_subtotal}'] = "Sub Total"
    ws[f'G{r_subtotal}'].font = styles['font_bold']
    ws[f'G{r_subtotal}'].alignment = styles['align_left']
    ws[f'J{r_subtotal}'] = f"=J{tot_days_r}"
    ws[f'J{r_subtotal}'].font = styles['font_bold']
    ws[f'J{r_subtotal}'].alignment = styles['align_right']
    ws[f'J{r_subtotal}'].number_format = '#,##0.00'

    r_cgst = r_subtotal + 1
    ws.row_dimensions[r_cgst].height = 16
    ws[f'G{r_cgst}'] = "CGST"
    ws[f'G{r_cgst}'].font = styles['font_bold']
    ws.merge_cells(f'H{r_cgst}:I{r_cgst}')
    ws[f'H{r_cgst}'] = "9%"
    ws[f'H{r_cgst}'].font = styles['font_bold']
    ws[f'H{r_cgst}'].alignment = styles['align_center']
    ws[f'J{r_cgst}'] = f"=ROUND(J{r_subtotal}*0.09, 2)"
    ws[f'J{r_cgst}'].font = styles['font_bold']
    ws[f'J{r_cgst}'].alignment = styles['align_right']
    ws[f'J{r_cgst}'].number_format = '#,##0.00'

    r_sgst = r_subtotal + 2
    ws.row_dimensions[r_sgst].height = 16
    ws[f'G{r_sgst}'] = "SGST"
    ws[f'G{r_sgst}'].font = styles['font_bold']
    ws.merge_cells(f'H{r_sgst}:I{r_sgst}')
    ws[f'H{r_sgst}'] = "9%"
    ws[f'H{r_sgst}'].font = styles['font_bold']
    ws[f'H{r_sgst}'].alignment = styles['align_center']
    ws[f'J{r_sgst}'] = f"=ROUND(J{r_subtotal}*0.09, 2)"
    ws[f'J{r_sgst}'].font = styles['font_bold']
    ws[f'J{r_sgst}'].alignment = styles['align_right']
    ws[f'J{r_sgst}'].number_format = '#,##0.00'

    r_grand = r_subtotal + 3
    ws.row_dimensions[r_grand].height = 18
    ws[f'G{r_grand}'] = "Grand Total"
    ws[f'G{r_grand}'].font = styles['font_bold']
    ws.merge_cells(f'H{r_grand}:I{r_grand}')
    ws[f'H{r_grand}'] = "(Rounded Off)"
    ws[f'H{r_grand}'].font = styles['font_bold']
    ws[f'H{r_grand}'].alignment = styles['align_center']
    ws[f'J{r_grand}'] = f"=ROUND(SUM(J{r_subtotal}:J{r_sgst}), 0)"
    ws[f'J{r_grand}'].font = styles['font_bold']
    ws[f'J{r_grand}'].alignment = styles['align_right']
    ws[f'J{r_grand}'].number_format = '#,##0.00'

    for r in range(r_subtotal, r_grand + 1):
        for c in range(1, 11):
            ws.cell(r, c).border = styles['box_border']

    # Amount in words row
    r_words = r_grand + 1
    ws.row_dimensions[r_words].height = 18
    ws.merge_cells(f'A{r_words}:J{r_words}')
    cgst_calc = round(subtotal_sum * 0.09, 2)
    sgst_calc = round(subtotal_sum * 0.09, 2)
    grand_tot_calc = round(subtotal_sum + cgst_calc + sgst_calc)
    words = num_to_indian_words(grand_tot_calc)
    ws[f'A{r_words}'] = words
    ws[f'A{r_words}'].font = styles['font_bold']
    ws[f'A{r_words}'].alignment = styles['align_center']
    for c in range(1, 11):
        ws.cell(r_words, c).border = styles['box_border']

    # Bank Details & Signatures attached directly after words row (no gap)
    start_bank_r = r_words + 1
    r_bank_end = write_bank_and_signature_block(ws, start_bank_r, styles)

    return r_bank_end, r_subtotal, r_grand

def render_fmc_invoice_block(ws, start_row, copy_type, invoice_no, invoice_date, po_number, metadata, activity_groups, service_charge_pct, styles):
    """
    Renders one full FMC (New Gen) Tax Invoice block (ORIGINAL or DUPLICATE) starting at start_row.
    Returns: (end_row, r_subtotal, r_grand)
    """
    write_sbt_header_block(ws, start_row, copy_type, styles)

    # Customer & Metadata block
    r_cust_start = start_row + 8
    for r_c in range(r_cust_start, r_cust_start + 9):
        ws.row_dimensions[r_c].height = 14

    ws[f'A{r_cust_start}'] = "TO:"
    ws[f'A{r_cust_start}'].font = styles['font_bold']
    ws[f'A{r_cust_start + 1}'] = "New Gen Crop Solutions Pvt. Ltd."
    ws[f'A{r_cust_start + 1}'].font = styles['font_bold']
    ws[f'A{r_cust_start + 2}'] = "#39-10-5, 3rd FLOOR"
    ws[f'A{r_cust_start + 2}'].font = styles['font_regular']
    ws[f'A{r_cust_start + 3}'] = "VNR Towers, opp: Water Tanks"
    ws[f'A{r_cust_start + 3}'].font = styles['font_regular']
    ws[f'A{r_cust_start + 4}'] = "Labbipet, Vijayawada"
    ws[f'A{r_cust_start + 4}'].font = styles['font_regular']
    ws[f'A{r_cust_start + 5}'] = "GSTIN: 37AADCN6445Q1ZR"
    ws[f'A{r_cust_start + 5}'].font = styles['font_bold']

    # Right metadata
    ws.merge_cells(f'I{r_cust_start}:J{r_cust_start}')
    ws[f'I{r_cust_start}'] = "MSME"
    ws[f'I{r_cust_start}'].font = styles['font_bold']
    ws[f'I{r_cust_start}'].alignment = styles['align_center']

    ws[f'G{r_cust_start + 1}'] = "Invoice no:"
    ws[f'G{r_cust_start + 1}'].font = styles['font_bold']
    ws.merge_cells(f'I{r_cust_start + 1}:J{r_cust_start + 1}')
    ws[f'I{r_cust_start + 1}'] = invoice_no
    ws[f'I{r_cust_start + 1}'].font = styles['font_bold']
    ws[f'I{r_cust_start + 1}'].alignment = styles['align_center']

    ws[f'G{r_cust_start + 2}'] = "Invoice Date:"
    ws[f'G{r_cust_start + 2}'].font = styles['font_bold']
    ws.merge_cells(f'I{r_cust_start + 2}:J{r_cust_start + 2}')
    ws[f'I{r_cust_start + 2}'] = invoice_date
    ws[f'I{r_cust_start + 2}'].font = styles['font_bold']
    ws[f'I{r_cust_start + 2}'].alignment = styles['align_center']

    ws[f'G{r_cust_start + 3}'] = "Place Of Supply:"
    ws[f'G{r_cust_start + 3}'].font = styles['font_bold']
    ws.merge_cells(f'I{r_cust_start + 3}:J{r_cust_start + 3}')
    ws[f'I{r_cust_start + 3}'] = "Andhra Pradesh"
    ws[f'I{r_cust_start + 3}'].font = styles['font_bold']
    ws[f'I{r_cust_start + 3}'].alignment = styles['align_center']

    ws[f'G{r_cust_start + 4}'] = "New Gen Po:"
    ws[f'G{r_cust_start + 4}'].font = styles['font_bold']
    ws.merge_cells(f'I{r_cust_start + 4}:J{r_cust_start + 4}')
    ws[f'I{r_cust_start + 4}'] = po_number
    ws[f'I{r_cust_start + 4}'].font = styles['font_bold']
    ws[f'I{r_cust_start + 4}'].alignment = styles['align_center']

    ws[f'G{r_cust_start + 5}'] = "Product:"
    ws[f'G{r_cust_start + 5}'].font = styles['font_bold']
    ws.merge_cells(f'I{r_cust_start + 5}:J{r_cust_start + 5}')
    ws[f'I{r_cust_start + 5}'] = metadata.get('product', '')
    ws[f'I{r_cust_start + 5}'].font = styles['font_bold']
    ws[f'I{r_cust_start + 5}'].alignment = styles['align_center']

    ws[f'G{r_cust_start + 6}'] = "Crop:"
    ws[f'G{r_cust_start + 6}'].font = styles['font_bold']
    ws.merge_cells(f'I{r_cust_start + 6}:J{r_cust_start + 6}')
    ws[f'I{r_cust_start + 6}'] = metadata.get('crop', '')
    ws[f'I{r_cust_start + 6}'].font = styles['font_bold']
    ws[f'I{r_cust_start + 6}'].alignment = styles['align_center']

    ws[f'G{r_cust_start + 7}'] = "AMM"
    ws[f'G{r_cust_start + 7}'].font = styles['font_bold']
    ws.merge_cells(f'I{r_cust_start + 7}:J{r_cust_start + 7}')
    ws[f'I{r_cust_start + 7}'] = metadata.get('amm', '') or metadata.get('zdgm', '')
    ws[f'I{r_cust_start + 7}'].font = styles['font_bold']
    ws[f'I{r_cust_start + 7}'].alignment = styles['align_center']

    # Row AREA
    ws[f'G{r_cust_start + 8}'] = "AREA"
    ws[f'G{r_cust_start + 8}'].font = styles['font_bold']
    ws.merge_cells(f'I{r_cust_start + 8}:J{r_cust_start + 8}')
    ws[f'I{r_cust_start + 8}'] = metadata.get('area') or metadata.get('territory', '')
    ws[f'I{r_cust_start + 8}'].font = styles['font_bold']
    ws[f'I{r_cust_start + 8}'].alignment = styles['align_center']

    # Spacer (Row 23 / header gap)
    r_hdr_spacer = r_cust_start + 9
    ws.row_dimensions[r_hdr_spacer].height = 27

    # Particulars Table Header
    r_tbl_hdr = r_hdr_spacer + 1
    ws.row_dimensions[r_tbl_hdr].height = 20
    ws[f'A{r_tbl_hdr}'] = "SI.NO"
    ws[f'A{r_tbl_hdr}'].font = styles['font_header']
    ws[f'A{r_tbl_hdr}'].alignment = styles['align_center']

    ws.merge_cells(f'B{r_tbl_hdr}:G{r_tbl_hdr}')
    ws[f'B{r_tbl_hdr}'] = "PARTICULARS"
    ws[f'B{r_tbl_hdr}'].font = styles['font_header']
    ws[f'B{r_tbl_hdr}'].alignment = styles['align_center']

    ws[f'H{r_tbl_hdr}'] = "HSN/SAC"
    ws[f'H{r_tbl_hdr}'].font = styles['font_header']
    ws[f'H{r_tbl_hdr}'].alignment = styles['align_center']

    ws[f'I{r_tbl_hdr}'] = "NO.OF.Quantity"
    ws[f'I{r_tbl_hdr}'].font = styles['font_header']
    ws[f'I{r_tbl_hdr}'].alignment = styles['align_center']

    ws[f'J{r_tbl_hdr}'] = "Amount"
    ws[f'J{r_tbl_hdr}'].font = styles['font_header']
    ws[f'J{r_tbl_hdr}'].alignment = styles['align_center']

    for c in range(1, 11):
        ws.cell(r_tbl_hdr, c).border = styles['box_border']

    # Particulars Data Rows (9 rows)
    start_act_r = r_tbl_hdr + 1
    num_act_rows = max(9, len(activity_groups))
    end_act_r = start_act_r + num_act_rows - 1

    curr_r = start_act_r
    raw_subtotal = 0.0
    total_qty = 0

    for idx, (act_name, act_info) in enumerate(activity_groups.items(), start=1):
        ws.row_dimensions[curr_r].height = 15
        ws.cell(curr_r, 1, idx).alignment = styles['align_center']
        ws.cell(curr_r, 1).font = styles['font_bold']
        
        ws.merge_cells(start_row=curr_r, start_column=2, end_row=curr_r, end_column=7)
        ws.cell(curr_r, 2, f"{act_name} Activities Expenses").font = styles['font_bold']
        ws.cell(curr_r, 2).alignment = styles['align_left']

        if idx == 1:
            ws.cell(curr_r, 8, "998596").alignment = styles['align_center']
            ws.cell(curr_r, 8).font = styles['font_bold']

        q = act_info.get('qty', len(act_info.get('rows', [])))
        total_qty += q
        c_q = ws.cell(curr_r, 9, q)
        c_q.alignment = styles['align_center']
        c_q.font = styles['font_bold']

        raw_amt = act_info.get('raw_amount', 0.0)
        raw_subtotal += raw_amt

        c_amt = ws.cell(curr_r, 10, raw_amt)
        c_amt.alignment = styles['align_right']
        c_amt.font = styles['font_bold']
        c_amt.number_format = '#,##0.00'

        for c in range(1, 11):
            ws.cell(curr_r, c).border = styles['box_border']

        curr_r += 1

    # Fill remaining empty rows up to end_act_r
    while curr_r <= end_act_r:
        ws.row_dimensions[curr_r].height = 15
        ws.merge_cells(start_row=curr_r, start_column=2, end_row=curr_r, end_column=7)
        
        c_q = ws.cell(curr_r, 9, 0)
        c_q.alignment = styles['align_center']
        c_q.font = styles['font_bold']

        c_amt = ws.cell(curr_r, 10, 0.0)
        c_amt.alignment = styles['align_right']
        c_amt.font = styles['font_bold']
        c_amt.number_format = '#,##0.00'

        for c in range(1, 11):
            ws.cell(curr_r, c).border = styles['box_border']
        curr_r += 1

    # Activities Subtotal Row
    r_act_total = end_act_r + 1
    ws.row_dimensions[r_act_total].height = 18
    ws.merge_cells(start_row=r_act_total, start_column=2, end_row=r_act_total, end_column=7)

    c_tot_q = ws.cell(r_act_total, 9, f"=SUM(I{start_act_r}:I{end_act_r})")
    c_tot_q.alignment = styles['align_center']
    c_tot_q.font = styles['font_bold']

    c_tot_amt = ws.cell(r_act_total, 10, f"=SUM(J{start_act_r}:J{end_act_r})")
    c_tot_amt.alignment = styles['align_right']
    c_tot_amt.font = styles['font_bold']
    c_tot_amt.number_format = '#,##0.00'

    for c in range(1, 11):
        ws.cell(r_act_total, c).border = styles['box_border']

    # Service Charges Block (Rows r_sc1 and r_sc2)
    r_sc1 = r_act_total + 1
    r_sc2 = r_act_total + 2
    ws.row_dimensions[r_sc1].height = 15
    ws.row_dimensions[r_sc2].height = 15
    ws[f'B{r_sc1}'] = "Service"
    ws[f'B{r_sc1}'].font = styles['font_bold']
    ws[f'B{r_sc2}'] = "Charges"
    ws[f'B{r_sc2}'].font = styles['font_bold']

    ws.merge_cells(f'E{r_sc1}:F{r_sc1}')
    ws[f'E{r_sc1}'] = f"{service_charge_pct:.2f}%"
    ws[f'E{r_sc1}'].font = styles['font_bold']
    ws[f'E{r_sc1}'].alignment = styles['align_center']

    ws[f'J{r_sc2}'] = f"=ROUND(J{r_act_total}*{service_charge_pct/100.0:.4f}, 2)"
    ws[f'J{r_sc2}'].font = styles['font_bold']
    ws[f'J{r_sc2}'].alignment = styles['align_right']
    ws[f'J{r_sc2}'].number_format = '#,##0.00'

    for r in range(r_sc1, r_sc2 + 1):
        for c in range(1, 11):
            ws.cell(r, c).border = styles['box_border']

    # Total with Service Charges Row
    r_tot_sc = r_sc2 + 1
    ws.row_dimensions[r_tot_sc].height = 18
    ws.merge_cells(start_row=r_tot_sc, start_column=2, end_row=r_tot_sc, end_column=7)
    ws[f'J{r_tot_sc}'] = f"=J{r_act_total}+J{r_sc2}"
    ws[f'J{r_tot_sc}'].font = styles['font_bold']
    ws[f'J{r_tot_sc}'].alignment = styles['align_right']
    ws[f'J{r_tot_sc}'].number_format = '#,##0.00'

    for c in range(1, 11):
        ws.cell(r_tot_sc, c).border = Border(
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000'),
            left=Side(style='thin', color='000000') if c in [1, 8, 9, 10] else None,
            right=Side(style='thin', color='000000') if c in [7, 8, 9, 10] else None
        )

    # Taxes & Grand Total Block
    r_subtotal = r_tot_sc + 1
    ws.row_dimensions[r_subtotal].height = 16
    ws.merge_cells(f'G{r_subtotal}:H{r_subtotal}')
    ws[f'G{r_subtotal}'] = "Sub Total"
    ws[f'G{r_subtotal}'].font = styles['font_bold']
    ws[f'G{r_subtotal}'].alignment = styles['align_left']
    ws[f'J{r_subtotal}'] = f"=J{r_tot_sc}"
    ws[f'J{r_subtotal}'].font = styles['font_bold']
    ws[f'J{r_subtotal}'].alignment = styles['align_right']
    ws[f'J{r_subtotal}'].number_format = '#,##0.00'

    r_cgst = r_subtotal + 1
    ws.row_dimensions[r_cgst].height = 16
    ws[f'G{r_cgst}'] = "CGST"
    ws[f'G{r_cgst}'].font = styles['font_bold']
    ws.merge_cells(f'H{r_cgst}:I{r_cgst}')
    ws[f'H{r_cgst}'] = "9%"
    ws[f'H{r_cgst}'].font = styles['font_bold']
    ws[f'H{r_cgst}'].alignment = styles['align_center']
    ws[f'J{r_cgst}'] = f"=ROUND(J{r_subtotal}*0.09, 2)"
    ws[f'J{r_cgst}'].font = styles['font_bold']
    ws[f'J{r_cgst}'].alignment = styles['align_right']
    ws[f'J{r_cgst}'].number_format = '#,##0.00'

    r_sgst = r_subtotal + 2
    ws.row_dimensions[r_sgst].height = 16
    ws[f'G{r_sgst}'] = "SGST"
    ws[f'G{r_sgst}'].font = styles['font_bold']
    ws.merge_cells(f'H{r_sgst}:I{r_sgst}')
    ws[f'H{r_sgst}'] = "9%"
    ws[f'H{r_sgst}'].font = styles['font_bold']
    ws[f'H{r_sgst}'].alignment = styles['align_center']
    ws[f'J{r_sgst}'] = f"=ROUND(J{r_subtotal}*0.09, 2)"
    ws[f'J{r_sgst}'].font = styles['font_bold']
    ws[f'J{r_sgst}'].alignment = styles['align_right']
    ws[f'J{r_sgst}'].number_format = '#,##0.00'

    r_grand = r_subtotal + 3
    ws.row_dimensions[r_grand].height = 18
    ws[f'G{r_grand}'] = "Grand Total"
    ws[f'G{r_grand}'].font = styles['font_bold']
    ws.merge_cells(f'H{r_grand}:I{r_grand}')
    ws[f'H{r_grand}'] = "(Rounded Off)"
    ws[f'H{r_grand}'].font = styles['font_bold']
    ws[f'H{r_grand}'].alignment = styles['align_center']
    ws[f'J{r_grand}'] = f"=ROUND(SUM(J{r_subtotal}:J{r_sgst}), 0)"
    ws[f'J{r_grand}'].font = styles['font_bold']
    ws[f'J{r_grand}'].alignment = styles['align_right']
    ws[f'J{r_grand}'].number_format = '#,##0.00'

    for r in range(r_subtotal, r_grand + 1):
        for c in range(1, 11):
            ws.cell(r, c).border = styles['box_border']

    # Amount in words row
    r_words = r_grand + 1
    ws.row_dimensions[r_words].height = 18
    ws.merge_cells(f'A{r_words}:J{r_words}')
    sc_amt = round(raw_subtotal * (service_charge_pct / 100.0), 2)
    tot_with_sc = raw_subtotal + sc_amt
    cgst_calc = round(tot_with_sc * 0.09, 2)
    sgst_calc = round(tot_with_sc * 0.09, 2)
    grand_tot_calc = round(tot_with_sc + cgst_calc + sgst_calc)
    words = num_to_indian_words(grand_tot_calc)
    ws[f'A{r_words}'] = words
    ws[f'A{r_words}'].font = styles['font_bold']
    ws[f'A{r_words}'].alignment = styles['align_center']
    for c in range(1, 11):
        ws.cell(r_words, c).border = styles['box_border']

    # Bank Details & Signatures attached directly after words row (no gap)
    start_bank_r = r_words + 1
    r_bank_end = write_bank_and_signature_block(ws, start_bank_r, styles)

    return r_bank_end, r_subtotal, r_grand

def build_corteva_sheet1_invoice(ws, invoice_no, invoice_date, po_number, metadata, activity_groups, service_charge_pct, styles):
    """
    Builds Sheet1 Tax Invoice for Corteva with 5 rows gap at top, both ORIGINAL and DUPLICATE copies, and balanced margins.
    """
    ws.views.sheetView[0].showGridLines = True

    # Exact column widths calibrated for 1-page A4 Portrait with balanced side margins
    col_widths = {'A': 5.5, 'B': 10.5, 'C': 10.5, 'D': 9.5, 'E': 9.5, 'F': 9.5, 'G': 12.5, 'H': 10.5, 'I': 13.5, 'J': 14.5}
    for col_let, w in col_widths.items():
        ws.column_dimensions[col_let].width = w

    # 5 top gap rows for Page 1
    for r_top in range(1, 6):
        ws.row_dimensions[r_top].height = 14

    # 1. Render ORIGINAL Invoice block starting at Row 6
    orig_start_r = 6
    orig_end_r, r_subtotal, r_grand = render_corteva_invoice_block(
        ws, orig_start_r, "ORIGINAL", invoice_no, invoice_date, po_number, metadata, activity_groups, service_charge_pct, styles
    )

    # 2. Insert Page Break right after ORIGINAL invoice block
    ws.row_breaks.append(Break(id=orig_end_r))

    # 5 top gap rows for Page 2
    for r_top2 in range(orig_end_r + 1, orig_end_r + 6):
        ws.row_dimensions[r_top2].height = 14

    # 3. Render DUPLICATE Invoice block starting after 5 gap rows
    dup_start_r = orig_end_r + 6
    dup_end_r, _, _ = render_corteva_invoice_block(
        ws, dup_start_r, "DUPLICATE", invoice_no, invoice_date, po_number, metadata, activity_groups, service_charge_pct, styles
    )

    # Set Print Area spanning both ORIGINAL and DUPLICATE pages with balanced side margins
    setup_page_print_fit(ws, print_area=f"A1:J{dup_end_r}", orientation="portrait")

    return r_subtotal, r_grand

def build_fmc_sheet1_invoice(ws, invoice_no, invoice_date, po_number, metadata, activity_groups, service_charge_pct, styles):
    """
    Builds Sheet1 Tax Invoice for FMC (New Gen) with 5 rows gap at top, both ORIGINAL and DUPLICATE copies, and balanced margins.
    """
    ws.views.sheetView[0].showGridLines = True

    # Exact column widths calibrated for 1-page A4 Portrait with balanced side margins
    col_widths = {'A': 5.5, 'B': 10.5, 'C': 10.5, 'D': 9.5, 'E': 9.5, 'F': 9.5, 'G': 12.5, 'H': 10.5, 'I': 13.5, 'J': 14.5}
    for col_let, w in col_widths.items():
        ws.column_dimensions[col_let].width = w

    # 5 top gap rows for Page 1
    for r_top in range(1, 6):
        ws.row_dimensions[r_top].height = 14

    # 1. Render ORIGINAL Invoice block starting at Row 6
    orig_start_r = 6
    orig_end_r, r_subtotal, r_grand = render_fmc_invoice_block(
        ws, orig_start_r, "ORIGINAL", invoice_no, invoice_date, po_number, metadata, activity_groups, service_charge_pct, styles
    )

    # 2. Insert Page Break right after ORIGINAL invoice block
    ws.row_breaks.append(Break(id=orig_end_r))

    # 5 top gap rows for Page 2
    for r_top2 in range(orig_end_r + 1, orig_end_r + 6):
        ws.row_dimensions[r_top2].height = 14

    # 3. Render DUPLICATE Invoice block starting after 5 gap rows
    dup_start_r = orig_end_r + 6
    dup_end_r, _, _ = render_fmc_invoice_block(
        ws, dup_start_r, "DUPLICATE", invoice_no, invoice_date, po_number, metadata, activity_groups, service_charge_pct, styles
    )

    # Set Print Area spanning both ORIGINAL and DUPLICATE pages with balanced side margins
    setup_page_print_fit(ws, print_area=f"A1:J{dup_end_r}", orientation="portrait")

    return r_subtotal, r_grand

def build_corteva_sheet2_details(ws, short_iv, records, service_charge_pct, styles):
    """
    Builds Sheet2 for Corteva matching SS2 layout with PO Number column included:
    Grouped by Activity, then by TBM, with IV NO : {IV} ( index ), subtotals, and 5% service charge row.
    """
    ws.views.sheetView[0].showGridLines = True
    setup_page_print_fit(ws, print_area=None, orientation="landscape")

    col_widths = {
        'A': 6, 'B': 12, 'C': 16, 'D': 16, 'E': 14, 'F': 14, 'G': 14, 'H': 12,
        'I': 12, 'J': 14, 'K': 12, 'L': 15, 'M': 14, 'N': 12, 'O': 12, 'P': 14, 'Q': 20
    }
    for col_let, w in col_widths.items():
        ws.column_dimensions[col_let].width = w

    act_groups = {}
    for r in records:
        act = clean_str(r.get('activity', '')).upper() or "GENERAL"
        tbm = clean_str(r.get('tbm', '')).title() or "TBM"
        terr = clean_str(r.get('territory', '')).title() or ""
        
        if act not in act_groups:
            act_groups[act] = {}
        tbm_key = (tbm, terr)
        if tbm_key not in act_groups[act]:
            act_groups[act][tbm_key] = []
        act_groups[act][tbm_key].append(r)

    current_r = 1
    iv_idx = 1

    headers = [
        "S.No", "Date", "ZDGM", "TBM", "MDO", "Territory", "Product", "Crop", "Activity", "Village",
        "No.of Farmers", "Tent/Hall Suppliers Charges", "Food Expenses", "Transport", "Others", "Total", "PO Number"
    ]

    for act_name, tbm_dict in act_groups.items():
        # Section Header: IV NO : 67 ( 1 ) across Columns 1 to 17 (A to Q)
        ws.merge_cells(start_row=current_r, start_column=1, end_row=current_r, end_column=17)
        c_iv = ws.cell(current_r, 1, f"IV NO : {short_iv} ( {iv_idx} )")
        c_iv.font = styles['font_green_title']
        c_iv.alignment = styles['align_center']
        ws.row_dimensions[current_r].height = 22
        current_r += 1
        iv_idx += 1

        act_table_tot_rows = []

        for (tbm_name, terr_name), rows_list in tbm_dict.items():
            ws.merge_cells(start_row=current_r, start_column=1, end_row=current_r, end_column=17)
            terr_str = f"{terr_name} " if terr_name else ""
            c_tbl = ws.cell(current_r, 1, f"Activities Expenses by {terr_str}Tbm {tbm_name}")
            c_tbl.font = styles['font_green_title']
            c_tbl.alignment = styles['align_center']
            ws.row_dimensions[current_r].height = 20
            current_r += 1

            hdr_r = current_r
            ws.row_dimensions[hdr_r].height = 24
            for c_idx, h in enumerate(headers, start=1):
                cell = ws.cell(hdr_r, c_idx, h)
                cell.font = styles['font_green_bold'] if c_idx in [1, 9, 10, 11, 12, 13, 14, 15, 16, 17] else styles['font_bold']
                cell.alignment = styles['align_center_wrap']
                cell.border = styles['thin_border']
            current_r += 1

            start_data_r = current_r
            for s_no, item in enumerate(rows_list, start=1):
                ws.row_dimensions[current_r].height = 18
                ws.cell(current_r, 1, s_no).alignment = styles['align_center']
                ws.cell(current_r, 1).font = styles['font_bold']

                ws.cell(current_r, 2, item.get('date', '')).alignment = styles['align_center']
                ws.cell(current_r, 3, item.get('zdgm', '')).alignment = styles['align_left']
                ws.cell(current_r, 4, item.get('tbm', '')).alignment = styles['align_left']
                ws.cell(current_r, 5, item.get('mdo', '')).alignment = styles['align_left']
                ws.cell(current_r, 6, item.get('territory', '')).alignment = styles['align_center']
                ws.cell(current_r, 7, item.get('product', '')).alignment = styles['align_center']
                ws.cell(current_r, 8, item.get('crop', '')).alignment = styles['align_center']
                ws.cell(current_r, 9, item.get('activity', '')).alignment = styles['align_center']
                ws.cell(current_r, 10, item.get('village', '')).alignment = styles['align_left']
                
                c_farm = ws.cell(current_r, 11, item.get('farmers', 0))
                c_farm.alignment = styles['align_center']
                c_farm.number_format = '#,##0'

                for c_off, key in enumerate(['tent', 'food', 'transport', 'others'], start=12):
                    v = item.get(key, 0.0)
                    cell_v = ws.cell(current_r, c_off, v if v > 0 else "")
                    cell_v.alignment = styles['align_right']
                    if v > 0: cell_v.number_format = '#,##0'

                c_row_tot = ws.cell(current_r, 16, f"=SUM(L{current_r}:O{current_r})")
                c_row_tot.alignment = styles['align_right']
                c_row_tot.font = styles['font_bold']
                c_row_tot.number_format = '#,##0'

                # Col 17 (Q): Full PO Number
                c_po = ws.cell(current_r, 17, item.get('po_number', ''))
                c_po.alignment = styles['align_center']
                c_po.font = styles['font_bold']

                for c in range(1, 18):
                    ws.cell(current_r, c).border = styles['thin_border']
                current_r += 1

            end_data_r = current_r - 1

            tot_r = current_r
            ws.row_dimensions[tot_r].height = 20
            ws.cell(tot_r, 15, "Total").font = styles['font_green_bold']
            ws.cell(tot_r, 15).alignment = styles['align_right']
            
            c_tbl_sum = ws.cell(tot_r, 16, f"=SUM(P{start_data_r}:P{end_data_r})")
            c_tbl_sum.font = styles['font_green_bold']
            c_tbl_sum.alignment = styles['align_right']
            c_tbl_sum.number_format = '#,##0'

            for c in range(1, 18):
                ws.cell(tot_r, c).border = styles['thin_border']
            
            act_table_tot_rows.append(tot_r)
            current_r += 1

        subtot_r = current_r
        ws.row_dimensions[subtot_r].height = 20
        ws.merge_cells(start_row=subtot_r, start_column=8, end_row=subtot_r, end_column=9)
        ws.cell(subtot_r, 8, f"Total {act_name}").font = styles['font_bold']
        ws.cell(subtot_r, 8).alignment = styles['align_center']
        
        sum_refs = [f"P{r}" for r in act_table_tot_rows]
        c_sub_val = ws.cell(subtot_r, 10, f"=SUM({','.join(sum_refs)})")
        c_sub_val.font = styles['font_bold']
        c_sub_val.alignment = styles['align_right']
        c_sub_val.number_format = '#,##0.00'

        sc_r = current_r + 1
        ws.row_dimensions[sc_r].height = 18
        c_sc_val = ws.cell(sc_r, 10, f"=J{subtot_r}*{service_charge_pct/100.0:.4f}")
        c_sc_val.font = styles['font_bold']
        c_sc_val.alignment = styles['align_right']
        c_sc_val.number_format = '#,##0.00'

        tot_sc_r = current_r + 2
        ws.row_dimensions[tot_sc_r].height = 20
        c_tot_sc = ws.cell(tot_sc_r, 10, f"=J{subtot_r}+J{sc_r}")
        c_tot_sc.font = styles['font_bold']
        c_tot_sc.alignment = styles['align_right']
        c_tot_sc.number_format = '#,##0.00'

        for r_box in range(subtot_r, tot_sc_r + 1):
            for c in range(1, 18):
                ws.cell(r_box, c).border = styles['thin_border']

        current_r = tot_sc_r + 2

def build_fmc_sheet2_details(ws, short_iv, records, styles):
    """
    Builds Sheet2 for FMC matching SS5 layout with PO Number column included:
    Grouped by Territory tables with header IV NO : {IV} and bottom sum formula =P7+P17.
    """
    ws.views.sheetView[0].showGridLines = True
    setup_page_print_fit(ws, print_area=None, orientation="landscape")

    col_widths = {
        'A': 6, 'B': 12, 'C': 16, 'D': 16, 'E': 14, 'F': 14, 'G': 14, 'H': 12,
        'I': 12, 'J': 14, 'K': 12, 'L': 15, 'M': 14, 'N': 12, 'O': 12, 'P': 14, 'Q': 20
    }
    for col_let, w in col_widths.items():
        ws.column_dimensions[col_let].width = w

    terr_groups = {}
    for r in records:
        terr = clean_str(r.get('territory', '')).title() or "General"
        if terr not in terr_groups:
            terr_groups[terr] = []
        terr_groups[terr].append(r)

    current_r = 1

    # Top header: IV NO : 72 across Columns 1 to 17 (A to Q)
    ws.merge_cells('A1:Q1')
    c_iv = ws['A1']
    c_iv.value = f"IV NO : {short_iv}"
    c_iv.font = styles['font_green_title']
    c_iv.alignment = styles['align_center']
    ws.row_dimensions[1].height = 24
    current_r = 2

    headers = [
        "SI No.", "Date", "Area Manager", "TBM/SC/SO", "MDO", "Territory", "Product", "Crop", "Activity", "Village",
        "No.of Farmers", "Tent/Hall Suppliers Charges", "Food Expenses", "Trans port", "Others /Gifts", "Total", "PO Number"
    ]

    terr_total_cells = []

    for terr_name, rows_list in terr_groups.items():
        ws.merge_cells(start_row=current_r, start_column=1, end_row=current_r, end_column=17)
        c_title = ws.cell(current_r, 1, f"Activities Expenses by {terr_name} Territory")
        c_title.font = styles['font_green_title']
        c_title.alignment = styles['align_center']
        ws.row_dimensions[current_r].height = 22
        current_r += 1

        hdr_r = current_r
        ws.row_dimensions[hdr_r].height = 24
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(hdr_r, c_idx, h)
            cell.font = styles['font_green_bold']
            cell.alignment = styles['align_center_wrap']
            cell.border = styles['thin_border']
        current_r += 1

        start_data_r = current_r
        for s_no, item in enumerate(rows_list, start=1):
            ws.row_dimensions[current_r].height = 18
            ws.cell(current_r, 1, s_no).alignment = styles['align_center']
            ws.cell(current_r, 1).font = styles['font_green_bold']

            ws.cell(current_r, 2, item.get('date', '')).alignment = styles['align_center']
            ws.cell(current_r, 3, item.get('zdgm', '') or item.get('amm', '')).alignment = styles['align_left']
            ws.cell(current_r, 4, item.get('tbm', '')).alignment = styles['align_left']
            ws.cell(current_r, 5, item.get('mdo', '')).alignment = styles['align_left']
            ws.cell(current_r, 6, item.get('territory', '')).alignment = styles['align_center']
            ws.cell(current_r, 7, item.get('product', '')).alignment = styles['align_center']
            ws.cell(current_r, 8, item.get('crop', '')).alignment = styles['align_center']
            ws.cell(current_r, 9, item.get('activity', '')).alignment = styles['align_center']
            ws.cell(current_r, 10, item.get('village', '')).alignment = styles['align_left']

            c_farm = ws.cell(current_r, 11, item.get('farmers', 0))
            c_farm.alignment = styles['align_center']
            c_farm.number_format = '#,##0'

            for c_off, key in enumerate(['tent', 'food', 'transport', 'others'], start=12):
                v = item.get(key, 0.0)
                cell_v = ws.cell(current_r, c_off, v if v > 0 else "")
                cell_v.alignment = styles['align_right']
                if v > 0: cell_v.number_format = '#,##0'

            c_tot = ws.cell(current_r, 16, f"=SUM(L{current_r}:O{current_r})")
            c_tot.alignment = styles['align_right']
            c_tot.font = styles['font_green_bold']
            c_tot.number_format = '#,##0'

            # Col 17 (Q): Full PO Number
            c_po = ws.cell(current_r, 17, item.get('po_number', ''))
            c_po.alignment = styles['align_center']
            c_po.font = styles['font_green_bold']

            for c in range(1, 18):
                ws.cell(current_r, c).border = styles['thin_border']
            current_r += 1

        end_data_r = current_r - 1

        tot_r = current_r
        ws.row_dimensions[tot_r].height = 20
        ws.cell(tot_r, 15, "Total").font = styles['font_green_bold']
        ws.cell(tot_r, 15).alignment = styles['align_right']
        
        c_sum = ws.cell(tot_r, 16, f"=SUM(P{start_data_r}:P{end_data_r})")
        c_sum.font = styles['font_green_bold']
        c_sum.alignment = styles['align_right']
        c_sum.number_format = '#,##0'

        for c in range(1, 18):
            ws.cell(tot_r, c).border = styles['thin_border']

        terr_total_cells.append(f"P{tot_r}")
        current_r += 2

    bottom_r = current_r
    ws.row_dimensions[bottom_r].height = 22
    ws.merge_cells(start_row=bottom_r, start_column=3, end_row=bottom_r, end_column=4)
    c_lbl = ws.cell(bottom_r, 3, "Total")
    c_lbl.font = styles['font_bold']
    c_lbl.alignment = styles['align_center']

    if terr_total_cells:
        formula_sum = f"={'+'.join(terr_total_cells)}"
    else:
        formula_sum = "0"
    
    c_grand = ws.cell(bottom_r, 5, formula_sum)
    c_grand.font = styles['font_bold']
    c_grand.alignment = styles['align_center']
    c_grand.number_format = '#,##0'

    for c in range(3, 6):
        ws.cell(bottom_r, c).border = styles['box_border']

def build_corteva_summary_sheet(ws, invoice_no, invoice_date, po_number, metadata, po_value, r_subtotal, r_grand, styles):
    """
    Builds the 11-column Corteva Invoice Summary sheet dynamically linking to Sheet1 Sub Total & Grand Total.
    """
    ws.views.sheetView[0].showGridLines = True
    setup_page_print_fit(ws, print_area=None, orientation="landscape")

    col_widths = {
        'A': 16, 'B': 16, 'C': 28, 'D': 24, 'E': 18, 'F': 16,
        'G': 24, 'H': 24, 'I': 18, 'J': 18, 'K': 32
    }
    for col_let, w in col_widths.items():
        ws.column_dimensions[col_let].width = w

    headers = [
        "Vendor Code", "Payment Term", "Entity", "Vendor Name", "Invoice No.",
        "Invoice Date", "Invoice Amount (EXC GST)", "Invoice Amount (INC GST)",
        "PO Number", "PO Value", "Name of Corteva Requester/Receiver"
    ]
    
    ws.row_dimensions[1].height = 26
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(1, c_idx, h)
        cell.font = styles['font_header']
        cell.alignment = styles['align_center_wrap']
        cell.border = styles['box_border']

    # Row 2 Data (Linking to Sheet1 J{r_subtotal} and J{r_grand})
    ws.row_dimensions[2].height = 22
    ws.cell(2, 1, "80141626").alignment = styles['align_center']
    ws.cell(2, 2, "45 Days").alignment = styles['align_center']
    ws.cell(2, 3, "Corteva Agriscience India Private Limited").alignment = styles['align_left']
    ws.cell(2, 4, "Radhadevi Kamisetty").alignment = styles['align_left']
    ws.cell(2, 5, invoice_no).alignment = styles['align_center']
    ws.cell(2, 6, invoice_date).alignment = styles['align_center']
    
    # Amount Exc GST linking to Sheet1 Sub Total
    c_exc = ws.cell(2, 7, f"=Sheet1!J{r_subtotal}")
    c_exc.alignment = styles['align_right']
    c_exc.number_format = '#,##0'

    # Amount Inc GST linking to Sheet1 Grand Total
    c_inc = ws.cell(2, 8, f"=Sheet1!J{r_grand}")
    c_inc.alignment = styles['align_right']
    c_inc.number_format = '#,##0'

    ws.cell(2, 9, po_number).alignment = styles['align_center']
    
    c_val = ws.cell(2, 10, po_value if po_value else 250000)
    c_val.alignment = styles['align_right']
    c_val.number_format = '#,##0'

    requester = metadata.get('zdgm', '') or metadata.get('amm', '') or "R.Bhaskar"
    ws.cell(2, 11, requester).alignment = styles['align_left']

    for c in range(1, 12):
        ws.cell(2, c).font = styles['font_regular']
        ws.cell(2, c).border = styles['box_border']

def generate_or_update_invoice(
    company,
    tbm_summary_path,
    save_folder_path,
    invoice_number,
    po_number,
    service_charge_pct=5.0,
    invoice_date=None,
    po_value=None,
    requester_name=None,
    area=None
):
    """
    Main function to generate or update a PO Tax Invoice workbook.
    """
    if not po_number or not str(po_number).strip():
        raise ValueError("PO Number is a mandatory field.")

    if not invoice_number or not str(invoice_number).strip():
        raise ValueError("Invoice Number is a mandatory field.")

    target_po = str(po_number).strip()
    inv_num_raw = str(invoice_number).strip()
    company_clean = str(company).strip().title() # "Corteva" or "Fmc"

    short_iv = format_short_iv(inv_num_raw)
    full_iv = format_full_iv(inv_num_raw)

    if not invoice_date:
        invoice_date = datetime.date.today().strftime("%d-%m-%Y")

    save_dir = Path(save_folder_path).resolve()
    save_dir.mkdir(parents=True, exist_ok=True)

    # 1. Extract activity records for this PO from All-TBMs-Summary.xlsx
    records, metadata = extract_tables_for_po(tbm_summary_path, target_po)
    if not records:
        raise ValueError(f"No activity expense records found for PO {target_po} in {Path(tbm_summary_path).name}")

    if requester_name:
        metadata['zdgm'] = requester_name
        metadata['amm'] = requester_name

    if area and str(area).strip():
        metadata['area'] = str(area).strip()
        metadata['territory'] = str(area).strip()

    # Group by Activity for Sheet1 Particulars
    activity_groups = {}
    for r in records:
        act = clean_str(r.get('activity', '')).upper() or "GENERAL"
        if act not in activity_groups:
            activity_groups[act] = {'qty': 0, 'raw_amount': 0.0, 'rows': []}
        activity_groups[act]['qty'] += 1
        activity_groups[act]['raw_amount'] += float(r.get('total', 0.0))
        activity_groups[act]['rows'].append(r)

    styles = get_base_styles()

    # 2. Check if an invoice file for this invoice number / PO already exists
    existing_file = None
    for f in save_dir.glob("*.xlsx"):
        if f.stem == short_iv or f.stem == full_iv or f.stem.lower() == inv_num_raw.lower() or target_po.upper() in f.stem.upper():
            existing_file = f
            break

    is_update = existing_file is not None and existing_file.exists()
    out_file_name = f"{short_iv}.xlsx" if not is_update else existing_file.name
    out_path = save_dir / out_file_name

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    ws_sheet1 = wb.create_sheet(title="Sheet1")
    ws_sheet2 = wb.create_sheet(title="Sheet2")

    if company_clean.startswith("Corteva"):
        r_subtotal, r_grand = build_corteva_sheet1_invoice(ws_sheet1, full_iv, invoice_date, target_po, metadata, activity_groups, service_charge_pct, styles)
        build_corteva_sheet2_details(ws_sheet2, short_iv, records, service_charge_pct, styles)
        ws_summary = wb.create_sheet(title="Sheet4")
        build_corteva_summary_sheet(ws_summary, full_iv, invoice_date, target_po, metadata, po_value, r_subtotal, r_grand, styles)
    else:
        # FMC / New Gen
        r_subtotal, r_grand = build_fmc_sheet1_invoice(ws_sheet1, full_iv, invoice_date, target_po, metadata, activity_groups, service_charge_pct, styles)
        build_fmc_sheet2_details(ws_sheet2, short_iv, records, styles)

    wb.save(out_path)
    wb.close()

    total_activities = len(records)
    total_raw_amount = sum(float(r.get('total', 0.0)) for r in records)
    total_with_sc = total_raw_amount * (1 + service_charge_pct / 100.0)
    cgst_amt = round(total_with_sc * 0.09, 2)
    sgst_amt = round(total_with_sc * 0.09, 2)
    grand_total = round(total_with_sc + cgst_amt + sgst_amt)

    return {
        "success": True,
        "isUpdate": is_update,
        "message": f"Invoice {full_iv} {'updated & appended' if is_update else 'generated successfully'} for PO {target_po} in {out_path.name}!",
        "outputPath": str(out_path),
        "invoiceNo": full_iv,
        "shortInvoiceNo": short_iv,
        "poNumber": target_po,
        "company": company_clean,
        "area": metadata.get('area') or metadata.get('territory', ''),
        "invoiceDate": invoice_date,
        "totalActivities": total_activities,
        "subTotalExcGst": round(total_with_sc, 2),
        "grandTotalIncGst": grand_total,
        "grandTotalWords": num_to_indian_words(grand_total)
    }
