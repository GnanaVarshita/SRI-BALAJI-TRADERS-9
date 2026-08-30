import os
import sys
import unittest
import shutil
from pathlib import Path
import openpyxl

BACKEND_DIR = Path(__file__).resolve().parent.parent
sys.path.append(str(BACKEND_DIR))

import details_of_bills_generator
import invoice_generator

class TestDetailsOfBills(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.test_dir = BACKEND_DIR / "test_scratch_details_of_bills"
        cls.test_dir.mkdir(parents=True, exist_ok=True)
        cls.invoices_dir = cls.test_dir / "generated_invoices"
        cls.invoices_dir.mkdir(parents=True, exist_ok=True)
        cls.details_excel_path = cls.test_dir / "Details of Bills 2026 TO 2027.xlsx"
        cls.cards_excel_path = cls.test_dir / "Nandyala FMC Budget.xlsx"

        # Create sample All-TBM summary to generate sample invoices
        cls.tbm_summary_path = cls.test_dir / "FMC-All-TBMs-Summary.xlsx"
        cls._create_sample_fmc_summary(cls.tbm_summary_path)

        # Generate sample invoice 72
        invoice_generator.generate_or_update_invoice(
            company="FMC",
            tbm_summary_path=str(cls.tbm_summary_path),
            save_folder_path=str(cls.invoices_dir),
            invoice_number="SBT26270072",
            po_number="500BB20260710166",
            service_charge_pct=5.0,
            invoice_date="09-04-2026",
            requester_name="Madhavareddy",
            area="Nandyala"
        )

        if cls.details_excel_path.exists():
            cls.details_excel_path.unlink()

        # Create sample Budget PO summary cards workbook
        cls._create_sample_budget_cards(cls.cards_excel_path)

    @classmethod
    def tearDownClass(cls):
        if cls.test_dir.exists():
            shutil.rmtree(cls.test_dir, ignore_errors=True)

    @classmethod
    def _create_sample_fmc_summary(cls, path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(1, 1, "ACTIVITIES EXPENSES BY NANDYALA TERRITORY")
        headers = ["SI No", "Date", "ZDGM", "TBM", "MDO", "Territory", "Product", "Crop", "Activity", "Village",
                   "No. of Farmers", "Tent/Hall/Chairs Suppliers Charges", "Food Expenses", "Transport", "Others/Gifts", "Total", "PO Number"]
        for c, h in enumerate(headers, 1):
            ws.cell(2, c, h)
        
        ws.cell(3, 1, 1)
        ws.cell(3, 2, "09-04-2026")
        ws.cell(3, 3, "Madhavareddy")
        ws.cell(3, 4, "Kiran Kumar")
        ws.cell(3, 5, "Suresh")
        ws.cell(3, 6, "Nandyala")
        ws.cell(3, 7, "Coragen")
        ws.cell(3, 8, "Paddy")
        ws.cell(3, 9, "OFM")
        ws.cell(3, 10, "Koilkuntla")
        ws.cell(3, 11, 25)
        ws.cell(3, 12, 1200)
        ws.cell(3, 13, 2500)
        ws.cell(3, 14, 800)
        ws.cell(3, 15, 0)
        ws.cell(3, 16, 4500)
        ws.cell(3, 17, "500BB20260710166")

        ws.cell(4, 1, "Total")
        ws.cell(4, 16, 4500)
        wb.save(path)
        wb.close()

    @classmethod
    def _create_sample_budget_cards(cls, path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "38604-10177"

        # Card 1: 500BB20260710166
        ws["A1"] = "500BB20260710166"
        ws["D1"] = "PRODUCT"
        ws["F1"] = "Coragen"
        ws["G1"] = "Nandyala"
        ws["A7"] = "I.V NO"
        ws["B7"] = "DATE"
        ws["C7"] = "AREA"
        ws["D7"] = "PO NUMBER"
        ws["E7"] = "BUDGET TYP"
        ws["F7"] = "PRODUCT"
        ws["G7"] = "CROP"
        ws["H7"] = "ACTIVITY"
        ws["I7"] = "ZDGM"
        ws["J7"] = "MIE"
        ws["K7"] = "NO.OF.Activities"

        # Initial row 8 without IV NO and DATE
        ws["C8"] = "Nandyala"
        ws["D8"] = "500BB20260710166"
        ws["E8"] = "Brand"
        ws["F8"] = "Coragen"
        ws["G8"] = "Paddy"
        ws["H8"] = "OFM"
        ws["I8"] = "Madhavareddy"
        ws["J8"] = "Kiran Kumar"
        ws["K8"] = 1

        wb.save(path)
        wb.close()

    def test_01_receivable_date_calculation(self):
        # 09-04-2026 + 45 days = 24-05-2026
        rec_date = details_of_bills_generator.calculate_receivable_date("09-04-2026", 45)
        self.assertEqual(rec_date, "24-05-2026")

        # 16-04-2026 + 45 days = 31-05-2026
        rec_date2 = details_of_bills_generator.calculate_receivable_date("16-04-2026", 45)
        self.assertEqual(rec_date2, "31-05-2026")

    def test_02_create_details_of_bills_structure(self):
        wb, ws, is_new = details_of_bills_generator.create_or_load_details_of_bills_wb(
            str(self.details_excel_path),
            financial_year="APRIL 2026 to MARCH 2027"
        )
        self.assertTrue(is_new)
        self.assertEqual(ws["H2"].value, "APRIL 2026 to MARCH 2027")
        self.assertEqual(ws["A4"].value, "I.V NO")
        self.assertEqual(ws["B4"].value, "DATE")
        self.assertEqual(ws["C4"].value, "AREA")
        self.assertEqual(ws["D4"].value, "PO NUMBER")
        self.assertEqual(ws["P4"].value, "RECEIVABL\nE DATE")
        self.assertEqual(ws["U4"].value, "Total Amount")

        # Check Row 3 top sum formulas
        self.assertIn("=SUM(L5:", ws["L3"].value)
        self.assertIn("=SUM(M5:", ws["M3"].value)
        self.assertIn("=SUM(N5:", ws["N3"].value)
        self.assertIn("=SUM(O5:", ws["O3"].value)

        wb.save(str(self.details_excel_path))
        wb.close()

    def test_03_scan_and_append_invoices_and_cards_sync(self):
        res = details_of_bills_generator.scan_and_append_invoices(
            details_excel_path=str(self.details_excel_path),
            invoices_folder_path=str(self.invoices_dir),
            budget_cards_path=str(self.cards_excel_path),
            financial_year="APRIL 2026 to MARCH 2027"
        )
        self.assertTrue(res["success"])
        self.assertEqual(res["totalInvoicesFound"], 1)
        self.assertIn("72", res["appendedInvoices"])
        self.assertEqual(res["totalRowsAdded"], 1)

        # Verify Details of Bills Excel
        wb = openpyxl.load_workbook(self.details_excel_path)
        ws = wb["Sheet1"]
        self.assertEqual(ws["A5"].value, "72")
        self.assertEqual(ws["B5"].value, "09-04-2026")
        self.assertEqual(ws["C5"].value, "Nandyala")
        self.assertEqual(ws["D5"].value, "500BB20260710166")
        self.assertEqual(ws["E5"].value, "Brand")
        self.assertEqual(ws["F5"].value, "Coragen")
        self.assertEqual(ws["G5"].value, "Paddy")
        self.assertEqual(ws["H5"].value, "OFM")
        self.assertEqual(ws["I5"].value, "Madhavareddy")
        self.assertEqual(ws["J5"].value, "Kiran Kumar")
        self.assertEqual(ws["K5"].value, 1) # 1 Activity event (NOT 25 farmers)
        self.assertEqual(ws["L5"].value, 4500)
        self.assertEqual(ws["M5"].value, 4725.0) # 4500 * 1.05
        self.assertEqual(ws["N5"].value, 5575.5) # 4725 * 1.18
        self.assertEqual(ws["O5"].value, 5576.0) # Grand Total
        self.assertEqual(ws["P5"].value, "24-05-2026") # 09-04-2026 + 45 days
        
        # Verify updated row 3 formulas
        self.assertEqual(ws["L3"].value, "=SUM(L5:L5)")
        self.assertEqual(ws["M3"].value, "=SUM(M5:M5)")
        self.assertEqual(ws["N3"].value, "=SUM(N5:N5)")
        self.assertEqual(ws["O3"].value, "=SUM(O5:O5)")
        wb.close()

        # Verify Budget Cards Excel was updated with IV NO and DATE
        wb_cards = openpyxl.load_workbook(self.cards_excel_path)
        ws_card = wb_cards["38604-10177"]
        self.assertEqual(ws_card["A8"].value, "72")
        self.assertEqual(ws_card["B8"].value, "09-04-2026")
        wb_cards.close()

    def test_04_deduplication_on_second_run(self):
        # Run again - should skip invoice 72 as already present
        res = details_of_bills_generator.scan_and_append_invoices(
            details_excel_path=str(self.details_excel_path),
            invoices_folder_path=str(self.invoices_dir),
            budget_cards_path=str(self.cards_excel_path),
            financial_year="APRIL 2026 to MARCH 2027"
        )
        self.assertTrue(res["success"])
        self.assertEqual(len(res["appendedInvoices"]), 0)
        self.assertIn("72", res["skippedInvoices"])

    def test_05_append_new_invoice_77(self):
        # Create invoice 77 in the invoices directory (matching screenshot 1)
        inv_77_path = self.invoices_dir / "77.xlsx"
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["A6"] = "SRI BALAJI TRADERS"
        ws1["G15"] = "Invoice no:"
        ws1["I15"] = "SBT26270077"
        ws1["G16"] = "Invoice Date:"
        ws1["I16"] = "01-09-2026"
        ws1["G18"] = "New Gen Po:"
        ws1["I18"] = "500BB2026038608"
        ws1["G19"] = "Product:"
        ws1["I19"] = "Velzo"
        ws1["G20"] = "Crop:"
        ws1["I20"] = "Tomato"
        ws1["G21"] = "AMM"
        ws1["I21"] = "U.Madhava Reddy"
        ws1["G22"] = "AREA"
        ws1["I22"] = "Nandyala"

        # Particulars in Sheet 1
        ws1["A25"] = 1
        ws1["B25"] = "DA Activities Expenses"
        ws1["I25"] = 6
        ws1["J25"] = 6600.00

        ws1["A26"] = 2
        ws1["B26"] = "FD Activities Expenses"
        ws1["I26"] = 2
        ws1["J26"] = 11800.00

        ws1["G41"] = "Grand Total"
        ws1["J41"] = 21712.00
        wb.save(inv_77_path)
        wb.close()

        # Run sync - should append invoice 77 with 2 activity rows
        res = details_of_bills_generator.scan_and_append_invoices(
            details_excel_path=str(self.details_excel_path),
            invoices_folder_path=str(self.invoices_dir),
            budget_cards_path=str(self.cards_excel_path),
            financial_year="APRIL 2026 to MARCH 2027"
        )
        self.assertTrue(res["success"])
        self.assertIn("77", res["appendedInvoices"])
        self.assertEqual(res["totalRowsAdded"], 2)

        # Verify Details of Bills rows for invoice 77 (rows 6 and 7)
        wb_bills = openpyxl.load_workbook(self.details_excel_path)
        ws_bills = wb_bills["Sheet1"]
        self.assertEqual(ws_bills["A6"].value, "77")
        self.assertEqual(ws_bills["B6"].value, "01-09-2026")
        self.assertEqual(ws_bills["D6"].value, "500BB2026038608")
        self.assertEqual(ws_bills["E6"].value, "Brand")
        self.assertEqual(ws_bills["F6"].value, "Velzo")
        self.assertEqual(ws_bills["G6"].value, "Tomato")
        self.assertEqual(ws_bills["H6"].value, "DA")
        self.assertEqual(ws_bills["K6"].value, 6) # DA Qty 6
        self.assertEqual(ws_bills["L6"].value, 6600.0)

        self.assertIn(ws_bills["A7"].value, [None, ""]) # Second row of group is empty in Col A
        self.assertEqual(ws_bills["H7"].value, "FD")
        self.assertEqual(ws_bills["K7"].value, 2) # FD Qty 2
        self.assertEqual(ws_bills["L7"].value, 11800.0)
        self.assertEqual(ws_bills["O7"].value, 21712.0) # Grand Total
        self.assertEqual(ws_bills["P7"].value, "16-10-2026") # 01-09-2026 + 45 days
        wb_bills.close()

if __name__ == "__main__":
    unittest.main()
