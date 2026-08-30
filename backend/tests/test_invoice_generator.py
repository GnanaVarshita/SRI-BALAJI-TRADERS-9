import os
import sys
import unittest
import shutil
from pathlib import Path
import openpyxl

BACKEND_DIR = Path(__file__).resolve().parent.parent
sys.path.append(str(BACKEND_DIR))

import invoice_generator

class TestInvoiceGenerator(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.test_dir = BACKEND_DIR / "test_scratch_invoices"
        cls.test_dir.mkdir(parents=True, exist_ok=True)
        cls.save_dir = cls.test_dir / "generated_invoices"
        cls.save_dir.mkdir(parents=True, exist_ok=True)

        # Create a sample Corteva All-TBMs Summary Excel
        cls.corteva_summary_path = cls.test_dir / "Corteva-All-TBMs-Summary.xlsx"
        cls._create_sample_corteva_summary(cls.corteva_summary_path)

        # Create a sample FMC All-TBMs Summary Excel
        cls.fmc_summary_path = cls.test_dir / "FMC-All-TBMs-Summary.xlsx"
        cls._create_sample_fmc_summary(cls.fmc_summary_path)

    @classmethod
    def tearDownClass(cls):
        if cls.test_dir.exists():
            shutil.rmtree(cls.test_dir, ignore_errors=True)

    def setUp(self):
        for f in self.save_dir.glob("*.xlsx"):
            try:
                f.unlink()
            except Exception:
                pass

    @classmethod
    def _create_sample_corteva_summary(cls, path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Table 1: LFM Activity by Manish under PO 4800108503
        ws.cell(1, 1, "MARKETING ACTIVITIES EXPENSES-MANISH-NELLORE")
        headers = ["SI No", "Date", "ZDGM", "TBM", "MDO", "Territory", "Product", "Crop", "Activity", "Village",
                   "No. of Farmers", "Tent/Hall/Chairs Suppliers Charges", "Food Expenses", "Transport", "Others/Gifts", "Total", "PO Number"]
        for c, h in enumerate(headers, 1):
            ws.cell(2, c, h)
        
        ws.cell(3, 1, 1)
        ws.cell(3, 2, "08-09-2026")
        ws.cell(3, 3, "SubbaRamiReddy")
        ws.cell(3, 4, "V Manish")
        ws.cell(3, 5, "Salman")
        ws.cell(3, 6, "Alluru")
        ws.cell(3, 7, "Herbicides")
        ws.cell(3, 8, "Paddy")
        ws.cell(3, 9, "LFM")
        ws.cell(3, 10, "Purni")
        ws.cell(3, 11, 105)
        ws.cell(3, 12, 2500)
        ws.cell(3, 13, 8925)
        ws.cell(3, 14, 1000)
        ws.cell(3, 15, "")
        ws.cell(3, 16, 12425)
        ws.cell(3, 17, "4800108503")

        ws.cell(4, 15, "TOTAL")
        ws.cell(4, 16, 12425)

        # Table 2: Field Days Activity by Manish under PO 4800108503
        ws.cell(6, 1, "MARKETING ACTIVITIES EXPENSES-MANISH-NELLORE")
        for c, h in enumerate(headers, 1):
            ws.cell(7, c, h)
        
        ws.cell(8, 1, 1)
        ws.cell(8, 2, "08-04-2026")
        ws.cell(8, 3, "SubbaRamiReddy")
        ws.cell(8, 4, "V Manish")
        ws.cell(8, 5, "Salman")
        ws.cell(8, 6, "Alluru")
        ws.cell(8, 7, "Herbicides")
        ws.cell(8, 8, "Paddy")
        ws.cell(8, 9, "FD")
        ws.cell(8, 10, "North Mopur")
        ws.cell(8, 11, 45)
        ws.cell(8, 12, 2100)
        ws.cell(8, 13, 3100)
        ws.cell(8, 14, 400)
        ws.cell(8, 15, "")
        ws.cell(8, 16, 5600)
        ws.cell(8, 17, "4800108503")

        ws.cell(9, 1, 2)
        ws.cell(9, 2, "08-05-2026")
        ws.cell(9, 3, "SubbaRamiReddy")
        ws.cell(9, 4, "V Manish")
        ws.cell(9, 5, "Tirumala")
        ws.cell(9, 6, "Rajupalem")
        ws.cell(9, 7, "Herbicides")
        ws.cell(9, 8, "Paddy")
        ws.cell(9, 9, "FD")
        ws.cell(9, 10, "Yellayapalem")
        ws.cell(9, 11, 135)
        ws.cell(9, 12, 3100)
        ws.cell(9, 13, 12825)
        ws.cell(9, 14, 1000)
        ws.cell(9, 15, "")
        ws.cell(9, 16, 16925)
        ws.cell(9, 17, "4800108503")

        ws.cell(10, 15, "TOTAL")
        ws.cell(10, 16, 22525)

        wb.save(path)
        wb.close()

    @classmethod
    def _create_sample_fmc_summary(cls, path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Table 1: Nellore Territory OFM under PO 500BB20260710166
        ws.cell(1, 1, "MARKETING ACTIVITIES EXPENSES-NELLORE")
        headers = ["SI No", "Date", "Area Manager", "TBM/SC/SO", "MDO", "Territory", "Product", "Crop", "Activity", "Village",
                   "No. of Farmers", "Tent/Hall Suppliers Charges", "Food Expenses", "Transport", "Others/Gifts", "Total", "PO Number"]
        for c, h in enumerate(headers, 1):
            ws.cell(2, c, h)

        ws.cell(3, 1, 1)
        ws.cell(3, 2, "22-07-2026")
        ws.cell(3, 3, "U.Madhavareddy")
        ws.cell(3, 4, "S.Samara Simha Reddy")
        ws.cell(3, 5, "G.Sumanth")
        ws.cell(3, 6, "Nellore")
        ws.cell(3, 7, "Resonex")
        ws.cell(3, 8, "Paddy")
        ws.cell(3, 9, "OFM")
        ws.cell(3, 10, "Ravur")
        ws.cell(3, 11, 25)
        ws.cell(3, 12, 400)
        ws.cell(3, 13, 500)
        ws.cell(3, 14, 400)
        ws.cell(3, 15, "")
        ws.cell(3, 16, 1300)
        ws.cell(3, 17, "500BB20260710166")

        ws.cell(4, 1, 2)
        ws.cell(4, 2, "21-07-2026")
        ws.cell(4, 3, "U.Madhavareddy")
        ws.cell(4, 4, "S.Samara Simha Reddy")
        ws.cell(4, 5, "M.siva")
        ws.cell(4, 6, "Nellore")
        ws.cell(4, 7, "Resonex")
        ws.cell(4, 8, "Paddy")
        ws.cell(4, 9, "OFM")
        ws.cell(4, 10, "Nagaleyturu")
        ws.cell(4, 11, 25)
        ws.cell(4, 12, 300)
        ws.cell(4, 13, 600)
        ws.cell(4, 14, 400)
        ws.cell(4, 15, "")
        ws.cell(4, 16, 1300)
        ws.cell(4, 17, "500BB20260710166")

        ws.cell(5, 15, "Total")
        ws.cell(5, 16, 2600)

        # Table 2: Kavali Territory OFM under PO 500BB20260710166
        ws.cell(7, 1, "MARKETING ACTIVITIES EXPENSES-KAVALI")
        for c, h in enumerate(headers, 1):
            ws.cell(8, c, h)

        ws.cell(9, 1, 1)
        ws.cell(9, 2, "04-07-2026")
        ws.cell(9, 3, "U.Madhava Reddy")
        ws.cell(9, 4, "G.Surendrababu")
        ws.cell(9, 5, "J.Ramakrishna")
        ws.cell(9, 6, "Kavali")
        ws.cell(9, 7, "Resonex")
        ws.cell(9, 8, "Paddy")
        ws.cell(9, 9, "OFM")
        ws.cell(9, 10, "North Mopuru")
        ws.cell(9, 11, 30)
        ws.cell(9, 12, 350)
        ws.cell(9, 13, 300)
        ws.cell(9, 14, "")
        ws.cell(9, 15, "")
        ws.cell(9, 16, 650)
        ws.cell(9, 17, "500BB20260710166")

        ws.cell(10, 1, 2)
        ws.cell(10, 2, "09-07-2026")
        ws.cell(10, 3, "U.Madhava Reddy")
        ws.cell(10, 4, "G.Surendrababu")
        ws.cell(10, 5, "V.Rambabu")
        ws.cell(10, 6, "Kavali")
        ws.cell(10, 7, "Resonex")
        ws.cell(10, 8, "Paddy")
        ws.cell(10, 9, "OFM")
        ws.cell(10, 10, "SVB kandriga")
        ws.cell(10, 11, 30)
        ws.cell(10, 12, 500)
        ws.cell(10, 13, 400)
        ws.cell(10, 14, "")
        ws.cell(10, 15, "")
        ws.cell(10, 16, 900)
        ws.cell(10, 17, "500BB20260710166")

        ws.cell(11, 15, "Total")
        ws.cell(11, 16, 1550)

        wb.save(path)
        wb.close()

    def test_indian_words_converter(self):
        self.assertEqual(invoice_generator.num_to_indian_words(126843), "One Lakh Twenty Six Thousand Eight Hundred and Forty Three Rupees Only")
        self.assertEqual(invoice_generator.num_to_indian_words(8989), "Eight Thousand Nine Hundred and Eighty Nine Rupees Only")
        self.assertEqual(invoice_generator.num_to_indian_words(8989.50), "Eight Thousand Nine Hundred and Eighty Nine Rupees and Fifty Paise Only")
        self.assertEqual(invoice_generator.num_to_indian_words(0), "Zero Rupees Only")

    def test_scan_pos_in_summary(self):
        # Create a workbook with 15 and 16-character POs
        test_wb_path = self.test_dir / "Test_POs_Summary.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["Q2"] = "PO"
        ws["Q3"] = "500BB2026018404"
        ws["Q4"] = "500BB20260710172"
        ws["Q5"] = "500BB20260710177"
        ws["Q6"] = "4800108503"
        wb.save(test_wb_path)
        wb.close()

        pos = invoice_generator.scan_pos_in_summary(str(test_wb_path))
        self.assertIn("500BB2026018404", pos)
        self.assertIn("500BB20260710172", pos)
        self.assertIn("500BB20260710177", pos)
        self.assertIn("4800108503", pos)
        # Ensure no truncated versions
        self.assertNotIn("500BB2026071017", pos)

    def test_corteva_invoice_generation(self):
        res = invoice_generator.generate_or_update_invoice(
            company="Corteva",
            tbm_summary_path=str(self.corteva_summary_path),
            save_folder_path=str(self.save_dir),
            invoice_number="SBT26270067",
            po_number="4800108503",
            service_charge_pct=5.0,
            invoice_date="20-08-2026",
            po_value=250000,
            requester_name="R.Bhaskar",
            area="Suryapet"
        )
        self.assertTrue(res["success"])
        self.assertEqual(res["invoiceNo"], "SBT26270067")
        self.assertEqual(res["shortInvoiceNo"], "67")
        self.assertEqual(res["poNumber"], "4800108503")
        self.assertEqual(res["area"], "Suryapet")
        
        # Verify generated Excel
        out_file = Path(res["outputPath"])
        self.assertTrue(out_file.exists())
        wb = openpyxl.load_workbook(out_file)
        self.assertIn("Sheet1", wb.sheetnames)
        self.assertIn("Sheet2", wb.sheetnames)
        self.assertIn("Sheet4", wb.sheetnames)

        # Check Sheet 1 header, area and metadata (Starts at Row 6 after 5 rows top gap)
        ws1 = wb["Sheet1"]
        self.assertEqual(ws1["A6"].value, "SRI BALAJI TRADERS")
        self.assertEqual(ws1["D6"].value, "SBT")
        self.assertEqual(ws1["I15"].value, "SBT26270067")
        self.assertEqual(ws1["I18"].value, "4800108503")
        # Check Sheet 1 ORIGINAL header and metadata
        self.assertEqual(ws1["I12"].value, "ORIGINAL")
        self.assertEqual(ws1["G21"].value, "AREA")
        self.assertEqual(ws1["I21"].value, "Suryapet")
        self.assertEqual(ws1["G22"].value, "ZDGM")
        self.assertEqual(ws1["I22"].value, "R.Bhaskar")

        # Check Bank details and signature block directly attached below words (Row 39 words, Rows 40-44 bank details)
        self.assertEqual(ws1["A40"].value, "Bank details:")
        self.assertEqual(ws1["A41"].value, "Bank name: Karnataka Bank")
        self.assertEqual(ws1["G41"].value, "For Sri Balaji Traders")
        self.assertEqual(ws1["B42"].value, ":6187000600001901")
        self.assertEqual(ws1["B43"].value, ":KARB0000618")
        self.assertEqual(ws1["B44"].value, ":Proddatur")
        self.assertEqual(ws1["G44"].value, "Authorised Signatory")

        # Check DUPLICATE invoice block in the same sheet (Starts at Row 51 after 5 rows gap)
        self.assertEqual(ws1["I57"].value, "DUPLICATE")
        self.assertEqual(ws1["A51"].value, "SRI BALAJI TRADERS")
        self.assertEqual(ws1["A85"].value, "Bank details:")
        self.assertEqual(ws1["G89"].value, "Authorised Signatory")
        self.assertTrue(len(ws1.row_breaks) >= 1)

        # Check Sheet 2 IV subheader & PO Number column in Col Q (17)
        ws2 = wb["Sheet2"]
        self.assertIn("IV NO : 67", ws2["A1"].value)
        self.assertEqual(ws2["Q3"].value, "PO Number")
        self.assertEqual(ws2["Q4"].value, "4800108503")

        # Check Sheet 4 summary
        ws4 = wb["Sheet4"]
        self.assertEqual(ws4["A2"].value, "80141626")
        self.assertEqual(ws4["E2"].value, "SBT26270067")
        self.assertEqual(ws4["I2"].value, "4800108503")
        self.assertEqual(ws4["G2"].value, "=Sheet1!J35")
        self.assertEqual(ws4["H2"].value, "=Sheet1!J38")

        wb.close()

    def test_fmc_invoice_generation(self):
        res = invoice_generator.generate_or_update_invoice(
            company="FMC",
            tbm_summary_path=str(self.fmc_summary_path),
            save_folder_path=str(self.save_dir),
            invoice_number="SBT26270072",
            po_number="500BB20260710166",
            service_charge_pct=4.5,
            invoice_date="22-08-2026",
            requester_name="Madhavareddy",
            area="Nandyala"
        )
        self.assertTrue(res["success"])
        self.assertEqual(res["invoiceNo"], "SBT26270072")
        self.assertEqual(res["shortInvoiceNo"], "72")
        self.assertEqual(res["poNumber"], "500BB20260710166")
        self.assertEqual(res["area"], "Nandyala")

        # Verify generated Excel
        out_file = Path(res["outputPath"])
        self.assertTrue(out_file.exists())
        wb = openpyxl.load_workbook(out_file)
        self.assertIn("Sheet1", wb.sheetnames)
        self.assertIn("Sheet2", wb.sheetnames)
        self.assertNotIn("Sheet4", wb.sheetnames) # FMC has no Sheet4

        # Check Sheet 1 FMC customer, area & service charges (Row 35-36)
        ws1 = wb["Sheet1"]
        self.assertEqual(ws1["I12"].value, "ORIGINAL")
        self.assertEqual(ws1["A15"].value, "New Gen Crop Solutions Pvt. Ltd.")
        self.assertEqual(ws1["G18"].value, "New Gen Po:")
        self.assertEqual(ws1["I18"].value, "500BB20260710166")
        self.assertEqual(ws1["G21"].value, "AMM")
        self.assertEqual(ws1["I21"].value, "Madhavareddy")
        self.assertEqual(ws1["G22"].value, "AREA")
        self.assertEqual(ws1["I22"].value, "Nandyala")
        self.assertEqual(ws1["B35"].value, "Service")
        self.assertEqual(ws1["B36"].value, "Charges")
        self.assertEqual(ws1["E35"].value, "4.50%")

        # Check Bank details at rows 43-47 directly attached below words (Row 42 words, Rows 43-47 bank details)
        self.assertEqual(ws1["A43"].value, "Bank details:")
        self.assertEqual(ws1["A44"].value, "Bank name: Karnataka Bank")
        self.assertEqual(ws1["G44"].value, "For Sri Balaji Traders")
        self.assertEqual(ws1["B45"].value, ":6187000600001901")
        self.assertEqual(ws1["B46"].value, ":KARB0000618")
        self.assertEqual(ws1["B47"].value, ":Proddatur")
        self.assertEqual(ws1["G47"].value, "Authorised Signatory")

        # Check DUPLICATE invoice block in the same sheet for FMC (Starts at Row 54 after 5 rows gap)
        self.assertEqual(ws1["I60"].value, "DUPLICATE")
        self.assertEqual(ws1["A54"].value, "SRI BALAJI TRADERS")
        self.assertEqual(ws1["A91"].value, "Bank details:")
        self.assertEqual(ws1["G95"].value, "Authorised Signatory")
        self.assertTrue(len(ws1.row_breaks) >= 1)

        # Check Sheet 2 FMC header & territory tables & PO Number column in Col Q (17)
        ws2 = wb["Sheet2"]
        self.assertEqual(ws2["A1"].value, "IV NO : 72")
        self.assertEqual(ws2["Q3"].value, "PO Number")
        self.assertEqual(ws2["Q4"].value, "500BB20260710166")

        wb.close()

    def test_incremental_append_update(self):
        # First generate FMC invoice 72
        res1 = invoice_generator.generate_or_update_invoice(
            company="FMC",
            tbm_summary_path=str(self.fmc_summary_path),
            save_folder_path=str(self.save_dir),
            invoice_number="72",
            po_number="500BB20260710166",
            service_charge_pct=4.5
        )
        self.assertTrue(res1["success"])
        self.assertFalse(res1["isUpdate"])

        # Run update again for invoice 72
        res2 = invoice_generator.generate_or_update_invoice(
            company="FMC",
            tbm_summary_path=str(self.fmc_summary_path),
            save_folder_path=str(self.save_dir),
            invoice_number="72",
            po_number="500BB20260710166",
            service_charge_pct=4.5
        )
        self.assertTrue(res2["success"])
        self.assertTrue(res2["isUpdate"])
        self.assertIn("updated & appended", res2["message"])

if __name__ == "__main__":
    unittest.main()
