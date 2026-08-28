import os
import re
import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import xlrd
except ImportError:
    xlrd = None

def clean_str(val):
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in ["none", "null", "nan"]:
        return ""
    return s

def format_date_val(val):
    if val is None:
        return ""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%d/%m/%Y")
    s = str(val).strip()
    if not s:
        return ""
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d\\%m\\%Y", "%d/%m/%y", "%d-%m-%y"]:
        try:
            dt = datetime.datetime.strptime(s, fmt)
            return dt.strftime("%d/%m/%Y")
        except ValueError:
            pass
    return s

def format_date_xlrd(val, datemode):
    if val is None or val == "":
        return ""
    if isinstance(val, float):
        try:
            dt = xlrd.xldate_as_datetime(val, datemode)
            return dt.strftime("%d/%m/%Y")
        except Exception:
            pass
    return format_date_val(val)

def parse_num(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0

FIELD_KEYWORDS = {
    'sl_no': ['sl no', 'sl.no', 's no', 's.no', 'sno', 'si no'],
    'date': ['date'],
    'zdgm': ['zdgm', 'zdgl', 'area manager', 'adgl', 'adg'],
    'tbm': ['tbm name', 'tbm', 'name of the tbm'],
    'mdo': ['mdo'],
    'territory': ['territory', 'tbm territory', 'area', 'place'],
    'product': ['product', 'item'],
    'crop': ['crop'],
    'activity': ['activity name', 'activity'],
    'village': ['village name', 'village', 'villages'],
    'farmers': ['no.of farmers', 'no of farmers', 'no. of farmers', 'n0 of farmers', 'farmers attended', 'farmers'],
    'tent': ['tent/hall/chairs', 'tent/ hall', 'tent', 'chairs', 'chairs/ table/tent', 'suppliers/charges', 'suppliers'],
    'food': ['food expenses', 'food/snacks', 'food', 'expenses'],
    'transport': ['transport', 'trans port', 'auto charges', 'auto'],
    'others': ['others/gifts', 'others / gifts', 'others / gift', 'others', 'gifts', 'saplaires'],
    'total': ['total amount', 'total', 'amount'],
    'po_number': ['po number', 'po.no', 'ponumber', 'po #']
}

import sys
sys.path.append(str(Path(__file__).resolve().parent))
try:
    import tbm_formatter
except ImportError:
    from . import tbm_formatter

def extract_activities_from_sheet(ws, tbm_folder_name, file_name):
    acts = tbm_formatter.extract_activities_from_sheet(ws, default_tbm_name=tbm_folder_name)
    for a in acts:
        a['source_file'] = file_name
    return acts

def extract_activities_from_xlrd_sheet(sh, datemode, tbm_folder_name, file_name):
    """
    Extracts activity records from legacy .xls sheets using xlrd.
    """
    best_row = None
    best_map = {}
    best_score = 0

    max_r = min(sh.nrows, 25)
    for r in range(max_r):
        curr_map = {}
        score = 0
        for c in range(sh.ncols):
            top_txt = clean_str(sh.cell_value(r - 1, c)).lower() if r > 0 else ""
            curr_txt = clean_str(sh.cell_value(r, c)).lower()
            comb_txt = f"{top_txt} {curr_txt}".strip()
            if not comb_txt:
                continue

            for field, kw_list in FIELD_KEYWORDS.items():
                if field not in curr_map:
                    if any(kw in comb_txt for kw in kw_list):
                        curr_map[field] = c
                        score += 1
                        break
        if score > best_score and score >= 3:
            best_score = score
            best_row = r
            best_map = curr_map

    if best_row is None:
        return []

    sheet_po_number = ""
    for r in range(min(10, sh.nrows)):
        for c in range(min(20, sh.ncols)):
            val = clean_str(sh.cell_value(r, c))
            m = re.search(r'5\d{2}[A-Z0-9]{8,12}', val, re.I)
            if m:
                sheet_po_number = m.group(0).upper()
                break
        if sheet_po_number:
            break

    activities = []
    for r in range(best_row + 1, sh.nrows):
        row_vals = [sh.cell_value(r, c) for c in range(sh.ncols)]
        if not any(row_vals):
            continue

        c1_val = clean_str(row_vals[0]).lower()
        c2_val = clean_str(row_vals[1]).lower() if len(row_vals) > 1 else ""
        date_col_idx = best_map.get('date', 2)
        date_c_val = clean_str(sh.cell_value(r, date_col_idx)).lower() if date_col_idx < len(row_vals) else ""

        if c1_val == 'total' or c2_val == 'total' or date_c_val == 'total':
            break

        date_val = format_date_xlrd(sh.cell_value(r, date_col_idx), datemode) if date_col_idx < len(row_vals) else ""
        
        tbm_idx = best_map.get('tbm', 0)
        tbm_val = clean_str(sh.cell_value(r, tbm_idx)) if tbm_idx < len(row_vals) else ""
        if not tbm_val:
            tbm_val = tbm_folder_name

        zdgm_idx = best_map.get('zdgm', 0)
        zdgm_val = clean_str(sh.cell_value(r, zdgm_idx)) if zdgm_idx < len(row_vals) else ""

        mdo_idx = best_map.get('mdo', 0)
        mdo_val = clean_str(sh.cell_value(r, mdo_idx)) if mdo_idx < len(row_vals) else ""

        terr_idx = best_map.get('territory', 0)
        territory_val = clean_str(sh.cell_value(r, terr_idx)) if terr_idx < len(row_vals) else ""

        prod_idx = best_map.get('product', 0)
        product_val = clean_str(sh.cell_value(r, prod_idx)) if prod_idx < len(row_vals) else ""

        crop_idx = best_map.get('crop', 0)
        crop_val = clean_str(sh.cell_value(r, crop_idx)) if crop_idx < len(row_vals) else ""

        act_idx = best_map.get('activity', 0)
        activity_val = clean_str(sh.cell_value(r, act_idx)) if act_idx < len(row_vals) else ""

        vlg_idx = best_map.get('village', 0)
        village_val = clean_str(sh.cell_value(r, vlg_idx)) if vlg_idx < len(row_vals) else ""

        farm_idx = best_map.get('farmers', 0)
        farmers_val = parse_num(sh.cell_value(r, farm_idx)) if farm_idx < len(row_vals) else 0.0

        tent_idx = best_map.get('tent', 0)
        tent_val = parse_num(sh.cell_value(r, tent_idx)) if tent_idx < len(row_vals) else 0.0

        food_idx = best_map.get('food', 0)
        food_val = parse_num(sh.cell_value(r, food_idx)) if food_idx < len(row_vals) else 0.0

        trans_idx = best_map.get('transport', 0)
        transport_val = parse_num(sh.cell_value(r, trans_idx)) if trans_idx < len(row_vals) else 0.0

        oth_idx = best_map.get('others', 0)
        others_val = parse_num(sh.cell_value(r, oth_idx)) if oth_idx < len(row_vals) else 0.0

        tot_idx = best_map.get('total', 0)
        total_val = parse_num(sh.cell_value(r, tot_idx)) if tot_idx < len(row_vals) else 0.0
        calc_total = tent_val + food_val + transport_val + others_val
        if total_val == 0.0 and calc_total > 0.0:
            total_val = calc_total

        po_idx = best_map.get('po_number', 0)
        po_val = clean_str(sh.cell_value(r, po_idx)) if po_idx < len(row_vals) else ""
        if not po_val and sheet_po_number:
            po_val = sheet_po_number

        if not activity_val and not product_val and total_val == 0.0 and not date_val:
            continue

        activities.append({
            'date': date_val,
            'zdgm': zdgm_val,
            'tbm': tbm_val,
            'mdo': mdo_val,
            'territory': territory_val,
            'product': product_val,
            'crop': crop_val,
            'activity': activity_val,
            'village': village_val,
            'farmers': int(farmers_val),
            'tent': tent_val,
            'food': food_val,
            'transport': transport_val,
            'others': others_val,
            'total': total_val,
            'po_number': po_val,
            'source_file': file_name
        })

    return activities

def parse_slip_format(ws, tbm_folder_name):
    kv = {}
    for r in range(1, 20):
        k = clean_str(ws.cell(r, 2).value).lower()
        v = ws.cell(r, 4).value
        if k:
            if 'date' in k: kv['date'] = format_date_val(v)
            elif 'product' in k: kv['product'] = clean_str(v)
            elif 'crop' in k: kv['crop'] = clean_str(v)
            elif 'activity' in k: kv['activity'] = clean_str(v)
            elif 'place' in k or 'territory' in k: kv['territory'] = clean_str(v)
            elif 'farmers' in k: kv['farmers'] = int(parse_num(v))
            elif 'food' in k: kv['food'] = parse_num(v)
            elif 'chairs' in k or 'tent' in k: kv['tent'] = parse_num(v)
            elif 'auto' in k or 'transport' in k: kv['transport'] = parse_num(v)
            elif 'others' in k: kv['others'] = parse_num(v)
            elif 'tbm' in k: kv['tbm'] = clean_str(v)

    if kv.get('activity') or kv.get('product') or kv.get('date'):
        tbm_val = kv.get('tbm') or tbm_folder_name
        tent = kv.get('tent', 0.0)
        food = kv.get('food', 0.0)
        transport = kv.get('transport', 0.0)
        others = kv.get('others', 0.0)
        total = tent + food + transport + others
        return {
            'date': kv.get('date', ''),
            'zdgm': '',
            'tbm': tbm_val,
            'mdo': '',
            'territory': kv.get('territory', ''),
            'product': kv.get('product', ''),
            'crop': kv.get('crop', ''),
            'activity': kv.get('activity', ''),
            'village': '',
            'farmers': kv.get('farmers', 0),
            'tent': tent,
            'food': food,
            'transport': transport,
            'others': others,
            'total': total,
            'po_number': '',
            'source_file': ws.title
        }
    return None

def create_table_styles():
    thin_border = Border(
        left=Side(style='thin', color='A6A6A6'),
        right=Side(style='thin', color='A6A6A6'),
        top=Side(style='thin', color='A6A6A6'),
        bottom=Side(style='thin', color='A6A6A6')
    )

    total_border = Border(
        left=Side(style='thin', color='A6A6A6'),
        right=Side(style='thin', color='A6A6A6'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )

    summary_border = Border(
        left=Side(style='medium', color='702000'),
        right=Side(style='medium', color='702000'),
        top=Side(style='medium', color='702000'),
        bottom=Side(style='double', color='702000')
    )

    subtotal_fill = PatternFill(start_color='F2DCDB', end_color='F2DCDB', fill_type='solid')
    po_total_fill = PatternFill(start_color='DDD9C4', end_color='DDD9C4', fill_type='solid')

    brown_color = '702000'

    title_font = Font(name='Calibri', size=11, bold=True, color=brown_color)
    hdr_font = Font(name='Calibri', size=10, bold=True, color=brown_color)
    brown_text_font = Font(name='Calibri', size=10, color=brown_color)
    date_font = Font(name='Calibri', size=10, color=brown_color)
    regular_font = Font(name='Calibri', size=10)
    bold_font = Font(name='Calibri', size=10, bold=True)
    po_font = Font(name='Calibri', size=10, bold=True, color=brown_color)
    subtotal_font = Font(name='Calibri', size=10, bold=True, color=brown_color)
    po_grand_font = Font(name='Calibri', size=11, bold=True, color=brown_color)

    return {
        'thin_border': thin_border,
        'total_border': total_border,
        'summary_border': summary_border,
        'subtotal_fill': subtotal_fill,
        'po_total_fill': po_total_fill,
        'title_font': title_font,
        'hdr_font': hdr_font,
        'brown_text_font': brown_text_font,
        'date_font': date_font,
        'regular_font': regular_font,
        'bold_font': bold_font,
        'po_font': po_font,
        'subtotal_font': subtotal_font,
        'po_grand_font': po_grand_font
    }

def write_table_to_sheet(ws, start_row, tbm_name, territory, rows_data, po_number, styles):
    """
    Writes a single formatted TBM activity table to worksheet at start_row matching screenshot layout.
    Returns (next_available_row, total_cell_row_idx).
    """
    title_row = start_row
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=16)
    title_cell = ws.cell(title_row, 1)
    terr_str = f"-{territory.upper()}" if territory else ""
    title_cell.value = f"MARKETING ACTIVITIES EXPENSES-{tbm_name.upper()}{terr_str}"
    title_cell.font = styles['title_font']
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[title_row].height = 24

    for c in range(1, 18):
        ws.cell(title_row, c).border = styles['thin_border']

    hdr_row = title_row + 1
    headers = [
        "SI No", "Date", "ZDGM", "TBM", "MDO", "Territory", "Product", "Crop", "Activity", "Village",
        "No. of Farmers", "Tent/Hall/Chairs Suppliers Charges", "Food Expenses", "Transport", "Others/Gifts", "Total", "PO Number"
    ]
    ws.row_dimensions[hdr_row].height = 32
    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws.cell(hdr_row, col_idx)
        cell.value = h_text
        cell.font = styles['hdr_font']
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = styles['thin_border']

    current_r = hdr_row + 1
    for idx, item in enumerate(rows_data, start=1):
        ws.row_dimensions[current_r].height = 20

        c_sno = ws.cell(current_r, 1, idx)
        c_sno.font = styles['bold_font']
        c_sno.alignment = Alignment(horizontal='center', vertical='center')

        c_date = ws.cell(current_r, 2, item.get('date', ''))
        c_date.font = styles['date_font']
        c_date.alignment = Alignment(horizontal='center', vertical='center')

        for c_idx, key in [(3, 'zdgm'), (4, 'tbm'), (5, 'mdo'), (6, 'territory'), (7, 'product'), (8, 'crop'), (9, 'activity'), (10, 'village')]:
            val = item.get(key, '')
            if key == 'tbm' and not val:
                val = tbm_name
            if key == 'territory' and not val:
                val = territory
            c_txt = ws.cell(current_r, c_idx, val)
            c_txt.font = styles['brown_text_font']
            c_txt.alignment = Alignment(horizontal='left' if c_idx in [3,4,5,10] else 'center', vertical='center')

        c_farmers = ws.cell(current_r, 11, item.get('farmers', 0))
        c_farmers.font = styles['regular_font']
        c_farmers.alignment = Alignment(horizontal='center', vertical='center')
        c_farmers.number_format = '#,##0'

        for c_offset, key in enumerate(['tent', 'food', 'transport', 'others'], start=12):
            amt_val = item.get(key, 0.0)
            c_amt = ws.cell(current_r, c_offset, amt_val if amt_val > 0 else "")
            c_amt.font = styles['regular_font']
            c_amt.alignment = Alignment(horizontal='right', vertical='center')
            if amt_val > 0:
                c_amt.number_format = '#,##0'

        c_tot = ws.cell(current_r, 16, f"=SUM(L{current_r}:O{current_r})")
        c_tot.font = styles['regular_font']
        c_tot.alignment = Alignment(horizontal='right', vertical='center')
        c_tot.number_format = '#,##0'

        row_po = item.get('po_number') or po_number or ''
        c_po = ws.cell(current_r, 17, row_po)
        c_po.font = styles['po_font']
        c_po.alignment = Alignment(horizontal='center', vertical='center')

        for c in range(1, 18):
            ws.cell(current_r, c).border = styles['thin_border']

        current_r += 1

    po_extra_row = current_r
    ws.row_dimensions[po_extra_row].height = 18
    for c in range(1, 18):
        ws.cell(po_extra_row, c).border = styles['thin_border']
    c_po_last = ws.cell(po_extra_row, 17, po_number or "")
    c_po_last.font = styles['po_font']
    c_po_last.alignment = Alignment(horizontal='center', vertical='center')
    current_r += 1

    tot_row = current_r
    ws.row_dimensions[tot_row].height = 22
    for c in range(1, 18):
        ws.cell(tot_row, c).border = styles['total_border']

    lbl_cell = ws.cell(tot_row, 15, "TOTAL")
    lbl_cell.font = styles['bold_font']
    lbl_cell.alignment = Alignment(horizontal='right', vertical='center')

    data_start_r = hdr_row + 1
    data_end_r = po_extra_row - 1

    sum_tot_cell = ws.cell(tot_row, 16, f"=SUM(P{data_start_r}:P{data_end_r})")
    sum_tot_cell.font = styles['bold_font']
    sum_tot_cell.alignment = Alignment(horizontal='right', vertical='center')
    sum_tot_cell.number_format = '#,##0'

    return tot_row + 3, tot_row

def generate_tbm_summary(tbm_folder_path, output_path=None, priority_po_list=None):
    """
    Main function to scan TBM summary folder, parse activities (supporting .xlsx and legacy .xls files),
    group by PO, Activity, TBM, format into max 9 tables per sheet, with Activity Subtotals and PO Grand Totals.
    """
    tbm_dir = Path(tbm_folder_path).resolve()
    if not tbm_dir.exists() or not tbm_dir.is_dir():
        raise ValueError(f"TBM Summary folder path does not exist: {tbm_folder_path}")

    if not output_path:
        territory_name = "Nandyala"
        for p in tbm_dir.parts:
            p_up = p.upper()
            if any(k in p_up for k in ['KURNOOL', 'NELLORE', 'NANDYAL', 'NANDYALA', 'SURYAPET']):
                territory_name = p.title()
                break
        if territory_name == "Nandyala" and tbm_dir.name and tbm_dir.name != "TBM s Summary":
            territory_name = tbm_dir.name.replace(" ", "-")

        output_filename = f"{territory_name}-All-TBMs-Summary.xlsx"
        output_path = tbm_dir / output_filename
    else:
        output_path = Path(output_path).resolve()

    priority_pos = set()
    if priority_po_list:
        if isinstance(priority_po_list, str):
            tokens = re.split(r'[\s,\n\r]+', priority_po_list)
        else:
            tokens = priority_po_list
        for tok in tokens:
            cleaned = str(tok).strip().upper()
            if cleaned:
                priority_pos.add(cleaned)

    all_extracted_rows = []
    tbm_subfolders = [d for d in tbm_dir.iterdir() if d.is_dir()]
    if not tbm_subfolders:
        tbm_subfolders = [tbm_dir]

    for subfolder in tbm_subfolders:
        tbm_name = subfolder.name if subfolder != tbm_dir else "TBM"
        # Find all files in subfolder (accepting .xlsx, .xls, .xlsm, .xlsb, .csv and files without extension)
        excel_files = [
            f for f in subfolder.rglob("*")
            if f.is_file() and (f.suffix.lower() in ['.xlsx', '.xls', '.xlsm', '.xlsb', '.csv', ''] or 'excel' in f.name.lower() or 'report' in f.name.lower() or 'activity' in f.name.lower())
        ]

        for ef in excel_files:
            if ef.resolve() == output_path.resolve():
                continue
            if "-All-TBMs-Summary" in ef.name:
                continue

            suf = ef.suffix.lower()

            # 1. Standard .xlsx and .xlsm files (openpyxl)
            if suf in ['.xlsx', '.xlsm']:
                try:
                    wb_in = openpyxl.load_workbook(ef, data_only=True)
                    # Prefer formatted sheet (Sheet 2) if it contains activities; otherwise use Sheet 1
                    target_sheets = []
                    if len(wb_in.sheetnames) >= 2:
                        target_sheets = [wb_in.sheetnames[1], wb_in.sheetnames[0]]
                    else:
                        target_sheets = wb_in.sheetnames

                    extracted_for_file = []
                    for sname in target_sheets:
                        ws_in = wb_in[sname]
                        sheet_activities = extract_activities_from_sheet(ws_in, tbm_name, ef.name)
                        if sheet_activities:
                            extracted_for_file = sheet_activities
                            break
                    
                    all_extracted_rows.extend(extracted_for_file)
                    wb_in.close()
                except Exception as e:
                    print(f"Error reading {suf} file {ef.name}: {e}")

            # 2. Legacy .xls files (xlrd)
            elif suf == '.xls':
                if xlrd is None:
                    print(f"Skipping .xls file {ef.name} because xlrd is not installed")
                    continue
                try:
                    wb_xls = xlrd.open_workbook(ef)
                    for sname in wb_xls.sheet_names():
                        sh_xls = wb_xls.sheet_by_name(sname)
                        sheet_activities = extract_activities_from_xlrd_sheet(sh_xls, wb_xls.datemode, tbm_name, ef.name)
                        all_extracted_rows.extend(sheet_activities)
                except Exception as e:
                    print(f"Error reading .xls file {ef.name}: {e}")

            # 3. CSV files
            elif suf == '.csv':
                try:
                    import csv
                    with open(ef, 'r', encoding='utf-8-sig', errors='replace') as csv_file:
                        reader = list(csv.reader(csv_file))
                        # Treat csv reader list as virtual sheet
                        class VirtualCsvSheet:
                            def __init__(self, data):
                                self.nrows = len(data)
                                self.ncols = max((len(r) for r in data), default=0)
                                self.data = data
                            def cell_value(self, r, c):
                                if r < len(self.data) and c < len(self.data[r]):
                                    return self.data[r][c]
                                return ""
                        v_sheet = VirtualCsvSheet(reader)
                        sheet_activities = extract_activities_from_xlrd_sheet(v_sheet, 0, tbm_name, ef.name)
                        all_extracted_rows.extend(sheet_activities)
                except Exception as e:
                    print(f"Error reading .csv file {ef.name}: {e}")

            # 4. Fallback for files without extension or unusual formats
            else:
                parsed = False
                # Try openpyxl first
                try:
                    wb_in = openpyxl.load_workbook(ef, data_only=True)
                    for sname in wb_in.sheetnames:
                        ws_in = wb_in[sname]
                        sheet_activities = extract_activities_from_sheet(ws_in, tbm_name, ef.name)
                        all_extracted_rows.extend(sheet_activities)
                    wb_in.close()
                    parsed = True
                except Exception:
                    pass

                # Try xlrd second
                if not parsed and xlrd is not None:
                    try:
                        wb_xls = xlrd.open_workbook(ef)
                        for sname in wb_xls.sheet_names():
                            sh_xls = wb_xls.sheet_by_name(sname)
                            sheet_activities = extract_activities_from_xlrd_sheet(sh_xls, wb_xls.datemode, tbm_name, ef.name)
                            all_extracted_rows.extend(sheet_activities)
                    except Exception:
                        pass

    if not all_extracted_rows:
        return {
            "success": False,
            "message": f"No activity records found in Excel files inside {tbm_dir.name}",
            "tables_count": 0,
            "sheets_count": 0
        }

    priority_category = []
    unlisted_category = []
    no_po_category = []

    for row in all_extracted_rows:
        po_num = str(row.get('po_number', '')).strip().upper()
        if not po_num:
            no_po_category.append(row)
        else:
            if priority_pos:
                if po_num in priority_pos:
                    priority_category.append(row)
                else:
                    unlisted_category.append(row)
            else:
                priority_category.append(row)

    def build_hierarchical_groups(rows_list):
        po_dict = {}
        for r in rows_list:
            po_num = str(r.get('po_number', '')).strip().upper() or "NO_PO"
            prod = str(r.get('product', '')).strip().title() or "General Product"
            act = str(r.get('activity', '')).strip().upper() or "GENERAL"
            tbm = str(r.get('tbm', '')).strip().title() or "TBM"

            if po_num not in po_dict:
                po_dict[po_num] = {}
            if act not in po_dict[po_num]:
                po_dict[po_num][act] = {}
            tbm_key = (prod, tbm)
            if tbm_key not in po_dict[po_num][act]:
                po_dict[po_num][act][tbm_key] = []
            po_dict[po_num][act][tbm_key].append(r)
        return po_dict

    styles = create_table_styles()

    if output_path.exists():
        try:
            wb = openpyxl.load_workbook(output_path)
        except Exception:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    def add_po_hierarchy_to_sheet_series(prefix, rows_list):
        if not rows_list:
            return 0

        po_dict = build_hierarchical_groups(rows_list)
        matching_sheets = [s for s in wb.sheetnames if s.startswith(prefix)]
        
        curr_sheet_idx = 1
        if matching_sheets:
            for ms in matching_sheets:
                m = re.search(r'\d+', ms)
                if m:
                    idx = int(m.group(0))
                    if idx > curr_sheet_idx:
                        curr_sheet_idx = idx

        target_sheet_name = f"{prefix}{curr_sheet_idx}"
        if target_sheet_name in wb.sheetnames:
            ws = wb[target_sheet_name]
        else:
            ws = wb.create_sheet(title=target_sheet_name)
            ws.views.sheetView[0].showGridLines = True

        existing_table_count = 0
        last_used_row = 1
        for r in range(1, ws.max_row + 1):
            val = clean_str(ws.cell(r, 1).value)
            if val.startswith("MARKETING ACTIVITIES EXPENSES"):
                existing_table_count += 1
                last_used_row = r

        col_widths = {
            'A': 8, 'B': 12, 'C': 16, 'D': 16, 'E': 14, 'F': 14, 'G': 14, 'H': 12,
            'I': 14, 'J': 14, 'K': 14, 'L': 24, 'M': 15, 'N': 12, 'O': 14, 'P': 14, 'Q': 18
        }
        for col_let, w in col_widths.items():
            ws.column_dimensions[col_let].width = w

        current_row = 1 if last_used_row == 1 and not ws.cell(1, 1).value else (ws.max_row + 3)
        total_tables_created = 0

        for po_num, act_dict in po_dict.items():
            display_po = po_num if po_num != "NO_PO" else ""
            po_subtotal_rows_on_sheet = []

            for act, tbm_dict in act_dict.items():
                act_table_rows_on_sheet = []

                for (prod, tbm_name), items in tbm_dict.items():
                    if existing_table_count >= 9:
                        if act_table_rows_on_sheet:
                            subtot_row = current_row
                            ws.row_dimensions[subtot_row].height = 24
                            ws.merge_cells(start_row=subtot_row, start_column=1, end_row=subtot_row, end_column=15)
                            lbl = ws.cell(subtot_row, 1, f"TOTAL EXPENSES FOR ACTIVITY: {act.upper()}")
                            lbl.font = styles['subtotal_font']
                            lbl.fill = styles['subtotal_fill']
                            lbl.alignment = Alignment(horizontal='right', vertical='center')
                            for c in range(1, 18):
                                ws.cell(subtot_row, c).border = styles['thin_border']
                                if c <= 15: ws.cell(subtot_row, c).fill = styles['subtotal_fill']
                            tot_refs = [f"P{r}" for r in act_table_rows_on_sheet]
                            c_tot = ws.cell(subtot_row, 16, f"=SUM({','.join(tot_refs)})")
                            c_tot.font = styles['subtotal_font']
                            c_tot.fill = styles['subtotal_fill']
                            c_tot.alignment = Alignment(horizontal='right', vertical='center')
                            c_tot.number_format = '#,##0'
                            ws.cell(subtot_row, 17, display_po).font = styles['subtotal_font']
                            ws.cell(subtot_row, 17).fill = styles['subtotal_fill']
                            ws.cell(subtot_row, 17).alignment = Alignment(horizontal='center', vertical='center')
                            po_subtotal_rows_on_sheet.append(subtot_row)
                            act_table_rows_on_sheet = []
                            current_row = subtot_row + 3

                        if po_subtotal_rows_on_sheet:
                            grand_row = current_row
                            ws.row_dimensions[grand_row].height = 26
                            ws.merge_cells(start_row=grand_row, start_column=1, end_row=grand_row, end_column=15)
                            lbl_g = ws.cell(grand_row, 1, f"GRAND TOTAL FOR PO: {display_po}" if display_po else "GRAND TOTAL FOR UNASSIGNED ACTIVITIES")
                            lbl_g.font = styles['po_grand_font']
                            lbl_g.fill = styles['po_total_fill']
                            lbl_g.alignment = Alignment(horizontal='right', vertical='center')
                            for c in range(1, 18):
                                ws.cell(grand_row, c).border = styles['summary_border']
                                if c <= 15: ws.cell(grand_row, c).fill = styles['po_total_fill']
                            po_sub_refs = [f"P{r}" for r in po_subtotal_rows_on_sheet]
                            c_g = ws.cell(grand_row, 16, f"=SUM({','.join(po_sub_refs)})")
                            c_g.font = styles['po_grand_font']
                            c_g.fill = styles['po_total_fill']
                            c_g.alignment = Alignment(horizontal='right', vertical='center')
                            c_g.number_format = '#,##0'
                            ws.cell(grand_row, 17, display_po).font = styles['po_grand_font']
                            ws.cell(grand_row, 17).fill = styles['po_total_fill']
                            ws.cell(grand_row, 17).alignment = Alignment(horizontal='center', vertical='center')
                            po_subtotal_rows_on_sheet = []

                        curr_sheet_idx += 1
                        target_sheet_name = f"{prefix}{curr_sheet_idx}"
                        ws = wb.create_sheet(title=target_sheet_name)
                        ws.views.sheetView[0].showGridLines = True
                        for col_let, w in col_widths.items():
                            ws.column_dimensions[col_let].width = w
                        existing_table_count = 0
                        current_row = 1

                    terr = items[0].get('territory', '') if items else ''
                    current_row, tbl_tot_r = write_table_to_sheet(
                        ws, current_row, tbm_name, terr, items, display_po, styles
                    )
                    act_table_rows_on_sheet.append(tbl_tot_r)
                    existing_table_count += 1
                    total_tables_created += 1

                if act_table_rows_on_sheet:
                    subtot_row = current_row
                    ws.row_dimensions[subtot_row].height = 24
                    ws.merge_cells(start_row=subtot_row, start_column=1, end_row=subtot_row, end_column=15)
                    lbl = ws.cell(subtot_row, 1, f"TOTAL EXPENSES FOR ACTIVITY: {act.upper()}")
                    lbl.font = styles['subtotal_font']
                    lbl.fill = styles['subtotal_fill']
                    lbl.alignment = Alignment(horizontal='right', vertical='center')
                    for c in range(1, 18):
                        ws.cell(subtot_row, c).border = styles['thin_border']
                        if c <= 15: ws.cell(subtot_row, c).fill = styles['subtotal_fill']
                    tot_refs = [f"P{r}" for r in act_table_rows_on_sheet]
                    c_tot = ws.cell(subtot_row, 16, f"=SUM({','.join(tot_refs)})")
                    c_tot.font = styles['subtotal_font']
                    c_tot.fill = styles['subtotal_fill']
                    c_tot.alignment = Alignment(horizontal='right', vertical='center')
                    c_tot.number_format = '#,##0'
                    ws.cell(subtot_row, 17, display_po).font = styles['subtotal_font']
                    ws.cell(subtot_row, 17).fill = styles['subtotal_fill']
                    ws.cell(subtot_row, 17).alignment = Alignment(horizontal='center', vertical='center')
                    po_subtotal_rows_on_sheet.append(subtot_row)
                    current_row = subtot_row + 3

            if po_subtotal_rows_on_sheet:
                grand_row = current_row
                ws.row_dimensions[grand_row].height = 26
                ws.merge_cells(start_row=grand_row, start_column=1, end_row=grand_row, end_column=15)
                lbl_g = ws.cell(grand_row, 1, f"GRAND TOTAL FOR PO: {display_po}" if display_po else "GRAND TOTAL FOR UNASSIGNED ACTIVITIES")
                lbl_g.font = styles['po_grand_font']
                lbl_g.fill = styles['po_total_fill']
                lbl_g.alignment = Alignment(horizontal='right', vertical='center')
                for c in range(1, 18):
                    ws.cell(grand_row, c).border = styles['summary_border']
                    if c <= 15: ws.cell(grand_row, c).fill = styles['po_total_fill']
                po_sub_refs = [f"P{r}" for r in po_subtotal_rows_on_sheet]
                c_g = ws.cell(grand_row, 16, f"=SUM({','.join(po_sub_refs)})")
                c_g.font = styles['po_grand_font']
                c_g.fill = styles['po_total_fill']
                c_g.alignment = Alignment(horizontal='right', vertical='center')
                c_g.number_format = '#,##0'
                ws.cell(grand_row, 17, display_po).font = styles['po_grand_font']
                ws.cell(grand_row, 17).fill = styles['po_total_fill']
                ws.cell(grand_row, 17).alignment = Alignment(horizontal='center', vertical='center')
                current_row = grand_row + 4

        return total_tables_created

    n_pri_tbls = add_po_hierarchy_to_sheet_series("Sheet", priority_category)
    n_unl_tbls = add_po_hierarchy_to_sheet_series("Unlisted POs ", unlisted_category)
    n_nopo_tbls = add_po_hierarchy_to_sheet_series("No PO ", no_po_category)

    create_tbm_amount_summary_sheet(wb, all_extracted_rows, styles)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()

    total_tables = n_pri_tbls + n_unl_tbls + n_nopo_tbls
    return {
        "success": True,
        "message": f"Successfully consolidated {len(all_extracted_rows)} activity records across {total_tables} table(s) with Activity Subtotals, PO Grand Totals, and TBM Amount Summary sheet in {output_path.name}!",
        "outputPath": str(output_path),
        "totalActivities": len(all_extracted_rows),
        "totalTables": total_tables,
        "priorityTables": n_pri_tbls,
        "unlistedTables": n_unl_tbls,
        "noPoTables": n_nopo_tbls,
        "sheetsCount": len(wb.sheetnames)
    }

def create_tbm_amount_summary_sheet(wb, all_extracted_rows, styles):
    """
    Creates a summary worksheet at the end of the workbook listing total amount used by TBMs,
    organized by: PO, TBM, Product, Crop, Activity, No. of Activities, Total Amount.
    """
    sheet_title = "TBM Amount Summary"
    if sheet_title in wb.sheetnames:
        wb.remove(wb[sheet_title])

    ws = wb.create_sheet(title=sheet_title)
    ws.views.sheetView[0].showGridLines = True

    col_widths = {
        'A': 22,  # PO Number
        'B': 26,  # TBM Name
        'C': 18,  # Product
        'D': 16,  # Crop
        'E': 20,  # Activity
        'F': 18,  # No. of Activities
        'G': 20   # Total Amount
    }
    for col_let, w in col_widths.items():
        ws.column_dimensions[col_let].width = w

    summary_map = {}
    for r in all_extracted_rows:
        po_num = str(r.get('po_number', '')).strip().upper() or "NO PO"
        tbm_name = str(r.get('tbm', '')).strip().title() or "TBM"
        prod = str(r.get('product', '')).strip().title() or "General Product"
        crop = str(r.get('crop', '')).strip().title() or "General Crop"
        act = str(r.get('activity', '')).strip().upper() or "GENERAL"
        tot_amt = float(r.get('total', 0.0) or 0.0)

        key = (po_num, tbm_name, prod, crop, act)
        if key not in summary_map:
            summary_map[key] = {
                'count': 0,
                'total_amount': 0.0
            }
        summary_map[key]['count'] += 1
        summary_map[key]['total_amount'] += tot_amt

    # Title Header
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    title_cell = ws.cell(1, 1, "SUMMARY OF EXPENSES & ACTIVITIES BY TBM")
    title_cell.font = styles['po_grand_font']
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    for c in range(1, 8):
        ws.cell(1, c).border = styles['thin_border']

    # Table Column Headers
    headers = ["PO", "TBM", "Product", "Crop", "Activity", "No. of Activities", "Total Amount"]
    ws.row_dimensions[2].height = 28
    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws.cell(2, col_idx, h_text)
        cell.font = styles['hdr_font']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = styles['thin_border']
        cell.fill = PatternFill(start_color='F2DCDB', end_color='F2DCDB', fill_type='solid')

    # Populate Data
    current_r = 3
    sorted_keys = sorted(summary_map.keys(), key=lambda k: (0 if k[0] != 'NO PO' else 1, k[0], k[1], k[2], k[4]))

    for key in sorted_keys:
        po_num, tbm_name, prod, crop, act = key
        data = summary_map[key]

        ws.row_dimensions[current_r].height = 20

        c_po = ws.cell(current_r, 1, po_num)
        c_po.font = styles['po_font']
        c_po.alignment = Alignment(horizontal='center', vertical='center')

        c_tbm = ws.cell(current_r, 2, tbm_name)
        c_tbm.font = styles['brown_text_font']
        c_tbm.alignment = Alignment(horizontal='left', vertical='center')

        c_prod = ws.cell(current_r, 3, prod)
        c_prod.font = styles['brown_text_font']
        c_prod.alignment = Alignment(horizontal='center', vertical='center')

        c_crop = ws.cell(current_r, 4, crop)
        c_crop.font = styles['brown_text_font']
        c_crop.alignment = Alignment(horizontal='center', vertical='center')

        c_act = ws.cell(current_r, 5, act)
        c_act.font = styles['brown_text_font']
        c_act.alignment = Alignment(horizontal='center', vertical='center')

        c_cnt = ws.cell(current_r, 6, data['count'])
        c_cnt.font = styles['regular_font']
        c_cnt.alignment = Alignment(horizontal='center', vertical='center')
        c_cnt.number_format = '#,##0'

        c_tot = ws.cell(current_r, 7, data['total_amount'])
        c_tot.font = styles['regular_font']
        c_tot.alignment = Alignment(horizontal='right', vertical='center')
        c_tot.number_format = '#,##0'

        for c in range(1, 8):
            ws.cell(current_r, c).border = styles['thin_border']

        current_r += 1

    # Grand Total Row
    grand_r = current_r
    ws.row_dimensions[grand_r].height = 24
    ws.merge_cells(start_row=grand_r, start_column=1, end_row=grand_r, end_column=5)
    
    lbl_g = ws.cell(grand_r, 1, "GRAND TOTAL")
    lbl_g.font = styles['po_grand_font']
    lbl_g.alignment = Alignment(horizontal='right', vertical='center')

    for c in range(1, 8):
        ws.cell(grand_r, c).border = styles['summary_border']
        ws.cell(grand_r, c).fill = styles['po_total_fill']

    c_cnt_tot = ws.cell(grand_r, 6, f"=SUM(F3:F{grand_r-1})")
    c_cnt_tot.font = styles['po_grand_font']
    c_cnt_tot.alignment = Alignment(horizontal='center', vertical='center')
    c_cnt_tot.number_format = '#,##0'

    c_amt_tot = ws.cell(grand_r, 7, f"=SUM(G3:G{grand_r-1})")
    c_amt_tot.font = styles['po_grand_font']
    c_amt_tot.alignment = Alignment(horizontal='right', vertical='center')
    c_amt_tot.number_format = '#,##0'

